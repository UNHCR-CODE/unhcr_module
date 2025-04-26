from fedot.core.pipelines.node import PipelineNode
from fedot.core.pipelines.pipeline import Pipeline
from fedot.utilities.ts_gapfilling import ModelGapFiller

import gc
import logging
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import ScatterChart, Reference, Series
import os
import pandas as pd
from sklearn.metrics import mean_absolute_error, mean_squared_error, median_absolute_error
import time



# Turn on interactive plotting mode
plt.ion()


from unhcr import constants as const
from unhcr import utils
from unhcr import db
# OPTIONAL: set your own environment
##ef = const.load_env(r'E:\_UNHCR\CODE\unhcr_module\.env')
## print(ef)
# OPTIONAL: set your own environment

mods = const.import_local_libs(mods=[["utils","utils"], ["constants", "const"], ["db", "db"]])
logger, *rest = mods
if const.LOCAL: # testing with local python files
    logger, utils, const, db = mods

utils.log_setup(level="INFO", log_file="unhcr.update_all.log", override=True)
logger.info(f"PROD: {const.PROD}, DEBUG: {const.DEBUG}, LOCAL: {const.LOCAL} {os.getenv('LOCAL')} .env file @: {const.environ_path}")

if not utils.is_version_greater_or_equal('0.4.8'):
    logger.error(
        "This version of the script requires at least version 0.4.6 of the unhcr module."
    )
    exit(47)

#!!! this is from E:\_UNHCR\CODE\DATA\gaps\eyedro_data_gaps.xlsx -- Azure gb tables with data gaps, that are not gensets
gbs_not_genset = ['gb_00980789', 'gb_0098082b', 'gb_00980858', 'gb_00980885', 'gb_0098088c', 'gb_0098088d', 'gb_00980890', 'gb_00980892', 'gb_00980898', 'gb_0098089a', 'gb_0098089c', 'gb_0098089e', 'gb_0098089f', 'gb_009808b0', 'gb_009808b1', 'gb_009808b6', 'gb_009808b9', 'gb_009808bb', 'gb_009808be', 'gb_009808bf', 'gb_009808f1', 'gb_0098090a', 'gb_0098090c', 'gb_00980912', 'gb_00980929', 'gb_00980958', 'gb_009809e9', 'gb_009809ea', 'gb_00980a21', 'gb_00980a2c', 'gb_00980a3e', 'gb_00980a4f', 'gb_00980a74', 'gb_00980aa1', 'gb_00980af4', 'gb_00980b2a', 'gb_00980b6e', 'gb_00980b81', 'gb_00980b89', 'gb_00980da0', 'gb_00980da2', 'gb_00980dfe', 'gb_00980e0d', 'gb_00980af5', 'gb_00980b11', 'gb_00980b35', 'gb_00980b6e', 'gb_00980da0', 'gb_00980da6', 'gb_00980db4', 'gb_00980dc4', 'gb_00980dc6', 'gb_00980dd7', 'gb_00980ddd', 'gb_00980df4', 'gb_00980dfe', 'gb_00980e0d', 'gb_00980e22']

# done = ['gb_00980789', 'gb_0098082b']
# gbs_not_genset = set(gbs_not_genset) - set(done)

engines = db.set_db_engines()

def print_gap_metrics(actual, ridge_predicted, composite_predicted, gap_number):
    """
    Print metrics for a specific gap
    
    :param actual: Actual values in the gap
    :param ridge_predicted: Ridge model predictions
    :param composite_predicted: Composite model predictions
    :param gap_number: Index of the current gap
    """
    print(f"\nGap {gap_number} Metrics:")
    
    model_labels = ['Inverted ridge regression', 'Composite model']
    for predicted, model_label in zip(
            [ridge_predicted, composite_predicted], model_labels):
        print(f"{model_label}")

        mae_metric = mean_absolute_error(actual, predicted)
        print(f"Mean absolute error - {mae_metric:.2f}")

        rmse_metric = (mean_squared_error(actual, predicted)) ** 0.5
        print(f"Root mean squared error - {rmse_metric:.2f}")

        median_ae_metric = median_absolute_error(actual, predicted)
        print(f"Median absolute error - {median_ae_metric:.2f}")


def plot_gap_result(subset, gap_indices, gap_number):
    """
    Plot results for a specific gap interactively
    
    :param subset: DataFrame containing the gap and surrounding data
    :param gap_indices: Indices of the gap within the subset
    :param gap_number: Index of the current gap
    """
    # Create masked array for visualizing the gap
    gap_mask = np.zeros(len(subset), dtype=bool)
    gap_mask[gap_indices] = True
    
    # Create a masked array for the gap
    masked_wh = np.ma.masked_array(subset['wh'].values, mask=~gap_mask)
    
    # Calculate the gap duration for the title
    gap_start = subset.iloc[gap_indices[0]]['date']
    gap_end = subset.iloc[gap_indices[-1]]['date']
    duration_minutes = len(gap_indices)
    
    # Create plot with lower DPI to reduce memory usage
    plt.figure(figsize=(10, 6), dpi=80)
    
    # Plot actual values
    plt.plot(subset['date'], subset['wh'], c='blue',
             alpha=0.5, label='Actual values', linewidth=1)
    
    # Plot the gap filled values
    # plt.plot(subset['date'], subset['ridge'], c='orange',
    #          alpha=0.8, label='Ridge gap-filling', linewidth=1)
    plt.plot(subset['date'], subset['composite'], c='red',
             alpha=0.8, label='Composite gap-filling', linewidth=1)
    
    # Plot the gap itself more prominently
    plt.plot(subset.iloc[gap_indices]['date'], masked_wh[gap_indices], 'ko', 
             markersize=3, label='Gap locations')
    
    # Add vertical lines to indicate gap boundaries
    plt.axvline(x=gap_start, color='gray', linestyle='--', alpha=0.7)
    plt.axvline(x=gap_end, color='gray', linestyle='--', alpha=0.7)
    
    plt.title(f'Gap {gap_number}: {duration_minutes} minutes ({gap_start} to {gap_end})')
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    
    # Show the plot interactively instead of saving
    plt.draw()
    plt.pause(0.1)  # Small pause to ensure the plot is drawn
    print(f"Displayed interactive plot for gap {gap_number}")
    # Add a small delay so the plot is visible
    time.sleep(0.5)  # Add a short delay to give time to view the plot


def get_composite_pipeline(window_size):
    """
    The function returns prepared pipeline of 5 models

    :return: Pipeline object
    """
    # Create node with a 150-minute window size
    node_1 = PipelineNode('lagged')
    node_1.parameters = {'window_size': window_size}
    
    # Create node with a 100-minute window size
    node_2 = PipelineNode('lagged')
    node_2.parameters = {'window_size': window_size}
    
    # Linear models from lagged features
    node_linear_1 = PipelineNode('linear', nodes_from=[node_1])
    node_linear_2 = PipelineNode('linear', nodes_from=[node_2])

    # Final ridge regression combining both linear models
    node_final = PipelineNode('ridge', nodes_from=[node_linear_1,
                                                  node_linear_2])
    return Pipeline(node_final)


def get_simple_pipeline(window_size):
    """ 
    Function returns simple pipeline with lagged features and ridge regression
    """
    # Create node with a 150-minute window size
    node_lagged = PipelineNode('lagged')
    node_lagged.parameters = {'window_size': window_size}
    
    # Ridge regression from lagged features
    node_ridge = PipelineNode('ridge', nodes_from=[node_lagged])
    return Pipeline(node_ridge)

def run_gapfilling_by_segments(file_path, plot_individual_gaps=True, min_gap_size=10, df=None):
    """
    Memory-efficient gap filling that processes one gap at a time
    and optionally plots each gap's results individually.

    :param file_path: path to the file
    :param plot_individual_gaps: whether to plot each gap individually
    :param min_gap_size: minimum gap size (in minutes) to plot individually
    :return: pandas dataframe with columns 'date','with_gap','ridge',
    'composite','wh'
    """
    # Load the data
    print("Loading data...")
    # Read the data with efficient dtypes
    if df is not None:
        dataframe = df
    else:
        dataframe = pd.read_csv(file_path, dtype={'wh': 'float32', 'with_gap': 'float32'})
    dataframe['date'] = pd.to_datetime(dataframe['date'], infer_datetime_format=True)
    
    print("Preprocessing time series...")
    # Convert date to index and fill missing timestamps
    dataframe = dataframe.set_index("date")
    full_range = pd.date_range(start=dataframe.index.min(), end=dataframe.index.max(), freq="T")
    dataframe = dataframe.reindex(full_range)
    
    # Process missing values
    missing_mask = dataframe["wh"].isna()
    dataframe["wh"] = dataframe["wh"].fillna(0)
    dataframe.loc[missing_mask, "with_gap"] = -100.0
    
    # Reset index
    dataframe = dataframe.reset_index().rename(columns={"index": "date"})
    
    print("Identifying gaps in the time series...")
    # Find all gaps (continuous sequences of -100.0 values)
    is_gap = dataframe['with_gap'] == -100.0
    gap_indices = np.where(is_gap)[0]
    
    if len(gap_indices) == 0:
        print("No gaps found in the data.")
        return dataframe  # No gaps to fill
    
    print(f"Found {len(gap_indices)} missing values across multiple gaps.")
    
    # Initialize result columns with actual values
    dataframe['ridge'] = dataframe['wh'].copy()
    dataframe['composite'] = dataframe['wh'].copy()
    
    # Group continuous gap indices to identify separate gaps
    print("Grouping continuous gaps...")
    gap_groups = []
    current_group = [gap_indices[0]]
    
    for i in range(1, len(gap_indices)):
        if gap_indices[i] == gap_indices[i-1] + 1:
            # This gap is continuous with the previous one
            current_group.append(gap_indices[i])
        else:
            # This is the start of a new gap
            gap_groups.append(current_group)
            current_group = [gap_indices[i]]
    
    # Add the last group if it exists
    if current_group:
        gap_groups.append(current_group)
    
    print(f"Identified {len(gap_groups)} separate gaps to process.")
    
    
    # Process each gap separately to minimize memory usage
    for i, gap_group in enumerate(gap_groups):
        gap_size = len(gap_group)
        if gap_size > 1440*7:
            print(f"Skipping gap {i+1} of {len(gap_groups)} (size: {gap_size} minutes): too long (> 60 day)")
            continue
        
        print(f"Processing gap {i+1} of {len(gap_groups)} (size: {gap_size} minutes)...")
        window_size = 300 #gap_size * 3 if gap_size * 3 < 5000 else 5000
        before_after = 1440 #gap_size * 1.5 if gap_size * 1.5 < 1440 else 1440
        # Determine the maximum window size used in any model
        # This ensures we have enough data before and after the gap for accurate prediction
        max_window_size = window_size  # Maximum window size used in any model
        buffer_size = max_window_size * 2  # Extra buffer for safety

        # Calculate the start and end indices with buffer for window size
        # Ensure we don't go out of bounds
        start_idx = max(0, gap_group[0] - buffer_size)
        end_idx = min(len(dataframe) - 1, gap_group[-1] + buffer_size)
        
        # Extract only the subset of data around the gap
        subset = dataframe.iloc[start_idx:end_idx+1].copy()
        
        # Apply gap filling algorithms to just this subset
        gap_array = subset['with_gap'].values
        
        print(f"  Applying ridge regression model to gap {i+1}...")
        # Ridge pipeline - simpler model
        ridge_pipeline = get_simple_pipeline(window_size)
        ridge_gapfiller = ModelGapFiller(gap_value=-100.0, pipeline=ridge_pipeline)
        subset['ridge'] = ridge_gapfiller.forward_inverse_filling(gap_array)
        #!!!! only use ridge for now
        subset['composite'] = ridge_gapfiller.forward_inverse_filling(gap_array)
        
        # Clean up to free memory
        del ridge_pipeline, ridge_gapfiller
        gc.collect()
        
        #!!!! only use ridge for now
        # print(f"  Applying composite model to gap {i+1}...")
        # # Composite pipeline - more complex model
        # composite_pipeline = get_composite_pipeline(window_size)
        # composite_gapfiller = ModelGapFiller(gap_value=-100.0, pipeline=composite_pipeline)
        # subset['composite'] = composite_gapfiller.forward_filling(gap_array)
        
        # # Clean up to free memory
        # del composite_pipeline, composite_gapfiller
        # gc.collect()
        
        # Identify the gap positions within the subset
        subset_gap_mask = subset['with_gap'] == -100.0
        subset_gap_indices = np.where(subset_gap_mask)[0]
        
        # Map subset positions back to original dataframe positions
        original_positions = subset_gap_indices + start_idx
        

        # After gap filling, replace negative values with small random positive numbers
        min_value = 0.1  # Lower bound of random range
        max_value = 2.0   # Upper bound of random range

        # Generate random replacements for ridge predictions
        negative_ridge_mask = subset['ridge'] < 0
        if negative_ridge_mask.any():
            random_values = np.random.uniform(min_value, max_value, size=negative_ridge_mask.sum())
            subset.loc[negative_ridge_mask, 'ridge'] = random_values

        # Generate random replacements for composite predictions
        negative_composite_mask = subset['composite'] < 0
        if negative_composite_mask.any():
            random_values = np.random.uniform(min_value, max_value, size=negative_composite_mask.sum())
            subset.loc[negative_composite_mask, 'composite'] = random_values
        
        
        
        
        
        
        
        
        # Copy filled values back to main dataframe
        # We only copy the actual gap values that were filled, not the entire subset
        dataframe.loc[original_positions, 'ridge'] = subset.iloc[subset_gap_indices]['ridge'].values
        dataframe.loc[original_positions, 'composite'] = subset.iloc[subset_gap_indices]['composite'].values
        
        # Calculate and print metrics for this specific gap
        if len(subset_gap_indices) > 0:
            actual = subset.iloc[subset_gap_indices]['wh'].values
            ridge_predicted = subset.iloc[subset_gap_indices]['ridge'].values
            composite_predicted = subset.iloc[subset_gap_indices]['composite'].values
            
            print_gap_metrics(actual, ridge_predicted, composite_predicted, i+1)
            
            # Plot this gap individually if it's large enough
            if plot_individual_gaps and gap_size >= min_gap_size:
                # We only plot context around the gap, not the entire subset
                context_size = min(buffer_size, before_after)  # Show at most 60 data points before/after
                plot_start = int(max(0, subset_gap_indices[0] - context_size))
                plot_end = int(min(len(subset) - 1, subset_gap_indices[-1] + context_size))
                
                plot_subset = subset.iloc[plot_start:plot_end+1].copy()
                
                # Adjust gap indices for the new reduced plotting subset
                plot_gap_indices = [idx - plot_start for idx in subset_gap_indices 
                                   if plot_start <= idx <= plot_end]
                
                # Plot this specific gap
                with plt.style.context('seaborn-v0_8-whitegrid'):  # Use a nice style
                    plot_gap_result(plot_subset, plot_gap_indices, i+1)
                    # Allow the user to press Enter to continue or just wait
                    # when in interactive mode
                    print("Showing plot... (close the plot window or press Enter to continue)")
                    try:
                        # This will return immediately if Enter is pressed
                        # or timeout after 5 seconds
                        plt.pause(2)
                    except Exception:
                        pass
            else:
                context_size = min(buffer_size, before_after)  # Show at most 60 data points before/after
                plot_start = int(max(0, subset_gap_indices[0] - context_size))
                plot_end = int(min(len(subset) - 1, subset_gap_indices[-1] + context_size))
                
                plot_subset = subset.iloc[plot_start:plot_end+1].copy()
                total_ridge = plot_subset.loc[plot_subset['with_gap'] == -100, 'ridge'].sum()
                print(total_ridge)
                file_name = f'{data_dir}\\{table}_subset_zero_gap.xlsx'
                if total_ridge > 0:
                    file_name = f'{data_dir}\\{table}_subset_gap.xlsx'
                
                plot_subset.to_excel(file_name, index=False, sheet_name="Data")
                # Load workbook and select sheet
                wb = load_workbook(file_name)
                ws = wb["Data"]

                # Define chart
                chart = ScatterChart()
                chart.title = "With Gap vs Ridge"
                chart.x_axis.title = "Date"
                chart.x_axis.number_format = "yyyy-mm-dd hh:mm"
                chart.x_axis.majorTimeUnit = "days"

                # Set X (date) and Y (ridge, with_gap) ranges
                x_values = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                y_values_ridge = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)  # Ridge
                y_values_gap = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)  # With Gap

                # Create Series
                series_ridge = Series(y_values_ridge, x_values, title="Ridge")
                series_gap = Series(y_values_gap, x_values, title="With Gap")

                # Set secondary axis for with_gap
                series_gap.graphicalProperties.line.solidFill = "FF0000"  # Red for differentiation
                series_gap.marker.symbol = "circle"
                series_gap.y_axis = 1  # Assign to secondary axis

                # Add series to chart
                chart.series.append(series_ridge)
                chart.series.append(series_gap)

                # Enable secondary axis
                chart.y_axis.axId = 100
                chart.y_axis2 = chart.y_axis.__class__()
                chart.y_axis2.axId = 200
                chart.y_axis2.title = "With Gap Values"
                chart.y_axis2.crossAx = 100  # Cross primary Y-axis
                chart.y_axis.crossAx = 200  # Cross secondary Y-axis
                chart.series[1].y_axis = chart.y_axis2  # Assign second series to secondary axis

                # Add chart to sheet
                ws.add_chart(chart, "E5")

                # Save the workbook
                wb.save(file_name)
                print(f"Excel file saved: {file_name}")
        
        # Clean up subset to free memory
        del subset, subset_gap_indices, original_positions
        gc.collect()
        
        print(f"  Completed gap {i+1}.")
    
    print("All gaps have been filled successfully.")
    return dataframe


def hyper_gaps(table, data_dir):


    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.chart import ScatterChart,LineChart, BarChart,Reference, Series
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.styles import NamedStyle
    from openpyxl.chart.text import RichText 
    from openpyxl.drawing.text import RichTextProperties


    csv_file = r'E:\_UNHCR\CODE\DATA\gaps\gb_00980789_subset_gap.csv'
    df = pd.read_csv(csv_file)

    # Convert date column to datetime type
    df["date"] = pd.to_datetime(df["date"])

    ####df[ "with_gap"] = df["with_gap"] / 100
    # Write data to Excel first, using 'openpyxl' engine
    file_name = r"E:\_UNHCR\CODE\DATA\gaps\gb_00980789_subset_gap.xlsx"
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name='Data', index=False)

    # Load workbook and select sheet (after saving data)
    wb = load_workbook(file_name)
    ws = wb["Data"]

    ws.delete_cols(5)
    ws.delete_cols(2)
    
    
    # from openpyxl.chart.axis import DateAxis, NumericAxis
    # from openpyxl.chart.text import RichText
    # from openpyxl.drawing.text import Paragraph,ParagraphProperties, CharacterProperties, RichTextProperties

    
    # # **Ensure Date format in Column A**
    # # Set the style for date formatting and make it a str
    # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    #     for cell in row:
    #         cell.value = cell.value.strftime("%Y-%m-%d %H:%M")

    # # Create primary Line Chart for the ridge method
    # chart1 = LineChart()
    # chart1.type = "line"
    # chart1.style = 10
    # chart1.title = "Gap filling comparison"

    # # Primary axis titles
    # chart1.y_axis.title = "Ridge (Wh)"
    # chart1.x_axis.title = "Date time"

    # # Primary axis tick marks
    # chart1.y_axis.majorTickMark = "out"
    # chart1.x_axis.majorTickMark = "out"

    # # X-axis labels position
    # chart1.x_axis.tickLblPos = "low"

    # # Reset axis title formatting
    # chart1.y_axis.title.txPr = None  

    # # Add data for the ridge method
    # ridge_data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=ws.max_row)
    # chart1.add_data(ridge_data, titles_from_data=True)

    # # Set X-axis categories
    # categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    # chart1.set_categories(categories)
    # chart1.x_axis.majorGridlines = None

    # # Create secondary chart for composite method
    # chart2 = LineChart()
    # chart2.type = "line"
    # chart2.style = 11  # Different style for differentiation

    # # Important: Set up the secondary axis
    # chart2.y_axis = NumericAxis(axId=200)  # Create a secondary y-axis with a unique ID
    # chart2.y_axis.title = "Composite (Wh)"
    # chart2.y_axis.crosses = "max"  # Position the axis at the maximum x value (right side)
    # chart2.y_axis.majorGridlines = None

    # # Add data for the composite method
    # composite_data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)
    # chart2.add_data(composite_data, titles_from_data=True)

    # # Calculate Y-axis scaling for primary axis (ridge)
    # ridge_values = [value for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=ws.max_row, values_only=True) 
    #             for value in row if value is not None]
    # max_ridge = max(ridge_values) if ridge_values else 1
    # min_ridge = min(ridge_values) if ridge_values else 0
    # chart1.y_axis.scaling.min = min_ridge
    # chart1.y_axis.scaling.max = max_ridge * 1.1
    # chart1.y_axis.majorUnit = round((max_ridge - min_ridge) / 10, 2)

    # # Calculate Y-axis scaling for secondary axis (composite)
    # composite_values = [value for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, max_row=ws.max_row, values_only=True) 
    #                 for value in row if value is not None]
    # max_composite = max(composite_values) if composite_values else 1
    # min_composite = min(composite_values) if composite_values else 0
    # chart2.y_axis.scaling.min = min_composite
    # chart2.y_axis.scaling.max = max_composite * 1.1
    # chart2.y_axis.majorUnit = round((max_composite - min_composite) / 10, 2)

    # # Primary axis settings
    # chart1.y_axis.majorTickMark = "out"
    # chart1.y_axis.minorTickMark = "none"
    # chart1.y_axis.tickLblPos = "nextTo"
    # chart1.y_axis.crosses = "autoZero"

    # # X-axis settings
    # chart1.x_axis.majorTickMark = "out"
    # chart1.x_axis.tickLblPos = "low"
    # chart1.x_axis.majorUnit = 1
    # chart1.x_axis.minorUnit = 1

    # # Rotate x-axis labels
    # from openpyxl.chart.text import RichText
    # from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties

    # chart1.x_axis.txPr = RichText()
    # paragraph = Paragraph()
    # paragraph.pPr = ParagraphProperties()
    # chart1.x_axis.txPr.p.append(paragraph)
    # chart1.x_axis.txPr.p[0].pPr.defRPr = CharacterProperties(rot="-2700000")
    # chart1.x_axis.title = None

    # # Link the charts - this is the key step for secondary axis
    # chart1 += chart2

    # # Chart size
    # chart1.width = 25
    # chart1.height = 15

    # # Layout settings
    # chart1.layout = Layout(
    #     manualLayout=ManualLayout(
    #         x=0.999,
    #         h=0.95,
    #         w=0.95
    #     )
    # )

    # # Add chart to worksheet
    # ws.add_chart(chart1, "G2")
    # wb.save(file_name)
    
    
    
    

    # **Ensure Date format in Column A**
    # Set the style for date formatting and make it a str
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = cell.value.strftime("%Y-%m-%d %H:%M")

    # Create Line Chart
    chart1 = LineChart()
    chart1.type = "line"
    chart1.style = 10
    chart1.title = "Gap filling"

    # Axis titles
    chart1.y_axis.title = "-100 = Gap Filled (with_gap)"
    chart1.x_axis.title = "Date time"

    # **Ensure axis labels (tick marks) are visible**
    chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart1.x_axis.majorTickMark = "out"  # Show tick marks outward

    # **Ensure X-axis labels are positioned at the bottom**
    chart1.x_axis.tickLblPos = "low"  # Position X-axis labels at the bottom

    # **Ensure axis titles are positioned correctly**
    #####chart1.x_axis.title.txPr = None  # Reset title formatting (places it outside)
    chart1.y_axis.title.txPr = None  

    # **Set Y-axis data range** (assuming the remaining column is the Y-values after removing columns B and E)
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=ws.max_row)  # Y-values in column B
    chart1.add_data(data, titles_from_data=True)

  # # **Set X-axis to the date values in column 1 (starting from row 2 to the last row)**
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Dates in column A
    chart1.x_axis.majorGridlines = None  # Remove gridlines for clarity
    chart1.set_categories(categories)  # Use set_categories if it works

    # **Increase chart size to fit labels properly**
    chart1.width = 25  # Increase width (default is ~15)
    chart1.height = 15  # Increase height (default is ~7.5)

    chart1.layout = Layout(
        manualLayout=ManualLayout(
            #x=1,  # Shift plot area slightly right
            h=0.9,   # Reduce height of the plot area
            w=0.95    # Reduce width of the plot area
        )
    )

    # Y-axis settings (auto-scaling, tick marks, etc.)
    y_values = [value for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=ws.max_row, values_only=True) 
                for value in row if value is not None]
    max_y = max(y_values) if y_values else 1
    min_y = min(y_values) if y_values else 0
    chart1.y_axis.scaling.min = min_y - 1
    chart1.y_axis.scaling.max = max_y + 1
    chart1.y_axis.majorUnit = round((max_y - min_y) / 10, 2)

    # Ensure Y-axis labels are displayed properly
    chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart1.y_axis.minorTickMark = "none"  # Hide minor tick marks
    chart1.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks
    chart1.y_axis.delete = False  # Ensure axis is visible

    # Ensure X-axis tick marks and labels are displayed
    chart1.x_axis.majorTickMark = "out"  # Show tick marks outward
    chart1.x_axis.tickLblPos = "low"  # Ensure labels are at the bottom
    chart1.x_axis.delete = False  # Make sure X-axis is visible
    chart1.x_axis.majorUnit = 1  # Display one tick per hour
    chart1.x_axis.minorUnit = 1  # Prevent minor ticks from interfering
    #chart1.x_axis.label_rotation = 45
    #ws['A2'] = str(ws['A2'].value) + ' '

    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph,ParagraphProperties, CharacterProperties, RichTextProperties

    chart1.x_axis.txPr = chart1.x_axis.title.text.rich
    chart1.x_axis.txPr.properties.rot = "-2700000"
    chart1.x_axis.title = None
        
        
        
# Create Line Chart
    chart2 = LineChart()
    chart2.type = "line"
    chart2.style = 10
    chart2.title = "Gap filling"

    # Axis titles
    chart2.y_axis.title = "Wh (ridge)"
    chart2.x_axis.title = "Date time"

    # **Ensure axis labels (tick marks) are visible**
    chart2.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart2.x_axis.majorTickMark = "out"  # Show tick marks outward

    # **Ensure X-axis labels are positioned at the bottom**
    chart2.x_axis.tickLblPos = "low"  # Position X-axis labels at the bottom

    # **Ensure axis titles are positioned correctly**
    #####chart2.x_axis.title.txPr = None  # Reset title formatting (places it outside)
    chart2.y_axis.title.txPr = None  

    # **Set Y-axis data range** (assuming the remaining column is the Y-values after removing columns B and E)
    data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)  # Y-values in column B
    chart2.add_data(data, titles_from_data=True)

  # # **Set X-axis to the date values in column 1 (starting from row 2 to the last row)**
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Dates in column A
    chart2.x_axis.majorGridlines = None  # Remove gridlines for clarity
    chart2.y_axis.majorGridlines = None 
    chart2.set_categories(categories)  # Use set_categories if it works

    # **Increase chart size to fit labels properly**
    chart2.width = 25  # Increase width (default is ~15)
    chart2.height = 15  # Increase height (default is ~7.5)

    chart2.layout = Layout(
        manualLayout=ManualLayout(
            #x=1,  # Shift plot area slightly right
            h=0.9,   # Reduce height of the plot area
            w=0.95    # Reduce width of the plot area
        )
    )

    # Y-axis settings (auto-scaling, tick marks, etc.)
    y_values = [value for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, max_row=ws.max_row, values_only=True) 
                for value in row if value is not None]
    max_y = max(y_values) if y_values else 1
    min_y = min(y_values) if y_values else 0
    chart2.y_axis.scaling.min = min_y - 1
    chart2.y_axis.scaling.max = max_y + 1
    chart2.y_axis.majorUnit = round((max_y - min_y) / 10, 2)

    # Ensure Y-axis labels are displayed properly
    chart2.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart2.y_axis.minorTickMark = "none"  # Hide minor tick marks
    chart2.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks
    chart2.y_axis.delete = False  # Ensure axis is visible

    # Ensure X-axis tick marks and labels are displayed
    chart2.x_axis.majorTickMark = "out"  # Show tick marks outward
    chart2.x_axis.tickLblPos = "low"  # Ensure labels are at the bottom
    chart2.x_axis.delete = False  # Make sure X-axis is visible
    chart2.x_axis.majorUnit = 1  # Display one tick per hour
    chart2.x_axis.minorUnit = 1  # Prevent minor ticks from interfering
    #chart2.x_axis.label_rotation = 45
    #ws['A2'] = str(ws['A2'].value) + ' '

    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph,ParagraphProperties, CharacterProperties, RichTextProperties

    chart2.x_axis.txPr = chart2.x_axis.title.text.rich
    chart2.x_axis.txPr.properties.rot = "-2700000"
    chart2.x_axis.title = 'XXXXX'
    chart2.y_axis.axId = 200
    chart1.legend.position = 'b'

        
        
    chart1.y_axis.crosses = "max"
    chart1 += chart2
        
        
        
    
    # chart_sheet_name = "Charts"
    # ws_ch
    # art = wb.create_sheet(chart_sheet_name)
    # # **Add chart to Excel**
    # ws_chart.add_chart(chart1, "C2")
    ws.add_chart(chart1, "G2")
    wb.save(file_name)

    
    
    
    
    
    
    
    
    
    
    pass



    # Load workbook and select sheet (after saving data)
    wb = load_workbook(file_name)
    ws = wb["Data"]

    ws.delete_cols(5, 1)
    ws.delete_cols(2, 2)
    
    # Define LineChart
    chart = ScatterChart()
    chart.title = "With Gap vs Ridge"
    chart.style = 13  # Optional style for better visual appearance
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Values"
    chart.x_axis.number_format = "yyyy-mm-dd hh:mm"
    chart.x_axis.majorTimeUnit = "days"

    # Set X (date) and Y (with_gap, ridge) ranges
    x_values = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    y_values1 = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)  # Ridge
    y_values2 = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)  # With Gap

    # Create Series for Ridge and With Gap data
    series1 = Series(y_values1, x_values, title="Ridge")
    series2 = Series(y_values2, x_values, title="With Gap")
    series2.graphicalProperties.line.solidFill = "FF0000"  # Change color for visibility

    # Add the series to the chart
    chart.series.append(series1)
    chart.series.append(series2)

    # Add chart to the sheet at the desired position
    ws.add_chart(chart, "E5")

    # Save the workbook after adding the chart
    wb.save(file_name)
    print(f"Excel file saved: {file_name}")




    if os.path.isfile(f'{data_dir}\\{table}_gaps.csv'):
        df = pd.read_csv(f'{data_dir}\\{table}_gaps.csv')
        df[["wh", "with_gap"]] = df[["wh", "with_gap"]].astype("float32")
        df["date"] = pd.to_datetime(df["date"])
        return df
    conn = engines[1].raw_connection()  # Get raw psycopg2 connection
    with conn.cursor() as cursor:
        # Step 1: Get all hypertables in the "eyedro" schema
        cursor.execute(f"""select ts date, (abs(wh_p1) + abs(wh_p2) + abs(wh_p3)) wh, (abs(wh_p1) + abs(wh_p2) + abs(wh_p3)) with_gap from eyedro.{table} order by ts;""")

        rows = cursor.fetchall()
        print(f"Processing hypertable {table} Rows: {len(rows)}")

        df = pd.DataFrame(rows, columns=["date", "wh", "with_gap"])
        df[["wh", "with_gap"]] = df[["wh", "with_gap"]].astype("float32")
        df["date"] = pd.to_datetime(df["date"])

        # Save results to CSV
        if not df.empty:
            path = f'{data_dir}\\{table}'
            df.to_csv(path + r"_gaps.csv", mode="w", index=False, header=True)
        # Close connection
        cursor.close()
        conn.close()
    return df


# Main function to run the example
if __name__ == '__main__':
    #try:
    # Start with a clean memory state
    gc.collect()
    data_dir = r'E:\_UNHCR\CODE\DATA\gaps'
    print("Starting memory-efficient gap filling process...")
    for table in gbs_not_genset:
        df = hyper_gaps(table,data_dir)
    
        file_path = f'{table}_gaps.csv'
        
        # Run the gap filling with our memory-efficient approach
        full_path = os.path.join(data_dir, file_path)
        dataframe = run_gapfilling_by_segments(full_path, plot_individual_gaps=False, min_gap_size=10, df=df)
        
        print("\nProcess completed successfully.")
        print("Interactive plots were displayed during processing.")
        
        # Keep the plots open until the user closes the program
        print("Press Ctrl+C to exit...")
        plt.ioff()  # Turn off interactive mode
        plt.show(block=True)  # This will block until all plot windows are closed
    
    # except Exception as e:
    #     print(f"An error occurred: {str(e)}")
    # finally:
    #     # Final cleanup to ensure memory is released
    #     print("Performing final memory cleanup...")
    #     gc.collect()
    #     print("Done.")