import csv
from datetime import datetime
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import logging
import openpyxl
from openpyxl.styles import Font
import pandas as pd
import requests
from unhcr import app_utils
from unhcr import err_handler
import unhcr.constants as const
from unhcr import db

mods = [
    ["app_utils", "app_utils"],
    ["err_handler", "err_handler"],
    ["constants", "const"],
    ["db", "db"],
]

res = app_utils.app_init(
    mods, "unhcr.gb_serial_nums.log", "0.4.7", level="INFO", override=True
)
if const.LOCAL:
    app_utils, err_handler, const, db = res

engines = db.set_db_engines()

if const.is_running_on_azure():
    istvan_csv = "~/code/DATA/gaps/new_top_20_2025-03-11_trimmed.csv"
    gaps_csv = "~/code/DATA/gaps/gaps.csv"
    gtb_gaps_excel = "~/code/DATA/gaps/gtb_gaps.xlsx"
    all_api_gbs_csv = "~/code/DATA/all_api_gbs.csv"
    unifier_gb_csv = "~/code/DATA/unifier_gb.csv"
    gb_merged_excel = "~/code/DATA/gaps/merged.xlsx"
else:
    istvan_csv = r"E:\_UNHCR\CODE\DATA\gaps\new_top_20_2025-03-11_trimmed.csv"
    gaps_csv = r"E:\_UNHCR\CODE\DATA\gaps\gaps.csv"
    gtb_gaps_excel = r"E:\_UNHCR\CODE\DATA\gaps\gtb_gaps.xlsx"
    all_api_gbs_csv = r"E:\_UNHCR\CODE\DATA\all_api_gbs.csv"
    unifier_gb_csv = r"E:\_UNHCR\CODE\DATA\unifier_gb.csv"
    gb_merged_excel = r"E:\_UNHCR\CODE\DATA\gaps\gb_merged_.xlsx"

#!!!!!!!!!!!!!

import pandas as pd
import sqlalchemy
import os
from sqlalchemy import create_engine, inspect
import psycopg2
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT
import logging
import traceback
import sys


def excel_to_postgres(
    excel_file_path,
    engine,
    schema="public",
    if_exists="replace",
    chunk_size=1000,
    table_prefix="gb_",
    table_suffix="",
):
    """
    Load all sheets from an Excel file into PostgreSQL database tables.

    Parameters:
    -----------
    excel_file_path : str
        Path to the Excel file
    schema : str, optional
        Schema name to use, if None uses the default schema
    if_exists : str, optional
        How to behave if table exists. Options: 'fail', 'replace', 'append'
    chunk_size : int, optional
        Number of rows to insert at once
    table_prefix : str, optional
        Prefix to add to all table names
    table_suffix : str, optional
        Suffix to add to all table names

    Returns:
    --------
    list
        List of table names created

    Raises:
    -------
    ValueError
        If input parameters are invalid
    FileNotFoundError
        If Excel file not found
    ConnectionError
        If cannot connect to the database
    PermissionError
        If database or schema creation fails due to permissions
    RuntimeError
        For processing errors
    IOError
        For file reading errors
    """
    created_tables = []

    # Input validation
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    if if_exists not in ["fail", "replace", "append"]:
        raise ValueError(
            f"Invalid if_exists value: {if_exists}. Must be 'fail', 'replace', or 'append'"
        )

    if not isinstance(chunk_size, int) or chunk_size <= 0:
        raise ValueError(
            f"Invalid chunk_size: {chunk_size}. Must be a positive integer"
        )

    # Check if database exists, if not create it
    try:
        with engine.connect() as conn:
            logging.info(f"Connected to database: {engine.url}")
        # Read all sheets from the Excel file
        try:
            excel_file = pd.ExcelFile(excel_file_path)
            sheet_names = excel_file.sheet_names
            logging.info(f"Found {len(sheet_names)} sheets in {excel_file_path}")
        except FileNotFoundError:
            raise
        except pd.errors.EmptyDataError:
            raise ValueError(f"Excel file is empty: {excel_file_path}")
        except Exception as e:
            raise IOError(f"Failed to read Excel file {excel_file_path}: {str(e)}")

        for sheet_name in sheet_names:
            try:
                # Read the sheet into a DataFrame
                logging.info(f"Reading sheet: {sheet_name}")
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                except pd.errors.EmptyDataError:
                    logging.warning(f"Sheet '{sheet_name}' is empty, skipping")
                    continue
                except Exception as e:
                    raise IOError(f"Failed to read sheet '{sheet_name}': {str(e)}")

                # Skip if empty
                if df.empty:
                    logging.warning(f"Sheet '{sheet_name}' has no data, skipping")
                    continue

                # Clean up column names for SQL compatibility
                original_columns = df.columns.tolist()
                df.columns = [
                    col.strip().lower().replace(" ", "_").replace("-", "_")
                    for col in df.columns
                ]

                # Log column name changes
                column_changes = {
                    original: new
                    for original, new in zip(original_columns, df.columns)
                    if original != new
                }
                if column_changes:
                    logging.info(
                        f"Column names changed for SQL compatibility: {column_changes}"
                    )

                # Create table name from sheet name
                table_name = f"{table_prefix}{sheet_name.strip().lower().replace(' ', '_').replace('-', '_')}{table_suffix}"

                # Create the full table name with schema if provided
                full_table_name = f"{schema}.{table_name}" if schema else table_name

                try:
                    # Write the DataFrame to PostgreSQL in chunks
                    logging.info(f"Writing {len(df)} rows to table {full_table_name}")
                    df.to_sql(
                        name=table_name,
                        schema=schema,
                        con=engine,
                        if_exists=if_exists,
                        index=False,
                        chunksize=chunk_size,
                    )

                    created_tables.append(full_table_name)
                    logging.info(
                        f"Successfully loaded sheet '{sheet_name}' to table '{full_table_name}'"
                    )
                except sqlalchemy.exc.OperationalError as e:
                    if "permission denied" in str(e).lower():
                        raise PermissionError(
                            f"Permission denied when creating table {full_table_name}: {str(e)}"
                        )
                    else:
                        raise RuntimeError(
                            f"Failed to create table {full_table_name}: {str(e)}"
                        )
                except Exception as e:
                    raise RuntimeError(
                        f"Failed to create/load table {full_table_name}: {str(e)}"
                    )
            except (
                FileNotFoundError,
                IOError,
                ValueError,
                ConnectionError,
                PermissionError,
            ):
                raise
            except Exception as e:
                raise RuntimeError(f"Failed to process sheet {sheet_name}: {str(e)}")

        if not created_tables:
            logging.warning("No tables were created from the Excel file")

        return created_tables

    except Exception as e:
        logging.error(f"Error: {str(e)}")
        logging.error(traceback.format_exc())
        raise


def mainline(engine, file_path):
    """Example usage of the excel_to_postgres function"""
    try:
        tables = excel_to_postgres(
            excel_file_path=file_path,
            engine=engine,
            schema="public",  # Use 'public' or your preferred schema
            if_exists="replace",  # 'fail', 'replace', or 'append'
            table_prefix="gb_",  # Optional prefix for tables
            chunk_size=5000,  # Adjust based on data size and memory constraints
        )

        logging.info(f"Successfully created {len(tables)} tables:")
        for table in tables:
            logging.info(f"- {table}")

    except FileNotFoundError as e:
        logging.error(f"File not found: {str(e)}")
        sys.exit(1)
    except ValueError as e:
        logging.error(f"Invalid input: {str(e)}")
        sys.exit(2)
    except ConnectionError as e:
        logging.error(f"Connection error: {str(e)}")
        sys.exit(3)
    except PermissionError as e:
        logging.error(f"Permission error: {str(e)}")
        sys.exit(4)
    except IOError as e:
        logging.error(f"I/O error: {str(e)}")
        sys.exit(5)
    except RuntimeError as e:
        logging.error(f"Runtime error: {str(e)}")
        sys.exit(6)
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        logging.error(traceback.format_exc())
        sys.exit(9)

merged_excel_path = gb_merged_excel.replace(
    "_.xlsx", f"_{datetime.now().date().isoformat()}.xlsx"
)
mainline(engines[1], merged_excel_path)
pass
#!!!!!!!!!!!!!!!!!!


def freeze_row1(file_path, insert_cols=[]):
    # Open the workbook with openpyxl (not pandas)
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.freeze_panes = "A2"  # Freeze the first row
        if "merged" in sheet_name:
            for col in insert_cols:
                c = col[0]
                v = col[1]
                sheet.insert_cols(c)
                for row in range(1, sheet.max_row + 1):
                    sheet.cell(row=row, column=c).value = v
                    cell = sheet.cell(row=row, column=c)
                    cell.font = Font(bold=True)
        elif "api_unifier" in sheet_name:
            sheet.insert_cols(0)
            sheet.insert_cols(20)
            for row in range(1, sheet.max_row + 1):
                sheet.cell(row=row, column=1).value = "unifier"
                sheet.cell(row=row, column=20).value = "api_matched"
                cell = sheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell = sheet.cell(row=row, column=20)
                cell.font = Font(bold=True)

    # Save the modified workbook
    wb.save(file_path)


def get_duplicate_elements(lst):
    seen_count = {}
    duplicates = {}

    for sublist in lst:
        # Trim each string in the sublist and convert to tuple
        trimmed_sublist = tuple(s.strip() if isinstance(s, str) else s for s in sublist)

        if trimmed_sublist in seen_count:
            seen_count[trimmed_sublist] += 1
            # Add to duplicates with current count if this is the first duplicate instance
            if trimmed_sublist not in duplicates:
                duplicates[trimmed_sublist] = seen_count[trimmed_sublist]
            else:
                duplicates[trimmed_sublist] = seen_count[trimmed_sublist]
        else:
            seen_count[trimmed_sublist] = 1

    # Convert dictionary keys back to lists for result
    result = []
    for trimmed_tuple, count in duplicates.items():
        result.append([list(trimmed_tuple), count])

    return result


def get_gb_user_info_data(csv_file):
    url = (
        const.GB_API_V1_API_BASE_URL
        + const.GB_API_V1_GET_DEVICE_LIST
        + const.GB_API_V1_USER_KEY
    )
    payload = {}
    headers = {}

    response, err = err_handler.request(
        lambda: requests.request("GET", url, headers=headers, data=payload)
    )
    if err:
        return None, err
    user_info_data, err = err_handler.parse_json(lambda: response.json())
    if err:
        return None, err
    return user_info_data, None


def parse_user_info_as_list(user_info_data):
    top_level = user_info_data.get("UserInfo", {}).get("DeviceSerialList", [])

    top_level_serials = []
    for sn in top_level:
        top_level_serials.append(str(sn)[:3] + "-" + str(sn)[3:])

    # Extract DeviceSerials and SiteLabels from DisplayGroupList
    device_data = []
    for site in user_info_data.get("UserInfo", {}).get("SiteList", []):
        site_label = site.get("SiteLabel")
        for group in site.get("DisplayGroupList", []):
            for device in group.get("DeviceList", []):
                sn = device.get("DeviceSerial")
                sn = str(sn)[:3] + "-" + str(sn)[3:]
                last_com = device.get("LastCommSecUtc")
                state = device.get("State")
                device_data.append((site_label, sn, last_com, state))

    # Combine top-level serials with device data
    return [(None, serial) for serial in top_level_serials] + device_data


def get_user_info_as_list():
    # Extract top-level DeviceSerialList
    user_info_data, err = get_gb_user_info_data(all_api_gbs_csv)
    if err or len(user_info_data["Errors"]) != 0:
        err_str = f"app_gb_serial_nums: Failed to get user info data ERROR: {err}  {user_info_data['Errors']}"
        logging.error(err_str)
        return None, err_str
    all_serials, err = err_handler.parse_json(
        lambda: parse_user_info_as_list(user_info_data)
    )
    if err:
        err_str = f"app_gb_serial_nums: Failed to parse user info data ERROR: {err}"
        logging.error(err_str)
        return None, err_str
    return all_serials, None


merged_excel_path = gb_merged_excel.replace(
    "_.xlsx", f"_{datetime.now().date().isoformat()}.xlsx"
)

all_serials, err = get_user_info_as_list()
if err:
    logging.error(err)
    exit(1)

# Remove duplicates if any
all_serials = list(set(all_serials))

# # Write to CSV file
with open(all_api_gbs_csv, "w", newline="") as file:
    writer = csv.writer(file)
    writer.writerow(["site_label", "gb_serial", "last_com", "state"])
    writer.writerows(all_serials)

# Print the result
x = 1
for item in all_serials:
    print(
        f" {x}  device_serial: {item[1]}  site_label: {item[0] if item[0] else 'N/A'}"
    )
    x += 1
# #!!!!! gen all_api_gbs csv with eyedro label & serial

df_serial = pd.read_csv(all_api_gbs_csv)
df_serial["site_label"] = df_serial["site_label"].fillna("N/A")

df_gaps = pd.read_csv(gaps_csv)
df_gaps["gb_serial"] = df_gaps["gb_serial"].str[:3] + "-" + df_gaps["gb_serial"].str[3:]

with pd.ExcelWriter(merged_excel_path, engine="openpyxl") as writer:
    df_serial.to_excel(writer, sheet_name="api_no_dups", index=False)

with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_gaps.to_excel(writer, sheet_name="gb_gaps", index=False)


merged_df = pd.merge(df_serial, df_gaps, on="gb_serial", how="right")
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    merged_df.to_excel(writer, sheet_name="merged_api_gaps", index=False)

df_top20 = pd.read_csv(istvan_csv, encoding="ISO-8859-1")
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_top20.to_excel(writer, sheet_name="top20", index=False)

# # Merge with all rows from df2 and matching rows from df1
# merged_df = pd.merge(df_serials, df_top20, left_on='device_serial', right_on='Meter Serial No.', how='left')
# merged_df['site_label'] = merged_df['site_label'].fillna('')
# #merged_df['device_serial'] = merged_df['device_serial'].fillna('')
# merged_df = merged_df.fillna('')

matching_rows = []
# Loop through each row in merged_df
for idx, row in merged_df.iterrows():
    device_serial = row["gb_serial"]
    # Find all matching rows in df_top20
    matches = df_top20[df_top20["Meter Serial No."] == device_serial]

    if not matches.empty:
        if len(matches) == 1:
            # Pick the first match (you can modify this to pick based on any condition)
            selected_match = matches.iloc[0]
        else:
            matches["Date Installed"] = matches["Date Installed"].fillna(
                pd.Timestamp("1900-01-01")
            )

            # Sort matches by 'Date Installed' to get the most recent first
            sorted_matches = matches.sort_values(by="Date Installed", ascending=False)

            # Pick the most recent match (first row after sorting)
            selected_match = sorted_matches.iloc[0]
    else:
        # If no match is found, append the original row from df_serials
        matching_rows.append(row)
        continue
    # Concatenate the original row from df_serials with the selected match
    merged_row = pd.concat([row, selected_match], axis=0)
    matching_rows.append(merged_row)

# Create a new DataFrame from the list of matching rows
matching_df = pd.DataFrame(matching_rows)
matching_df.to_excel(gtb_gaps_excel, index=False)

with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    matching_df.to_excel(writer, sheet_name="gtb_merged_top20", index=False)

pass


#!!!!!!!!!!!!!!!!!!!!!!!!!!!


# Load the Excel file
###file_path = r"D:\Downloads\device_serials.xlsx"  # Update with your path
##xls = pd.ExcelFile(gtb_gaps_excel)

df1 = pd.read_csv(unifier_gb_csv)
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df1.to_excel(writer, sheet_name="gb_unifier", index=False)

df2 = pd.read_csv(all_api_gbs_csv)

# Drop NaN values
device_serials = df1["Serial Number"].dropna()
site_sns = df2["gb_serial"].dropna()

# Perform fuzzy matching
matches = []
for serial in device_serials:
    best_match = process.extractOne(serial, site_sns, scorer=fuzz.ratio)
    matches.append((serial, best_match[0], best_match[1]))

# Convert matches to DataFrame
matches_df = pd.DataFrame(matches, columns=["DeviceSerial", "MatchedSerial", "Score"])

# Merge df1 and matches_df
merged_df = df1.merge(matches_df, left_on="Serial Number", right_on="DeviceSerial")

# Merge with df2 to get more details from serials sheet
df_final_output = merged_df.merge(
    df2, left_on="MatchedSerial", right_on="gb_serial", how="left"
)

# Filter matches with a high score
df_final_output = df_final_output[df_final_output["Score"] > 99]
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_final_output.to_excel(writer, sheet_name="api_unifier_matched", index=False)

# Remove duplicate rows based on all columns
df_no_dups_output = df_final_output.drop_duplicates()
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_no_dups_output.to_excel(writer, sheet_name="api_unifier_no_dups", index=False)
freeze_row1(merged_excel_path, [[11, "top20"], [5, "gaps"], [1, "api_no_dups"]])


# Save the results to an Excel file with two sheets
output_path = "fuzzy_matches.xlsx"
with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
    df_final_output.to_excel(writer, sheet_name="fuzzy", index=False)
    df_no_dups_output.to_excel(writer, sheet_name="no-dups", index=False)

print(f"Matches saved to {output_path}")

pass
