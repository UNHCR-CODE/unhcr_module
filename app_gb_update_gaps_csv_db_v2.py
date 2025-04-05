from datetime import datetime, timedelta
import logging
import openpyxl
from openpyxl.styles import Font
import os
import pandas as pd
import plyer
from rapidfuzz import fuzz
from rapidfuzz import process
import requests
import sqlalchemy
import sys
import traceback

from app_time_series_gapfilling_gb_v3 import concurrent_update_gaps
from unhcr import app_utils
from unhcr import err_handler
from unhcr import utils
import unhcr.constants as const
from unhcr import db

run_dt = datetime.now().date() #####- timedelta(days=1)

mods = [
    ["app_utils", "app_utils"],
    ["err_handler", "err_handler"],
    ["utils", "utils"],
    ["constants", "const"],
    ["db", "db"],
    ['gb_eyedro', 'gb_eyedro'],
]

res = app_utils.app_init(
    mods, "unhcr.gb_serial_nums.log", "0.4.7", level="INFO", override=True
)
if const.LOCAL:
    logger,app_utils, err_handler, utils, const, db, gb_eyedro = res
else:
    logger = res

engines = db.set_db_engines()

MERGED_EXCEL_PATH = const.add_xlsx_dt(const.GB_MERGED_EXCEL_PATH, run_dt.isoformat())
FILTERED_GB_SN_PATH=const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt.isoformat())
GAPS_CSV_PATH=const.add_csv_dt(const.GAPS_CSV_PATH, run_dt.isoformat())
UNIFIER_GB_CSV_PATH = const.add_csv_dt(const.UNIFIER_CSV, run_dt.isoformat())

#!!!!!!!!!!
# if os.path.exists(FILTERED_GB_SN_PATH):
#     filtered_gb_sn_df = pd.read_csv(FILTERED_GB_SN_PATH)
# else:
#     all_gb_api_sn_df, err = err_handler.error_wrapper(lambda: gb_eyedro.api_get_user_info_as_df())
#     if err:
#         logger.error(err)
#         exit(2)

#     filtered_gb_sn_df = all_gb_api_sn_df[0][
#         ~all_gb_api_sn_df[0]["gb_serial"].str.startswith(const.GB_GATEWAY_PREFIX)
#     ]
#     # Save to CSV
#     filtered_gb_sn_df.to_csv(FILTERED_GB_SN_PATH, index=False)

# sns = filtered_gb_sn_df["gb_serial"].str.replace('-', '').tolist()

# eng = db.set_azure_defaultdb_engine()

# gb_eyedro.create_tables(sorted(sns), engine=eng)
# pass
#!!!!!!!!!!!!

def freeze_row1(file_path, insert_cols=[]):
    # Open the workbook with openpyxl (not pandas)
    """
    Open an Excel file and freeze the first row in all sheets, then perform some sheet-specific modifications.

    :param file_path: The path to the Excel file to modify
    :param insert_cols: A list of two-element tuples. The first element is the column number to insert (0-based)
                        and the second element is the value to insert into all cells of that column.
    :return: None
    """
    
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.freeze_panes = "A2"  # Freeze the first row
        if "merged" in sheet_name:
            for col in insert_cols:
                c = col[0]
                v = col[1]
                sheet.insert_cols(c)
                for row in range(1, sheet.max_row + 1):
                    sheet.cell(row=row, column=c).value = v
                    cell = sheet.cell(row=row, column=c)
                    cell.font = Font(bold=True)
        elif "api_unifier" in sheet_name:
            sheet.insert_cols(0)
            sheet.insert_cols(20)
            for row in range(1, sheet.max_row + 1):
                sheet.cell(row=row, column=1).value = "unifier"
                sheet.cell(row=row, column=20).value = "api_matched"
                cell = sheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell = sheet.cell(row=row, column=20)
                cell.font = Font(bold=True)

    # Save the modified workbook
    wb.save(file_path)


def get_duplicate_elements(lst):
    """
    Identify duplicate sublists in a list and return them with their occurrence count.

    This function processes a list of sublists, trimming whitespace from string elements
    and converting each sublist to a tuple for immutability. It tracks the occurrence
    count of each unique sublist. If a sublist appears more than once, it is considered
    a duplicate. The function returns a list of these duplicate sublists along with their
    respective occurrence counts.

    Parameters
    ----------
    lst : list of lists
        The input list containing sublists to be processed.

    Returns
    -------
    list of lists
        A list where each element is a list containing a duplicate sublist and its count.
    """

    seen_count = {}
    duplicates = {}

    for sublist in lst:
        # Trim each string in the sublist and convert to tuple
        trimmed_sublist = tuple(s.strip() if isinstance(s, str) else s for s in sublist)

        if trimmed_sublist in seen_count:
            seen_count[trimmed_sublist] += 1
            # Add to duplicates with current count if this is the first duplicate instance
            if trimmed_sublist not in duplicates:
                duplicates[trimmed_sublist] = seen_count[trimmed_sublist]
            else:
                duplicates[trimmed_sublist] = seen_count[trimmed_sublist]
        else:
            seen_count[trimmed_sublist] = 1

    # Convert dictionary keys back to lists for result
    result = []
    for trimmed_tuple, count in duplicates.items():
        result.append([list(trimmed_tuple), count])

    return result


def excel_to_postgres(
    excel_file_path,
    engine,
    schema="public",
    if_exists="replace",
    chunks_size=1000,
    table_prefix="gb_",
    table_suffix="",
):
    """
    Load all sheets from an Excel file into PostgreSQL database tables.

    Parameters:
    -----------
    excel_file_path : str
        Path to the Excel file
    engine : sqlalchemy.engine.Engine
        The database connection engine
    schema : str, optional
        Schema name to use, if None uses the default schema
    if_exists : str, optional
        How to behave if table exists. Options: 'fail', 'replace', 'append'
    chunk_size : int, optional
        Number of rows to insert at once
    table_prefix : str, optional
        Prefix to add to all table names
    table_suffix : str, optional
        Suffix to add to all table names

    Returns:
    --------
    list
        List of table names created

    Raises:
    -------
    ValueError
        If input parameters are invalid
    FileNotFoundError
        If Excel file not found
    ConnectionError
        If cannot connect to the database
    PermissionError
        If database or schema creation fails due to permissions
    RuntimeError
        For processing errors
    IOError
        For file reading errors
    """

    created_tables = []

    # Input validation
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    if if_exists not in ["fail", "replace", "append"]:
        raise ValueError(
            f"Invalid if_exists value: {if_exists}. Must be 'fail', 'replace', or 'append'"
        )

    if not isinstance(chunks_size, int) or chunks_size <= 0:
        raise ValueError(
            f"Invalid chunk_size: {chunks_size}. Must be a positive integer"
        )

    # Check if database exists, if not create it
    try:
        with engine.connect() as conn:
            logger.info(f"Connected to database: {engine.url}")
        # Read all sheets from the Excel file
        try:
            excel_file = pd.ExcelFile(excel_file_path)
            sheet_names = excel_file.sheet_names
            logger.info(f"Found {len(sheet_names)} sheets in {excel_file_path}")
        except FileNotFoundError:
            raise
        except pd.errors.EmptyDataError:
            raise ValueError(f"Excel file is empty: {excel_file_path}")
        except Exception as e:
            raise IOError(f"Failed to read Excel file {excel_file_path}: {str(e)}")

        for sheet_name in sheet_names:
            try:
                # Read the sheet into a DataFrame
                logger.info(f"Reading sheet: {sheet_name}")
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                except pd.errors.EmptyDataError:
                    logger.warning(f"Sheet '{sheet_name}' is empty, skipping")
                    continue
                except Exception as e:
                    raise IOError(f"Failed to read sheet '{sheet_name}': {str(e)}")

                # Skip if empty
                if df.empty:
                    logger.warning(f"Sheet '{sheet_name}' has no data, skipping")
                    continue

                # Clean up column names for SQL compatibility
                original_columns = df.columns.tolist()
                df.columns = [
                    col.strip().lower().replace(" ", "_").replace("-", "_")
                    for col in df.columns
                ]

                # Log column name changes
                column_changes = {
                    original: new
                    for original, new in zip(original_columns, df.columns)
                    if original != new
                }
                if column_changes:
                    logger.info(
                        f"Column names changed for SQL compatibility: {column_changes}"
                    )

                # Create table name from sheet name
                table_name = f"{table_prefix}{sheet_name.strip().lower().replace(' ', '_').replace('-', '_')}{table_suffix}"

                # Create the full table name with schema if provided
                full_table_name = f"{schema}.{table_name}" if schema else table_name

                try:
                    # Write the DataFrame to PostgreSQL in chunks
                    logger.info(f"Writing {len(df)} rows to table {full_table_name}")
                    df.to_sql(
                        name=table_name,
                        schema=schema,
                        con=engine,
                        if_exists=if_exists,
                        index=False,
                        chunksize=chunks_size,
                    )

                    created_tables.append(full_table_name)
                    logger.info(
                        f"Successfully loaded sheet '{sheet_name}' to table '{full_table_name}'"
                    )
                except sqlalchemy.exc.OperationalError as e:
                    if "permission denied" in str(e).lower():
                        raise PermissionError(
                            f"Permission denied when creating table {full_table_name}: {str(e)}"
                        )
                    else:
                        raise RuntimeError(
                            f"Failed to create table {full_table_name}: {str(e)}"
                        )
                except Exception as e:
                    raise RuntimeError(
                        f"Failed to create/load table {full_table_name}: {str(e)}"
                    )
            except (
                FileNotFoundError,
                IOError,
                ValueError,
                ConnectionError,
                PermissionError,
            ):
                raise
            except Exception as e:
                raise RuntimeError(f"Failed to process sheet {sheet_name}: {str(e)}")

        if not created_tables:
            logger.warning("No tables were created from the Excel file")

        return created_tables

    except Exception as e:
        logger.error(f"Error: {str(e)}")
        logger.error(traceback.format_exc())
        raise


def save_to_postgres(engine, file_path, prefix="gb_"):
    """
    Saves data from an Excel file to PostgreSQL database tables.

    This function utilizes the `excel_to_postgres` function to load data from
    an Excel file into PostgreSQL tables. It handles exceptions related to file
    access, database connection, and data processing. Logging is used to track
    the process and any errors encountered.

    Parameters
    ----------
    engine : sqlalchemy.engine.Engine
        The SQLAlchemy database engine to connect to the PostgreSQL database.
    file_path : str
        Path to the Excel file containing the data to be saved.
    prefix : str, optional
        Prefix to prepend to all table names created in the database, by default "gb_".

    Raises
    ------
    FileNotFoundError
        If the specified Excel file is not found.
    ValueError
        If input parameters are invalid or data cannot be processed.
    ConnectionError
        If there is an issue connecting to the database.
    PermissionError
        If table creation fails due to insufficient permissions.
    IOError
        For errors in reading the file or writing data to the database.
    RuntimeError
        For errors during table creation or data loading.
    """

    try:
        tables = excel_to_postgres(
            excel_file_path=file_path,
            engine=engine,
            schema="public",  # Use 'public' or your preferred schema
            if_exists="replace",  # 'fail', 'replace', or 'append'
            table_prefix=prefix,  # Optional prefix for tables
            chunks_size=5000,  # Adjust based on data size and memory constraints
        )

        logger.info(f"Successfully created {len(tables)} tables:")
        for table in tables:
            logger.info(f"- {table}")

    except FileNotFoundError as e:
        logger.error(f"File not found: {str(e)}")
        sys.exit(1)
    except ValueError as e:
        logger.error(f"Invalid input: {str(e)}")
        sys.exit(2)
    except ConnectionError as e:
        logger.error(f"Connection error: {str(e)}")
        sys.exit(3)
    except PermissionError as e:
        logger.error(f"Permission error: {str(e)}")
        sys.exit(4)
    except IOError as e:
        logger.error(f"I/O error: {str(e)}")
        sys.exit(5)
    except RuntimeError as e:
        logger.error(f"Runtime error: {str(e)}")
        sys.exit(6)
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        logger.error(traceback.format_exc())
        sys.exit(9)


r""" 
Prior to running this script:
1. Download unifier report from https://eu1.unifier.oraclecloud.com/unhcr/bp/route/home/i-unifier?__uref=uuu986636683
    Got to "Unifier Properties"
    Select Reports, User Defined
    Select "smh_GB_Meters-report-v1" owner Steve Hermes
    Run as csv  (downloads as "Report.csv")
    Copy download to "CODE\DATA\gb_unifier_2024-03-18.csv" (rename and add current date (use your local root dir))
    (NOTE: this is handled by outlook classic that moves the daily report from my inbox to inbox->unifier and then saves as a CSV)
2. Run gapfilling script "CODE\unhcr_module\app_time_series_gapfilling_gb_v3.py"

3.

4.
"""

#!!!!!!!!!!!!!!!!!!!
MERGED_EXCEL_PATH = const.add_xlsx_dt(const.GB_MERGED_EXCEL_PATH, run_dt.isoformat())
FILTERED_GB_SN_PATH=const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt.isoformat())
GAPS_CSV_PATH=const.add_csv_dt(const.GAPS_CSV_PATH, run_dt.isoformat())
UNIFIER_GB_CSV_PATH = const.add_csv_dt(const.UNIFIER_CSV, run_dt.isoformat())
msg = None
if os.path.exists(UNIFIER_GB_CSV_PATH):
    # Get file timestamps
    file_stats = os.stat(UNIFIER_GB_CSV_PATH)
    # Get the creation time
    creation_date = datetime.fromtimestamp(file_stats.st_birthtime).date()
    if creation_date < run_dt - timedelta(days=1):
        msg = 'Unifier GB CSV file is older than today'
else:
    msg = 'Unifier GB CSV file does not exist'
if msg:
    res = utils.msgbox_yes_no(msg=f'{msg}, select a different file ?')
    if not res:
        exit(1)
    res = utils.show_dropdown_from_directory(f'{const.DATA_DIR_PATH}','unifier*.csv')
    if utils.selected_file is None:
        exit(1)
    UNIFIER_GB_CSV_PATH = utils.selected_file
    utils.selected_file = None

#!!!!!!!!  AZURE
eng = db.set_azure_defaultdb_engine()

#1. get all valid GB serial numbers, from csv if updated today, otherwise from API and save to csv
# filtered_gb_sn_df ['009-808E0', 'Iraq - CO Baghdad', 1742659998.0, 'Active'] ['gb_serial', 'site_label', 'epoch_utc', 'status']
if os.path.exists(FILTERED_GB_SN_PATH):
    filtered_gb_sn_df = pd.read_csv(FILTERED_GB_SN_PATH)
else:
    all_gb_api_sn_df, err = err_handler.error_wrapper(lambda: gb_eyedro.api_get_user_info_as_df())
    if err:
        logger.error(err)
        exit(2)

    filtered_gb_sn_df = all_gb_api_sn_df[0][
        ~all_gb_api_sn_df[0]["gb_serial"].str.startswith(const.GB_GATEWAY_PREFIX)
    ]
    # Save to CSV
    filtered_gb_sn_df.to_csv(FILTERED_GB_SN_PATH, index=False)

# this is updated in the lines above
df_serial = filtered_gb_sn_df.copy()
df_serial["site_label"] = df_serial["site_label"].fillna("N/A")

#!!!!!!!!! GAPS !!!!!!!!
#2. get gaps for all GB serial numbers
res, err = err_handler.error_wrapper(
    lambda: concurrent_update_gaps(local=True, csv_path=GAPS_CSV_PATH, run_dt=run_dt, force=True)) #concurrent_update_gaps(local=True) ###get_gb_gaps(df_serial.values, const.GB_GAPS_DATA_DIR, run_dt.isoformat()=run_dt.isoformat())
if err:
    logger.error(err)
    exit(3)


# Run the gapfilling script
df_gaps = res[0] #pd.read_csv(GAPS_CSV_PATH)
df_gaps.rename(columns={'gb_sn': 'gb_serial'}, inplace=True)
df_gaps["gb_serial"] = df_gaps["gb_serial"].str[:3] + "-" + df_gaps["gb_serial"].str[3:]

with pd.ExcelWriter(MERGED_EXCEL_PATH, engine="openpyxl") as writer:
    df_serial.to_excel(writer, sheet_name="api_no_dups", index=False)

with pd.ExcelWriter(MERGED_EXCEL_PATH, mode="a", engine="openpyxl") as writer:
    df_gaps.to_excel(writer, sheet_name="gb_gaps", index=False)

merged_df = pd.merge(df_serial, df_gaps, on="gb_serial", how="right")
with pd.ExcelWriter(MERGED_EXCEL_PATH, mode="a", engine="openpyxl") as writer:
    merged_df.to_excel(writer, sheet_name="merged_api_gaps", index=False)

# Create output directory if it doesn't exist
os.makedirs(os.path.dirname(const.TOP20_CSV), exist_ok=True)
# Read the Excel file without filtering
df_top20 = pd.read_excel(const.TOP20_ONEDRIVE_PATH, header=3,sheet_name="GENERAL")
# Save to CSV
df_top20.to_csv(const.TOP20_CSV, index=False)

df_top20 = pd.read_csv(const.TOP20_CSV, encoding="ISO-8859-1")
with pd.ExcelWriter(MERGED_EXCEL_PATH, mode="a", engine="openpyxl") as writer:
    df_top20.to_excel(writer, sheet_name="top20", index=False)

matching_rows = []
# Loop through each row in merged_df
for idx, row in merged_df.iterrows():
    device_serial = row["gb_serial"]
    # Find all matching rows in df_top20
    matches = df_top20[df_top20["Meter Serial No."] == device_serial]

    if not matches.empty:
        if len(matches) == 1:
            # Pick the first match (you can modify this to pick based on any condition)
            selected_match = matches.iloc[0]
        else:
            matches["Date Installed"] = matches["Date Installed"].fillna(
                pd.Timestamp("1900-01-01")
            )

            # Sort matches by 'Date Installed' to get the most recent first
            sorted_matches = matches.sort_values(by="Date Installed", ascending=False)

            # Pick the most recent match (first row after sorting)
            selected_match = sorted_matches.iloc[0]
    else:
        # If no match is found, append the original row from df_serials
        matching_rows.append(row)
        continue
    # Concatenate the original row from df_serials with the selected match
    merged_row = pd.concat([row, selected_match], axis=0)
    matching_rows.append(merged_row)

# Create a new DataFrame from the list of matching rows
matching_df = pd.DataFrame(matching_rows)
matching_df.to_excel(const.GTB_GAPS_EXCEL, index=False)

with pd.ExcelWriter(MERGED_EXCEL_PATH, mode="a", engine="openpyxl") as writer:
    matching_df.to_excel(writer, sheet_name="gtb_merged_top20", index=False)

df1 = pd.read_csv(UNIFIER_GB_CSV_PATH)
df1 = df1.rename(columns={'Status': 'stat_unifier'})
with pd.ExcelWriter(MERGED_EXCEL_PATH, mode="a", engine="openpyxl") as writer:
    df1.to_excel(writer, sheet_name="gb_unifier", index=False)

df2 = pd.read_csv(FILTERED_GB_SN_PATH)

# Drop NaN values
device_serials = df1["Serial Number"].dropna()
site_sns = df2["gb_serial"].dropna()

# Perform fuzzy matching
matches = []
for serial in device_serials:
    best_match = process.extractOne(serial, site_sns, scorer=fuzz.ratio)
    matches.append((serial, best_match[0], best_match[1]))

# Convert matches to DataFrame
matches_df = pd.DataFrame(matches, columns=["DeviceSerial", "MatchedSerial", "Score"])

# Merge df1 and matches_df
merged_df = df1.merge(matches_df, left_on="Serial Number", right_on="DeviceSerial")

# Merge with df2 to get more details from serials sheet
df_final_output = merged_df.merge(
    df2, left_on="MatchedSerial", right_on="gb_serial", how="left"
)

# Filter matches with a high score
df_final_output = df_final_output[df_final_output["Score"] > 99]
with pd.ExcelWriter(MERGED_EXCEL_PATH, mode="a", engine="openpyxl") as writer:
    df_final_output.to_excel(writer, sheet_name="api_unifier_matched", index=False)

# Remove duplicate rows based on all columns
df_no_dups_output = df_final_output.drop_duplicates()
with pd.ExcelWriter(MERGED_EXCEL_PATH, mode="a", engine="openpyxl") as writer:
    df_no_dups_output.to_excel(writer, sheet_name="api_unifier_no_dups", index=False)
freeze_row1(MERGED_EXCEL_PATH, [[11, "top20"], [5, "gaps"], [1, "api_no_dups"]])


# Save the results to an Excel file with two sheets
output_path = "fuzzy_matches.xlsx"
with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
    df_final_output.to_excel(writer, sheet_name="fuzzy", index=False)
    df_no_dups_output.to_excel(writer, sheet_name="no-dups", index=False)

print(f"Matches saved to {output_path}")

MERGED_EXCEL_PATH = const.add_xlsx_dt(const.GB_MERGED_EXCEL_PATH, run_dt.isoformat())
save_to_postgres(eng, MERGED_EXCEL_PATH)
pass

#!!!! TODO: get from DB
