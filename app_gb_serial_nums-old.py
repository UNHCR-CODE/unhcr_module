import csv
from datetime import datetime
import logging
import openpyxl
from openpyxl.styles import Font
import os
import pandas as pd
from rapidfuzz import fuzz
from rapidfuzz import process
import requests
import sqlalchemy
import sys
import traceback
from unhcr import app_utils
from unhcr import err_handler
import unhcr.constants as const
from unhcr import db

run_dt = datetime.now().date().isoformat()

mods = [
    ["app_utils", "app_utils"],
    ["err_handler", "err_handler"],
    ["constants", "const"],
    ["db", "db"],
]

res = app_utils.app_init(
    mods, "unhcr.gb_serial_nums.log", "0.4.7", level="INFO", override=True
)
if const.LOCAL:
    app_utils, err_handler, const, db = res

engines = db.set_db_engines()


# top20_master_path = r"E:\UNHCR\OneDrive - UNHCR\Green Data Team\07 Greenbox Management\Green Box - TOP 20 Countries\New Top 20.xlsx"
# # Create output directory if it doesn't exist
# os.makedirs(os.path.dirname(top20_csv), exist_ok=True)

# # Read the Excel file without filtering
# df_top20 = pd.read_excel(top20_master_path, header=3,sheet_name="GENERAL")

# # Save to CSV
# df.to_csv(top20_csv, index=False)
# pass

def freeze_row1(file_path, insert_cols=[]):
    # Open the workbook with openpyxl (not pandas)
    wb = openpyxl.load_workbook(file_path)

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet.freeze_panes = "A2"  # Freeze the first row
        if "merged" in sheet_name:
            for col in insert_cols:
                c = col[0]
                v = col[1]
                sheet.insert_cols(c)
                for row in range(1, sheet.max_row + 1):
                    sheet.cell(row=row, column=c).value = v
                    cell = sheet.cell(row=row, column=c)
                    cell.font = Font(bold=True)
        elif "api_unifier" in sheet_name:
            sheet.insert_cols(0)
            sheet.insert_cols(20)
            for row in range(1, sheet.max_row + 1):
                sheet.cell(row=row, column=1).value = "unifier"
                sheet.cell(row=row, column=20).value = "api_matched"
                cell = sheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell = sheet.cell(row=row, column=20)
                cell.font = Font(bold=True)

    # Save the modified workbook
    wb.save(file_path)


def get_duplicate_elements(lst):
    seen_count = {}
    duplicates = {}

    for sublist in lst:
        # Trim each string in the sublist and convert to tuple
        trimmed_sublist = tuple(s.strip() if isinstance(s, str) else s for s in sublist)

        if trimmed_sublist in seen_count:
            seen_count[trimmed_sublist] += 1
            # Add to duplicates with current count if this is the first duplicate instance
            if trimmed_sublist not in duplicates:
                duplicates[trimmed_sublist] = seen_count[trimmed_sublist]
            else:
                duplicates[trimmed_sublist] = seen_count[trimmed_sublist]
        else:
            seen_count[trimmed_sublist] = 1

    # Convert dictionary keys back to lists for result
    result = []
    for trimmed_tuple, count in duplicates.items():
        result.append([list(trimmed_tuple), count])

    return result


def get_gb_user_info_data():
    url = (
        const.GB_API_V1_API_BASE_URL
        + const.GB_API_V1_GET_DEVICE_LIST
        + const.GB_API_V1_USER_KEY
    )
    payload = {}
    headers = {}

    response, err = err_handler.request(
        lambda: requests.request("GET", url, headers=headers, data=payload)
    )
    if err:
        return None, err
    user_info_data, err = err_handler.error_wrapper(lambda: response.json())
    if err:
        return None, err
    return user_info_data, None


def parse_user_info_as_df(user_info_data, csv_path = const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt)):
    top_level = user_info_data.get("UserInfo", {}).get("DeviceSerialList", [])

    top_level_serials = []
    for sn in top_level:
        top_level_serials.append(str(sn)[:3] + "-" + str(sn)[3:])

    # Extract DeviceSerials and SiteLabels from DisplayGroupList
    device_data = []
    for site in user_info_data.get("UserInfo", {}).get("SiteList", []):
        site_label = site.get("SiteLabel")
        for group in site.get("DisplayGroupList", []):
            for device in group.get("DeviceList", []):
                sn = device.get("DeviceSerial")
                sn = str(sn)[:3] + "-" + str(sn)[3:]
                last_com = device.get("LastCommSecUtc")
                state = device.get("State")
                device_data.append((site_label, sn, last_com, state))

    # Convert device_data into a dictionary for quick lookup
    serial_info = {item[1]: item for item in device_data}

    # Step 1: Merge data while keeping all of top_level_serials
    merged_data = []
    added_serials = set()

    # Add serials from top_level_serials, along with their details if available
    for sn in top_level_serials:
        if sn in serial_info:
            merged_data.append([sn] + list(serial_info[sn][0:1]) + list(serial_info[sn][2:]))
            added_serials.add(sn)
        else:
            merged_data.append([sn])
            added_serials.add(sn)

    # Step 2: Add any row from device_data that did not match top_level_serials
    for item in serial_info.values():
        serial_num = item[1]
        if serial_num not in added_serials:
            merged_data.append([serial_num, item[0], item[2], item[3]])

    # Step 3: Remove standalone serials if a detailed entry exists
    final_data = []
    seen_serials = set()

    for row in merged_data:
        serial_num = row[0]
        if serial_num in seen_serials:
            continue  # Skip duplicate standalone serials
        seen_serials.add(serial_num)
        if len(row) == 1 and any(r[0] == serial_num and len(r) > 1 for r in merged_data):
            continue  # Remove standalone if detailed version exists
        final_data.append(row)

    # Optionally: Get the count of final rows
    logging.debug(f"Final row count: {len(final_data)}")
    # Convert merged data to a pandas DataFrame
    final_df = pd.DataFrame(final_data, columns=const.GB_SN_COLS)

    # Save to CSV
    if csv_path:
        final_df.to_csv(csv_path, index=False)
    return final_df


def get_user_info_as_df():
    # Extract top-level DeviceSerialList
    """
    Gets the user info data from the API, parses it into a list of tuples of
    (site_label, device_serial), and returns the list or an error string.

    Returns:
        list of tuples: (site_label, device_serial) or
        str: error string
    """
    user_info_data, err = get_gb_user_info_data()
    if err or len(user_info_data["Errors"]) != 0:
        err_str = f"app_gb_serial_nums: Failed to get user info data ERROR: {err}  {user_info_data['Errors']}"
        logging.error(err_str)
        return None, err_str
    all_serials_df, err = err_handler.error_wrapper(
        lambda: parse_user_info_as_df(user_info_data)
    )
    if err:
        err_str = f"app_gb_serial_nums: Failed to parse user info data ERROR: {err}"
        logging.error(err_str)
        return None, err_str
    return all_serials_df, None


def excel_to_postgres(
    excel_file_path,
    engine,
    schema="public",
    if_exists="replace",
    chunks_size=1000,
    table_prefix="gb_",
    table_suffix="",
):
    """
    Load all sheets from an Excel file into PostgreSQL database tables.

    Parameters:
    -----------
    excel_file_path : str
        Path to the Excel file
    schema : str, optional
        Schema name to use, if None uses the default schema
    if_exists : str, optional
        How to behave if table exists. Options: 'fail', 'replace', 'append'
    chunk_size : int, optional
        Number of rows to insert at once
    table_prefix : str, optional
        Prefix to add to all table names
    table_suffix : str, optional
        Suffix to add to all table names

    Returns:
    --------
    list
        List of table names created

    Raises:
    -------
    ValueError
        If input parameters are invalid
    FileNotFoundError
        If Excel file not found
    ConnectionError
        If cannot connect to the database
    PermissionError
        If database or schema creation fails due to permissions
    RuntimeError
        For processing errors
    IOError
        For file reading errors
    """
    created_tables = []

    # Input validation
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Excel file not found: {excel_file_path}")

    if if_exists not in ["fail", "replace", "append"]:
        raise ValueError(
            f"Invalid if_exists value: {if_exists}. Must be 'fail', 'replace', or 'append'"
        )

    if not isinstance(chunks_size, int) or chunks_size <= 0:
        raise ValueError(
            f"Invalid chunk_size: {chunks_size}. Must be a positive integer"
        )

    # Check if database exists, if not create it
    try:
        with engine.connect() as conn:
            logging.info(f"Connected to database: {engine.url}")
        # Read all sheets from the Excel file
        try:
            excel_file = pd.ExcelFile(excel_file_path)
            sheet_names = excel_file.sheet_names
            logging.info(f"Found {len(sheet_names)} sheets in {excel_file_path}")
        except FileNotFoundError:
            raise
        except pd.errors.EmptyDataError:
            raise ValueError(f"Excel file is empty: {excel_file_path}")
        except Exception as e:
            raise IOError(f"Failed to read Excel file {excel_file_path}: {str(e)}")

        for sheet_name in sheet_names:
            try:
                # Read the sheet into a DataFrame
                logging.info(f"Reading sheet: {sheet_name}")
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                except pd.errors.EmptyDataError:
                    logging.warning(f"Sheet '{sheet_name}' is empty, skipping")
                    continue
                except Exception as e:
                    raise IOError(f"Failed to read sheet '{sheet_name}': {str(e)}")

                # Skip if empty
                if df.empty:
                    logging.warning(f"Sheet '{sheet_name}' has no data, skipping")
                    continue

                # Clean up column names for SQL compatibility
                original_columns = df.columns.tolist()
                df.columns = [
                    col.strip().lower().replace(" ", "_").replace("-", "_")
                    for col in df.columns
                ]

                # Log column name changes
                column_changes = {
                    original: new
                    for original, new in zip(original_columns, df.columns)
                    if original != new
                }
                if column_changes:
                    logging.info(
                        f"Column names changed for SQL compatibility: {column_changes}"
                    )

                # Create table name from sheet name
                table_name = f"{table_prefix}{sheet_name.strip().lower().replace(' ', '_').replace('-', '_')}{table_suffix}"

                # Create the full table name with schema if provided
                full_table_name = f"{schema}.{table_name}" if schema else table_name

                try:
                    # Write the DataFrame to PostgreSQL in chunks
                    logging.info(f"Writing {len(df)} rows to table {full_table_name}")
                    df.to_sql(
                        name=table_name,
                        schema=schema,
                        con=engine,
                        if_exists=if_exists,
                        index=False,
                        chunksize=chunks_size,
                    )

                    created_tables.append(full_table_name)
                    logging.info(
                        f"Successfully loaded sheet '{sheet_name}' to table '{full_table_name}'"
                    )
                except sqlalchemy.exc.OperationalError as e:
                    if "permission denied" in str(e).lower():
                        raise PermissionError(
                            f"Permission denied when creating table {full_table_name}: {str(e)}"
                        )
                    else:
                        raise RuntimeError(
                            f"Failed to create table {full_table_name}: {str(e)}"
                        )
                except Exception as e:
                    raise RuntimeError(
                        f"Failed to create/load table {full_table_name}: {str(e)}"
                    )
            except (
                FileNotFoundError,
                IOError,
                ValueError,
                ConnectionError,
                PermissionError,
            ):
                raise
            except Exception as e:
                raise RuntimeError(f"Failed to process sheet {sheet_name}: {str(e)}")

        if not created_tables:
            logging.warning("No tables were created from the Excel file")

        return created_tables

    except Exception as e:
        logging.error(f"Error: {str(e)}")
        logging.error(traceback.format_exc())
        raise


def save_to_postgres(engine, file_path, prefix='gb_'):
    """Example usage of the excel_to_postgres function"""
    try:
        tables = excel_to_postgres(
            excel_file_path=file_path,
            engine=engine,
            schema="public",  # Use 'public' or your preferred schema
            if_exists="replace",  # 'fail', 'replace', or 'append'
            table_prefix=prefix,  # Optional prefix for tables
            chunks_size=5000,  # Adjust based on data size and memory constraints
        )

        logging.info(f"Successfully created {len(tables)} tables:")
        for table in tables:
            logging.info(f"- {table}")

    except FileNotFoundError as e:
        logging.error(f"File not found: {str(e)}")
        sys.exit(1)
    except ValueError as e:
        logging.error(f"Invalid input: {str(e)}")
        sys.exit(2)
    except ConnectionError as e:
        logging.error(f"Connection error: {str(e)}")
        sys.exit(3)
    except PermissionError as e:
        logging.error(f"Permission error: {str(e)}")
        sys.exit(4)
    except IOError as e:
        logging.error(f"I/O error: {str(e)}")
        sys.exit(5)
    except RuntimeError as e:
        logging.error(f"Runtime error: {str(e)}")
        sys.exit(6)
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        logging.error(traceback.format_exc())
        sys.exit(9)

#!!!!! save greening the blue 2024 spreadsheet & merged spreadsheet
# gtb_excel_path = r'E:\UNHCR\OneDrive - UNHCR\Energy Team\Concept development\AZURE DATA\Greening the Blue\20240319_2024_GB_Data_v5.xlsx'
# save_to_postgres(engines[1], gtb_excel_path, prefix = '')
# save_to_postgres(engines[0], gtb_excel_path, prefix = '')

# merged_excel_path = const.add_xlsx_dt(const.GB_MERGED_EXCEL_PATH, run_dt)
# )
# save_to_postgres(engines[1], merged_excel_path, prefix = 'gb_')
# save_to_postgres(engines[0], merged_excel_path, prefix = 'gb_')

# pass

#!!!!!!!!!!!!!!!!!!!
r""" 
Prior to running this script:
1. Download unifier report from https://eu1.unifier.oraclecloud.com/unhcr/bp/route/home/i-unifier?__uref=uuu986636683
    Got to "Unifier Properties"
    Select Reports, User Defined
    Select "smh_GB_Meters-report-v1" owner Steve Hermes
    Run as csv  (downloads as "Report.csv")
    Copy download to "CODE\DATA\gb_unifier_2024-03-18.csv" (rename and add current date (use your local root dir))
    
2. Run gapfilling script "CODE\unhcr_module\app_time_series_gapfilling_gb_v3.py"

3.

4.
"""

#!!!!!!!!!!!!!!!!!!!

merged_excel_path = const.add_xlsx_dt(const.GB_MERGED_EXCEL_PATH, run_dt)

all_gb_api_sn_df, err = err_handler.error_wrapper(lambda: get_user_info_as_df())
if err:
    logging.error(err)
    exit(1)

filtered_sn_df =  all_gb_api_sn_df[0][~all_gb_api_sn_df[0]['gb_serial'].str.startswith(const.GB_GATEWAY_PREFIX)]


# Write to CSV file
path_api_gb_sn = const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt)
with open(path_api_gb_sn, "w", newline="") as file:
    writer = csv.writer(file)
    writer.writerow(["site_label", "gb_serial", "last_com", "state"])
    writer.writerows(all_gb_api_sn)

# Print the result
x = 1
for item in all_gb_api_sn:
    print(
        f" {x}  device_serial: {item[1]}  site_label: {item[0] if item[0] else 'N/A'}"
    )
    x += 1
# #!!!!! gen all_api_gbs csv with eyedro label & serial

df_serial = pd.read_csv(path_api_gb_sn)
df_serial["site_label"] = df_serial["site_label"].fillna("N/A")

df_gaps = pd.read_csv(const.GAPS_CSV)
df_gaps["gb_serial"] = df_gaps["gb_serial"].str[:3] + "-" + df_gaps["gb_serial"].str[3:]

with pd.ExcelWriter(merged_excel_path, engine="openpyxl") as writer:
    df_serial.to_excel(writer, sheet_name="api_no_dups", index=False)

with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_gaps.to_excel(writer, sheet_name="gb_gaps", index=False)


merged_df = pd.merge(df_serial, df_gaps, on="gb_serial", how="right")
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    merged_df.to_excel(writer, sheet_name="merged_api_gaps", index=False)

df_top20 = pd.read_csv(const.TOP20_CSV, encoding="ISO-8859-1")
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_top20.to_excel(writer, sheet_name="top20", index=False)

matching_rows = []
# Loop through each row in merged_df
for idx, row in merged_df.iterrows():
    device_serial = row["gb_serial"]
    # Find all matching rows in df_top20
    matches = df_top20[df_top20["Meter Serial No."] == device_serial]

    if not matches.empty:
        if len(matches) == 1:
            # Pick the first match (you can modify this to pick based on any condition)
            selected_match = matches.iloc[0]
        else:
            matches["Date Installed"] = matches["Date Installed"].fillna(
                pd.Timestamp("1900-01-01")
            )

            # Sort matches by 'Date Installed' to get the most recent first
            sorted_matches = matches.sort_values(by="Date Installed", ascending=False)

            # Pick the most recent match (first row after sorting)
            selected_match = sorted_matches.iloc[0]
    else:
        # If no match is found, append the original row from df_serials
        matching_rows.append(row)
        continue
    # Concatenate the original row from df_serials with the selected match
    merged_row = pd.concat([row, selected_match], axis=0)
    matching_rows.append(merged_row)

# Create a new DataFrame from the list of matching rows
matching_df = pd.DataFrame(matching_rows)
matching_df.to_excel(const.GTB_GAPS_EXCEL, index=False)

with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    matching_df.to_excel(writer, sheet_name="gtb_merged_top20", index=False)

#!!!!!!!!!!!!!!!!!!!!!!!!!!!

df1 = pd.read_csv(const.UNIFIER_CSV)
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df1.to_excel(writer, sheet_name="gb_unifier", index=False)

df2 = pd.read_csv(const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt))

# Drop NaN values
device_serials = df1["Serial Number"].dropna()
site_sns = df2["gb_serial"].dropna()

# Perform fuzzy matching
matches = []
for serial in device_serials:
    best_match = process.extractOne(serial, site_sns, scorer=fuzz.ratio)
    matches.append((serial, best_match[0], best_match[1]))

# Convert matches to DataFrame
matches_df = pd.DataFrame(matches, columns=["DeviceSerial", "MatchedSerial", "Score"])

# Merge df1 and matches_df
merged_df = df1.merge(matches_df, left_on="Serial Number", right_on="DeviceSerial")

# Merge with df2 to get more details from serials sheet
df_final_output = merged_df.merge(
    df2, left_on="MatchedSerial", right_on="gb_serial", how="left"
)

# Filter matches with a high score
df_final_output = df_final_output[df_final_output["Score"] > 99]
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_final_output.to_excel(writer, sheet_name="api_unifier_matched", index=False)

# Remove duplicate rows based on all columns
df_no_dups_output = df_final_output.drop_duplicates()
with pd.ExcelWriter(merged_excel_path, mode="a", engine="openpyxl") as writer:
    df_no_dups_output.to_excel(writer, sheet_name="api_unifier_no_dups", index=False)
freeze_row1(merged_excel_path, [[11, "top20"], [5, "gaps"], [1, "api_no_dups"]])


# Save the results to an Excel file with two sheets
output_path = "fuzzy_matches.xlsx"
with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
    df_final_output.to_excel(writer, sheet_name="fuzzy", index=False)
    df_no_dups_output.to_excel(writer, sheet_name="no-dups", index=False)

print(f"Matches saved to {output_path}")

merged_excel_path = const.add_xlsx_dt(const.GB_MERGED_EXCEL_PATH, run_dt)
save_to_postgres(engines[1], merged_excel_path)
pass

#!!!! TODO: get from DB
