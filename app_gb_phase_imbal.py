from concurrent.futures import ProcessPoolExecutor
from datetime import datetime, timezone
import re
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import os
import pandas as pd
from pathlib import Path
from psycopg2 import DatabaseError
from sklearn.cluster import KMeans
from sklearn.preprocessing import MinMaxScaler

from sqlalchemy import text
from unhcr import app_utils
from unhcr import constants as const
from unhcr import db
from unhcr import gb_eyedro

mods=[["app_utils", "app_utils"], ["constants", "const"], ["db", "db"], ["gb_eyedro", "gb_eyedro"]]
res = app_utils.app_init(mods=mods, log_file="unhcr.gb_phase_imbal.log", version="0.4.7", level="INFO", override=True)
logger = res[0]
if const.LOCAL:
    logger, app_utils, const, db, gb_eyedro = res

def process_phase_dataset(df, site_name, normalize=False, n_clusters=4):
    """
    Process power load data to analyze consumption patterns and phase imbalance.
    
    Parameters:
    -----------
    df : pandas DataFrame
        Input dataframe containing hourly power data
    site_name : str
        Name of the site being analyzed
    normalize : bool
        Whether to normalize features before clustering
    n_clusters : int
        Number of clusters to use in KMeans algorithm
        
    Returns:
    --------
    DataFrame with added analysis columns
    """
    # Rename columns to match our processing
    df = df.copy()
    
    # Map the columns to the expected format
    column_mapping = {
        'avg_wh_p1': 'P1_Wh',
        'avg_wh_p2': 'P2_Wh', 
        'avg_wh_p3': 'P3_Wh'
    }
    
    df = df.rename(columns=column_mapping)
    
    # Calculate phase columns
    phase_cols = ["P1_Wh", "P2_Wh", "P3_Wh"]
    df["total_load"] = df[phase_cols].sum(axis=1)
    df["avg_load"] = df[phase_cols].mean(axis=1)
    
    # Calculate additional metrics
    df["std_load"] = df[phase_cols].std(axis=1)
    df["cv_load"] = (df["std_load"] / df["avg_load"] * 100).round(1)  # Coefficient of variation
    
    # Convert percentage string to float if exists, otherwise calculate
    if 'phase_imbalance' in df.columns:
        df["imbalance"] = df["phase_imbalance"]######.str.rstrip('%').astype(float) / 100
    else:
        df["imbalance"] = (
            (df[phase_cols].max(axis=1) - df[phase_cols].min(axis=1)) /
            df[phase_cols].max(axis=1)
        ).round(3).fillna(0)
    
    # Add time-based features if hr_utc is present
    if 'hr_utc' in df.columns:
        df["hour"] = df["hr_utc"]
        df["time_category"] = pd.cut(
            df["hour"], 
            bins=[0, 6, 12, 18, 24], 
            labels=["Night", "Morning", "Afternoon", "Evening"],
            include_lowest=True,
            right=False
        )
    
    # Calculate max-to-avg ratio (peak factor)
    df["peak_factor"] = (df[phase_cols].max(axis=1) / df["avg_load"]).round(2)
    
    # Prepare features for clustering
    features = df[["total_load", "imbalance"]].copy()
    if normalize:
        scaler = MinMaxScaler()
        features[["total_load", "imbalance"]] = scaler.fit_transform(features)
    
    # Apply KMeans clustering
    kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
    df["cluster"] = kmeans.fit_predict(features)
    centroids = kmeans.cluster_centers_
    
    # Classify clusters
    #median_load = df["total_load"].median()
    #median_imb = df["imbalance"].median()
    high_limit = 5000
    imb_limit = 0.5
    cluster_category_map = {}
    for i, (load, imb) in enumerate(centroids):
        if load < high_limit and imb < imb_limit:
            category = "Low Load & Low Imbalance"
        elif load >= high_limit and imb < imb_limit:
            category = "High Load & Low Imbalance"
        elif load < high_limit and imb >= imb_limit:
            category = "Low Load & High Imbalance"
        else:
            category = "High Load & High Imbalance"
        cluster_category_map[i] = category
    
    # Add category and site information
    df["category"] = df["cluster"].map(cluster_category_map)
    df["site_name"] = site_name
    
    # Calculate site summary metrics
    df.attrs["site_summary"] = {
        "site_name": site_name,
        "avg_total_load": df["total_load"].mean(),
        "max_total_load": df["total_load"].max(),
        "avg_imbalance": df["imbalance"].mean(),
        "max_imbalance": df["imbalance"].max(),
        "source": df["brand"].iloc[0] if "brand" in df.columns else 'N/A',
        "worst_hour": df.loc[df["imbalance"].idxmax(), "hr_utc"] if "hr_utc" in df.columns else None,
        "dominant_category": df["category"].value_counts().index[0],
        "high_imbalance_hours": (df["imbalance"] > 0.5).sum(),
        "peak_load_hour": df.loc[df["total_load"].idxmax(), "hr_utc"] if "hr_utc" in df.columns else None
    }
    
    return df

def process_phase_csv(csv_path, normalize=False, n_clusters=4):
    """Helper function to process a single file for parallel execution"""
    site_name = csv_path.stem
    try:
        # Read CSV with custom format
        df = pd.read_csv(csv_path, encoding="latin-1")
        
        # Check for required columns
        required_cols = {"avg_wh_p1", "avg_wh_p2", "avg_wh_p3"}
        if not required_cols.issubset(df.columns):
            return None, f"⚠️ Skipping {site_name}: missing required columns. Columns found: {df.columns}"
        
        df_result = process_phase_dataset(df, site_name, normalize=normalize, n_clusters=n_clusters)
        return df_result, f"✅ Processed: {site_name}"
    except Exception as e:
        return None, f"❌ Error processing {site_name}: {e}"


def apply_percent_format(ws, col_letter, start_row, end_row):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=col_letter, max_col=col_letter):
        for cell in row:
            cell.number_format = '0.0%'


def apply_percent_format_by_name(ws, df, col_names):
    for col_name in col_names:
        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = get_column_letter(col_idx)
        apply_percent_format(ws, col_idx, 2, len(df)+1) 


def custom_format_sheet(ws, df, col_name, colors):
    col_idx = df.columns.get_loc(col_name) + 1  # openpyxl is 1-based
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{len(df)+1}"  # +1 for header row

    rule = ColorScaleRule(
        start_type='min', start_color=colors[0],
        mid_type='percentile', mid_value=50, mid_color=colors[1],
        end_type='max', end_color=colors[2],
    )

    ws.conditional_formatting.add(cell_range, rule)


def create_phase_summary_sheet(writer, all_summaries):
    """Create a summary sheet with key metrics from all sites"""
    summary_df = pd.DataFrame(all_summaries)
    
    # Sort by imbalance and total load
    summary_df = summary_df.sort_values(by=["avg_imbalance", "avg_total_load"], ascending=[False, False])
    # Add conditional formatting

    workbook = writer.book
    worksheet = writer.sheets["Site_Summary"]
    custom_format_sheet(worksheet, summary_df, 'avg_imbalance', ['63BE7B', 'FFEB84', 'F8696B'])
    custom_format_sheet(worksheet, summary_df, 'max_imbalance', ['63BE7B', 'FFEB84', 'F8696B'])
    custom_format_sheet(worksheet, summary_df, 'avg_total_load', ['F8696B', 'FFEB84', '63BE7B'])
    apply_percent_format_by_name(worksheet, summary_df, ['avg_imbalance', 'max_imbalance'])
    
    # workbook = writer.book 
    # percent_format = workbook.add_format({'num_format': '0.0%'})
    # worksheet.set_column(3, 3, None, percent_format)
    # worksheet.set_column(4, 4, None, percent_format)
    return summary_df

def generate_phase_plots(df_dict, output_folder):
    """Generate and save analysis plots for each site"""
    os.makedirs(output_folder, exist_ok=True)
    
    for site_name, df in df_dict.items():
        try:
            print(f"Generating plots for {site_name}...")
            # Create figure with subplots
            fig, axes = plt.subplots(2, 2, figsize=(16, 12))
            fig.suptitle(f"Analysis for {site_name}", fontsize=16)
            
            # Plot 1: Phase loads over time
            phase_cols = ["P1_Wh", "P2_Wh", "P3_Wh"]
            for col in phase_cols:
                axes[0, 0].plot(df["hr_utc"] if "hr_utc" in df.columns else df.index, 
                               df[col], label=col)
            axes[0, 0].set_title("Phase Loads Over Time")
            axes[0, 0].set_xlabel("Hour")
            axes[0, 0].set_ylabel("Power (Wh)")
            axes[0, 0].legend()
            axes[0, 0].grid(True)
            
            # Plot 2: Imbalance over time
            axes[0, 1].plot(df["hr_utc"] if "hr_utc" in df.columns else df.index, 
                           df["imbalance"], color='red')
            axes[0, 1].set_title("Phase Imbalance Over Time")
            axes[0, 1].set_xlabel("Hour")
            axes[0, 1].set_ylabel("Imbalance (ratio)")
            axes[0, 1].grid(True)
            
            # Plot 3: Cluster scatter plot
            scatter = axes[1, 0].scatter(df["total_load"], df["imbalance"], 
                                        c=df["cluster"], cmap="viridis", s=50, alpha=0.7)
            axes[1, 0].set_title("Load vs Imbalance Clusters")
            axes[1, 0].set_xlabel("Total Load (Wh)")
            axes[1, 0].set_ylabel("Imbalance (ratio)")
            legend1 = axes[1, 0].legend(*scatter.legend_elements(),
                                      title="Clusters")
            axes[1, 0].add_artist(legend1)
            axes[1, 0].grid(True)
            
            # Plot 4: Category distribution
            df["category"].value_counts().plot(kind='pie', ax=axes[1, 1], autopct='%1.1f%%')
            axes[1, 1].set_title("Distribution of Categories")
            
            plt.tight_layout()
            plt.subplots_adjust(top=0.92)
            
            # Save the figure
            fig_path = os.path.join(output_folder, f"{site_name}_analysis.png")
            plt.savefig(fig_path)
            plt.close(fig)
            
        except Exception as e:
            print(f"Could not generate plots for {site_name}: {e}")

def classify_imbalance(
    ver: str,
    folder_path: str,
    output_xlsx: str = "classified_sites_report.xlsx",
    normalize: bool = False,
    top_n: int = 10,
    n_clusters: int = 4,
    parallel: bool = True,
    generate_images: bool = True
):
    """
    Process multiple CSV files containing power data and generate comprehensive reports.
    
    Parameters:
    -----------
    folder_path : str
        Path to folder containing CSV files
    output_xlsx : str
        Path for output Excel report
    normalize : bool
        Whether to normalize features before clustering
    top_n : int
        Number of top problematic sites to highlight
    n_clusters : int
        Number of clusters to use in KMeans algorithm
    parallel : bool 
        Whether to use parallel processing
    generate_images : bool
        Whether to generate analysis plots for sites
        
    Returns:
    --------
    Path to saved Excel report
    """
    start_time = datetime.now()
    folder = Path(folder_path)
    csv_files = list(folder.glob("*.csv"))
    print(f"Found {len(csv_files)} CSV files to process")
    
    base_wb = load_workbook(output_xlsx.replace('_report', '_template'))
    output_xlsx = output_xlsx.replace('.xlsx', f'_{ver}.xlsx')
    base_wb.save(output_xlsx)
    pass
    
    all_results = []
    result_dict = {}
    top_sites = []
    all_summaries = []
    
    # Process files (either in parallel or sequentially)
    if parallel and len(csv_files) > 1:
        with ProcessPoolExecutor() as executor:
            futures = {executor.submit(process_phase_csv, csv, normalize, n_clusters): csv for csv in csv_files}
            for future in futures:
                df_result, message = future.result()
                print(message)
                if df_result is not None:
                    site_name = df_result["site_name"].iloc[0]
                    result_dict[site_name] = df_result
                    all_results.append(df_result)
                    all_summaries.append(df_result.attrs["site_summary"])
                    
                    # Log top-N worst (high imbalance & high load)
                    df_top = df_result.sort_values(["category",  "imbalance", "total_load"], ascending=[True, False, False])
                    top_sites.append(df_top.head(top_n))
    else:
        for csv in csv_files:
            df_result, message = process_phase_csv(csv, normalize, n_clusters)
            print(message)
            if df_result is not None:
                site_name = df_result["site_name"].iloc[0]
                result_dict[site_name] = df_result
                all_results.append(df_result)

                all_summaries.append(df_result.attrs["site_summary"])

                # Log top-N worst (high imbalance & high load)
                df_top = df_result.sort_values(["category", "imbalance", "total_load"], ascending=[True, False, False])
                top_sites.append(df_top.head(top_n))

    # If no valid results were processed, exit
    if not all_results:
        print("No valid results were processed. Check the error messages above.")
        return None

    # Generate Excel report
    with pd.ExcelWriter(output_xlsx, engine='openpyxl', mode="a", if_sheet_exists="overlay") as writer:
        # Create summary sheet first
        summary_df = pd.DataFrame(all_summaries)
        # Sort by avg_total_load (descending) and avg_imbalance (ascending)
        summary_df = summary_df.sort_values(["avg_imbalance", "avg_total_load"], ascending=[False, False])
        summary_df.to_excel(writer, sheet_name="Site_Summary", index=False)
        create_phase_summary_sheet(writer, all_summaries)
        worksheet = writer.sheets["Site_Summary"]
        worksheet.freeze_panes = worksheet["A2"]

        # Combine all results
        workbook = writer.book
        #####percent_format = workbook.add_format({'num_format': '0.0%'})
        combined_df = pd.concat(all_results)
        combined_df.drop(columns=["phase_imbalance"], inplace=True)
        combined_df.to_excel(writer, sheet_name="All_Sites", index=False)
        worksheet = writer.sheets["All_Sites"]
        worksheet.freeze_panes = worksheet["A2"]
        custom_format_sheet(worksheet, combined_df, 'avg_load', ['F8696B', 'FFEB84', '63BE7B'])
        custom_format_sheet(worksheet, combined_df, 'imbalance', ['63BE7B', 'FFEB84', 'F8696B'])
        apply_percent_format_by_name(worksheet, combined_df, ['imbalance', 'peak_factor'])
        # worksheet.set_column(9, 9, None, percent_format)
        # worksheet.set_column(12, 12, None, percent_format)
        # Combine worst sites from all datasets
        if top_sites:
            worst_df = pd.concat(top_sites).sort_values(["category", "total_load", "imbalance"], ascending=[True, False, False])
            worst_df = worst_df.head(top_n)
            worst_df.drop(columns=["phase_imbalance"], inplace=True)
            worst_df.to_excel(writer, sheet_name="Top_Worst_Sites", index=False)
            worksheet = writer.sheets["Top_Worst_Sites"]
            worksheet.freeze_panes = worksheet["A2"]
            custom_format_sheet(worksheet, worst_df, 'avg_load', ['F8696B', 'FFEB84', '63BE7B'])
            custom_format_sheet(worksheet, worst_df, 'imbalance', ['63BE7B', 'FFEB84', 'F8696B'])
            apply_percent_format_by_name(worksheet, worst_df, ['imbalance', 'peak_factor'])
            # worksheet.set_column(9, 9, None, percent_format)
            # worksheet.set_column(12, 12, None, percent_format)
                # Add individual site sheets
        for df_result in all_results:
            df_result.drop(columns=["phase_imbalance"], inplace=True)
            site_name = df_result["site_name"].iloc[0]
            sheet_name = site_name[:31]  # Excel sheet name length limit
            df_result.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Add conditional formatting to the imbalance column
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Get the column index for imbalance
            custom_format_sheet(worksheet, df_result, 'avg_load', ['F8696B', 'FFEB84', '63BE7B'])
            custom_format_sheet(worksheet, df_result, 'imbalance', ['63BE7B', 'FFEB84', 'F8696B'])
            apply_percent_format_by_name(worksheet, df_result, ['imbalance', 'peak_factor'])
            # imb_col_idx = df_result.columns.get_loc("imbalance")
            # worksheet.set_column(imb_col_idx, imb_col_idx, None, percent_format)
            # pf_col_idx = df_result.columns.get_loc("peak_factor")
            # worksheet.set_column(pf_col_idx, pf_col_idx, None, percent_format)

    # Generate plots if requested
    if generate_images:
        images_folder = os.path.splitext(output_xlsx)[0] + "_plots"
        generate_phase_plots(result_dict, images_folder)
        print(f"📊 Analysis plots saved to: {os.path.abspath(images_folder)}")
    
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    
    print(f"\n🎉 Batch classification complete! Report saved to: {os.path.abspath(output_xlsx)}")
    print(f"⏱️ Processing completed in {duration:.2f} seconds")
    print(f"📑 Processed {len(all_results)} sites successfully")
    
    return output_xlsx

# -----------------------------
# 🧪 Usage Example:
# classify_imbalance(
#     "datasets/",
#     normalize=True,
#     top_n=10,
#     parallel=True,
#     generate_images=True
# )
# -----------------------------




def sanitize_filename(filename, replacement="_"):
    # Define illegal characters for Windows filenames
    illegal_chars = r'[<>:"/\\|?*\x00-\x1F\[\]]'  # Includes control characters (0-31)
    
    # Replace illegal characters with the specified replacement (default: "_")
    sanitized = re.sub(illegal_chars, replacement, filename)
    
    # Trim trailing dots and spaces (not allowed in Windows filenames)
    return sanitized.strip(" .")

def phase_imbalance(db_eng, sn_array,out_path):
    
    gb_office_path = out_path + r'\gb-office.csv'
    # Load CSV into DataFrame
    df_gb_office = pd.read_csv(gb_office_path, dtype={"serial": str}, encoding="ISO-8859-1")  # Ensure serial is treated as a string

    #out_path = r'E:\_UNHCR\CODE\DATA\archive'    #r'E:\UNHCR\OneDrive - UNHCR\Energy Team\Concept development\AZURE DATA\phase_imbalance\gensets'

    conn = None
    try:
        conn = db_eng.raw_connection()
        with conn.cursor() as cur:
            for item in sn_array:
                if item[0] == '0098082c':
                    pass
                try:
                    # last 180 days, more than 1 kwh in an hour
                    sql = f"""
                    WITH wh AS (
                        SELECT 
                            DATE_TRUNC('hour', ts) AS ts_hour,
                            DATE(ts) AS dt,
                            EXTRACT(HOUR FROM ts) AS hr_utc,
                            ROUND(SUM(wh_p1)::numeric, 3) AS wh_p1, 
                            ROUND(SUM(wh_p2)::numeric, 3) AS wh_p2, 
                            ROUND(SUM(wh_p3)::numeric, 3) AS wh_p3
                        FROM eyedro.gb_{item[0]}
                        where ts > now() - interval '180 days'
                        GROUP BY 1, 2, 3
                    ),
                    hourly_avg AS (
                        SELECT 
                            hr_utc,
                            ROUND(AVG(wh_p1)::numeric, 3) AS avg_wh_p1,
                            ROUND(AVG(wh_p2)::numeric, 3) AS avg_wh_p2,
                            ROUND(AVG(wh_p3)::numeric, 3) AS avg_wh_p3,
                            ROUND(
                            ((GREATEST(AVG(wh_p1)::numeric, AVG(wh_p2)::numeric, AVG(wh_p3)::numeric) - 
                            LEAST(AVG(wh_p1)::numeric, AVG(wh_p2)::numeric, AVG(wh_p3)::numeric)) / 
                            NULLIF(GREATEST(AVG(wh_p1)::numeric, AVG(wh_p2)::numeric, AVG(wh_p3)::numeric), 0)), 3
                        ) AS phase_imbalance
                        FROM wh
                        GROUP BY hr_utc
                    )
                    select * from hourly_avg 
                    where avg_wh_p1 + avg_wh_p2 + avg_wh_p3 > 1000
                    order by hr_utc
                    """
                    #print(sql)
                    cur.execute(sql)
                    res = cur.fetchall()
                    conn.commit()

                    # Convert to DataFrame only if there is data
                    if res:
                        df_result = pd.DataFrame(res, columns=[desc[0] for desc in cur.description])
                        #print (df_result)
                        # Search for a specific serial
                        matching_row = df_gb_office[df_gb_office["serial"] == item[0].upper()]
                        print (matching_row)
                        prefix=f'sn-{item[0]}'
                        if matching_row.empty:
                            fn = f'{out_path}\\{item[0]}_phase_imbalance_{run_dt}.xlsx'
                            df_result.to_csv(fn_csv, index=False)  # ✅ Save to CSVcsv'
                        else:
                            prefix = sanitize_filename(f'{matching_row.values[0][4]}_{matching_row.values[0][5]}_')
                            fn = f'{out_path}\\{prefix}{item[0]}_phase_imbalance_{run_dt}.xlsx'
                        df_result["source"] = item[1]
                        df_result["phase_imbalance"] = pd.to_numeric(df_result["phase_imbalance"])
                        df_result["avg_wh_p1"] = pd.to_numeric(df_result["avg_wh_p1"])
                        df_result["avg_wh_p2"] = pd.to_numeric(df_result["avg_wh_p2"])
                        df_result["avg_wh_p3"] = pd.to_numeric(df_result["avg_wh_p3"])
                        df_result["hr_utc"] = pd.to_numeric(df_result["hr_utc"])

                        fn_csv = fn.replace('.xlsx', '.csv')
                        df_result.to_csv(fn_csv, index=False)
                        #!!!! 
                        ######continue   # ✅ Save to CSV

                        with pd.ExcelWriter(fn, engine="openpyxl") as writer:
                            df_result.to_excel(writer, sheet_name=prefix, index=False)  # ✅ Save to CSV
                        # Load workbook and sheet
                        wb = load_workbook(fn)
                        ws = wb[prefix]

                        # Apply style to "phase_imbalance" column (column E, starting from row 2)
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
                            for cell in row:
                                cell.number_format = "0.0%"

                        # Create Column Chart
                        chart = BarChart()
                        chart.type = "col"
                        chart.style = 10
                        chart.title = "Phase Imbalance"

                        # **Ensure axis titles are positioned correctly**
                        chart.y_axis.title = "Phase Imbalance (%)"
                        chart.x_axis.title = "Hour (UTC)"

                        # **Ensure axis labels (tick marks) are visible**
                        #chart.y_axis.majorGridlines = None  # Remove unnecessary gridlines
                        #chart.y_axis.minorGridlines = None  # Ensure no minor gridlines
                        chart.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart.x_axis.majorTickMark = "out"  # Show tick marks outward

                        # **Ensure X-axis title is positioned properly**
                        chart.x_axis.title.txPr = None  # Reset title formatting (places it outside)
                        chart.y_axis.title.txPr = None  

                        # **Ensure data is treated as a single series**
                        data = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
                        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

                        chart.add_data(data, titles_from_data=False)  # Single series, no multi-color
                        chart.set_categories(categories)  # Ensure hours appear on X-axis

                        # **Increase chart size to fit labels properly**
                        chart.width = 20  # Increase width (default is ~15)
                        chart.height = 10  # Increase height (default is ~7.5)

                        chart.layout = Layout(
                            manualLayout=ManualLayout(
                                x=0.999,  # Shift plot area slightly right
                                #y=1.05,  # Shift plot area slightly up
                                h=0.95,  # Reduce height of the plot area
                                w=0.95   # Reduce width of the plot area
                            )
                        )

                        # Enable auto-scaling (default behavior in openpyxl)
                        chart.y_axis.scaling.min = None  # Allow automatic minimum value
                        chart.y_axis.scaling.max = None  # Allow automatic maximum value
                        # !!!!!values = [row[0] for row in ws.iter_rows(min_col=5, min_row=2, max_row=ws.max_row, max_col=5, values_only=True)]
                        # max_value = max(values) if values else 1  # Avoids errors if empty
                        # chart.y_axis.scaling.max = max_value * 1.1  # Scale up by 10%

                        # Ensure Y-axis major tick marks are visible (auto-scaled)
                        chart.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart.y_axis.minorTickMark = "none"  # Hide minor tick marks

                        # Force display of axis values (ensures ticks are labeled)
                        chart.y_axis.majorGridlines = None  # Keep clean gridlines
                        chart.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks

                        # Extract Y-axis values correctly
                        # y_values = [row[0] for row in ws.iter_rows(min_col=5, max_col=5, min_row=2, max_row=ws.max_row, values_only=True) if row[0] is not None]
                        # # Prevent errors when max() is called on empty data
                        # max_y = max(y_values) if y_values else 1  
                        # min_y = min(y_values) if y_values else 0  # Ensure we get the lowest value  
                        # Apply auto-scaling while ensuring labels are displayed
                        # chart.y_axis.scaling.min = min_y  # Start from lowest value
                        # chart.y_axis.scaling.max = max_y * 1.1  # Scale up by 10%
                        # chart.y_axis.majorUnit = round((max_y - min_y) / 10, 2)
                        chart.y_axis.scaling.min = 0  # Start from lowest value
                        chart.y_axis.scaling.max = 1.05
                        chart.y_axis.majorUnit = round(1 / 10, 2)  # Ensure reasonable spacing

                        # Ensure tick labels are displayed properly
                        chart.y_axis.majorTickMark = "out"
                        chart.y_axis.minorTickMark = "none"
                        chart.y_axis.tickLblPos = "nextTo"  # Position labels properly

                        # Force Y-axis labels to display (Excel sometimes hides them)
                        chart.y_axis.delete = False  # Ensure axis is visible

                        chart.x_axis.majorTickMark = "out"  # Show tick marks outward
                        chart.x_axis.tickLblPos = "nextTo"  # Ensure labels are placed properly
                        chart.x_axis.delete = False  # Make sure X-axis is visible
                        chart.x_axis.majorUnit = 1  # Ensure every hour is displayed
                        chart.x_axis.minorUnit = 1  # Prevent minor ticks from interfering
                        # Get the first data series (bars)
                        series = chart.series[0]  

                        # Set a solid fill color (change "0000FF" to your desired color)
                        series.graphicalProperties.solidFill = "0000FF"  # Blue

                        # **Improve styling**
                        chart.gapWidth = 30  # Adjust spacing between bars
                        chart.overlap = 0  # Ensure no overlap
                        chart.legend = None  # Remove legend (single series)

                        # Create Line Chart
                        chart1 = LineChart()
                        chart1.type = "line"
                        chart1.style = 10
                        chart1.title = "Wh by Phase"

                        # **Ensure axis titles are positioned correctly**
                        chart1.y_axis.title = "Phase (Wh)"
                        chart1.x_axis.title = "Hour (UTC)"

                        # **Ensure axis labels (tick marks) are visible**
                        chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart1.x_axis.majorTickMark = "out"  # Show tick marks outward

                        # **Ensure X-axis title is positioned properly**
                        chart1.x_axis.title.txPr = None  # Reset title formatting (places it outside)
                        chart1.y_axis.title.txPr = None  

                        # **Ensure data is treated as a single series**
                        data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=ws.max_row)

                        ####categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

                        chart1.add_data(data, titles_from_data=True)  # Single series, no multi-color
                        ####chart1.set_categories(categories)  # Ensure hours appear on X-axis

                        # **Increase chart size to fit labels properly**
                        chart1.width = 20  # Increase width (default is ~15)
                        chart1.height = 10  # Increase height (default is ~7.5)

                        chart1.layout = Layout(
                            manualLayout=ManualLayout(
                                x=0.999,  # Shift plot area slightly right
                                #y=1.05,  # Shift plot area slightly up
                                h=0.95,  # Reduce height of the plot area
                                w=0.95   # Reduce width of the plot area
                            )
                        )

                        # Enable auto-scaling (default behavior in openpyxl)
                        chart1.y_axis.scaling.min = None  # Allow automatic minimum value
                        chart1.y_axis.scaling.max = None  # Allow automatic maximum value
                        # !!!!!!! values = [row[0] for row in ws.iter_rows(min_col=5, min_row=2, max_row=ws.max_row, max_col=5, values_only=True)]
                        # max_value = max(values) if values else 1  # Avoids errors if empty
                        # chart1.y_axis.scaling.max = max_value * 1.1  # Scale up by 10%

                        # Ensure Y-axis major tick marks are visible (auto-scaled)
                        chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart1.y_axis.minorTickMark = "none"  # Hide minor tick marks

                        # Force display of axis values (ensures ticks are labeled)
                        chart1.y_axis.majorGridlines = None  # Keep clean gridlines
                        chart1.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks

                        # Extract Y-axis values correctly
                        # Extract all values from columns 2, 3, and 4
                        y_values = [value for row in ws.iter_rows(min_col=2, max_col=4, min_row=2, max_row=ws.max_row, values_only=True) 
                            for value in row if value is not None]
                        # Prevent errors when max() is called on empty data
                        max_y = max(y_values) if y_values else 1
                        min_y = min(y_values) if y_values else 0  # Ensure we get the lowest value
                        #Apply auto-scaling while ensuring labels are displayed
                        chart1.y_axis.scaling.min = min_y  # Start from lowest value
                        chart1.y_axis.scaling.max = max_y * 1.1  # Scale up by 10%
                        chart1.y_axis.majorUnit = round((max_y - min_y) / 10, 2)
                        # chart1.y_axis.scaling.min = 0  # Start from lowest value
                        # chart1.y_axis.scaling.max = 1.05
                        # chart1.y_axis.majorUnit = round(1 / 10, 2)  # Ensure reasonable spacing

                        # Ensure tick labels are displayed properly
                        chart1.y_axis.majorTickMark = "out"
                        chart1.y_axis.minorTickMark = "none"
                        chart1.y_axis.tickLblPos = "nextTo"  # Position labels properly

                        # Force Y-axis labels to display (Excel sometimes hides them)
                        chart1.y_axis.delete = False  # Ensure axis is visible

                        chart1.x_axis.majorTickMark = "out"  # Show tick marks outward
                        chart1.x_axis.tickLblPos = "nextTo"  # Ensure labels are placed properly
                        chart1.x_axis.delete = False  # Make sure X-axis is visible
                        chart1.x_axis.majorUnit = 1  # Ensure every hour is displayed
                        chart1.x_axis.minorUnit = 1  # Prevent minor ticks from interfering

                        chart_sheet_name = "Charts"
                        ws_chart = wb.create_sheet(chart_sheet_name)
                        # **Add chart to Excel**
                        ws_chart.add_chart(chart, "C2")
                        ws_chart.add_chart(chart1, "C23")

                        wb.save(fn)

                        logger.info(f"Saved CSV & XLSX for {item}: {fn}")
                    else:
                        logger.info(f"No data for {item}, skipping CSV creation.")
                        conn.commit()

                except DatabaseError as e:
                    logger.error(f"Database error for {item}: {e}")
                    conn.commit()
                except Exception as e:
                    logger.error(f"Unexpected error for {item}: {e}")
    except DatabaseError as e:
        logger.error(f"Database connection error: {e}")
        conn.commit()
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
    finally:
        if conn:
            conn.close()  # ✅ Always close the connection


run_dt = datetime.now(timezone.utc).date()
#ILTERED_GB_SN_PATH=const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, (run_dt - timedelta(days=1)).isoformat()) #const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt.isoformat())
FILTERED_GB_SN_PATH=const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, (run_dt).isoformat()) #const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt.isoformat())

if os.path.exists(FILTERED_GB_SN_PATH):
    filtered_gb_sn_df = pd.read_csv(FILTERED_GB_SN_PATH)
else:
    df, err = gb_eyedro.api_get_user_info_as_df()
    if err:
        logger.error(f"api_get_user_info_as_df ERROR: {err}")
        exit(1)

    filtered_gb_sn_df = df[df['gb_serial'].str.startswith('B') | df['gb_serial'].str.startswith('009')]['gb_serial'].drop_duplicates()

sn_array = sorted(filtered_gb_sn_df.str.replace('-', '').tolist())

eng = db.set_local_defaultdb_engine()
sql = "select serial_number, brand from public.gtb_gb_unifier where brand ilike '%gen%' and brand not ilike '%grid%'"
res, err = db.sql_execute(sql, eng)
if err:
    logger.error(f"SQL execution error: {err}")
    exit(1)
sn_array = [[row[0].replace('-', ''), row[1]] for row in res]
sn_array = sorted(sn_array)


out_path = r'E:\_UNHCR\CODE\DATA\archive' 
# !!!!!!! phase_imbalance
#phase_imbalance(eng, sn_array[:500], out_path)
# pass

# -----------------------------
# 🧪 Usage Example:
in_path = out_path #######r'E:\_UNHCR\CODE\DATA\phase\input/'
classify_imbalance('v2', in_path, normalize=False, top_n=25, parallel=False, generate_images=False)
# # -----------------------------

pass

