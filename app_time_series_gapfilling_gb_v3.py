from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from functools import partial
from fedot.core.pipelines.node import PipelineNode
from fedot.core.pipelines.pipeline import Pipeline
from fedot.utilities.ts_gapfilling import ModelGapFiller
import gc
import logging
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart,LineChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
import os
import pandas as pd
from sklearn.metrics import mean_absolute_error, mean_squared_error, median_absolute_error
import time

from unhcr import utils
from unhcr import constants as const
from unhcr import app_utils

run_dt = datetime.now().date()
FILTERED_GB_SN_PATH=const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt.isoformat())
filtered_gb_sn_df = pd.read_csv(FILTERED_GB_SN_PATH)
filtered_gb_sn_df['gb_serial'] = filtered_gb_sn_df['gb_serial'].str.replace('-', '', regex=True)

arr = filtered_gb_sn_df[['gb_serial','epoch_utc']].copy()
arr['epoch_utc'] = pd.to_numeric(arr['epoch_utc'], errors='coerce').fillna(-1).astype(np.int32)

arr = arr.values.tolist()

pass

# OPTIONAL: set your own environment
##ef = const.load_env(r'E:\_UNHCR\CODE\unhcr_module\.env')
## print(ef)
# OPTIONAL: set your own environment
from unhcr import db

mods = [
    ["app_utils", "app_utils"],
    ["constants", "const"],
    ["db", "db"],
]

res = app_utils.app_init(mods, "unhcr.update_all.log", "0.4.7", level="INFO", override=True)
if const.LOCAL:
    app_utils, const, db = res

engines = db.set_db_engines()

data_dir = const.GB_GAPS_DATA_DIR

# Turn on interactive plotting mode
plt.ion()


# Method 1: Using standard deviation
def replace_outliers_std(df, column='wh', threshold=3):
    """Replace outliers based on standard deviation with the mean"""
    mean = df[column].mean()
    std = df[column].std()
    
    # Create a copy of the dataframe
    result = df.copy()
    
    # Identify outliers
    outliers = (result[column] > mean + threshold * std) | (result[column] < mean - threshold * std)
    
    # Replace outliers with the mean
    result.loc[outliers, column] = mean
    
    return result

# Method 2: Using IQR (Interquartile Range)
def replace_outliers_iqr(df, column='wh', threshold=1.5):
    """Replace outliers based on IQR with the median"""
    q1 = df[column].quantile(0.25)
    q3 = df[column].quantile(0.75)
    iqr = q3 - q1
    
    # Create a copy of the dataframe
    result = df.copy()
    
    # Identify outliers
    lower_bound = q1 - threshold * iqr
    upper_bound = q3 + threshold * iqr
    outliers = (result[column] < lower_bound) | (result[column] > upper_bound)
    
    # Replace outliers with the median
    result.loc[outliers, column] = df[column].median()
    
    return result

# Method 3: Using rolling statistics (best for time series)
def replace_outliers_rolling(df, column='wh', window=10, threshold=3):
    """Replace outliers based on rolling statistics with the local rolling median"""
    # Calculate rolling median and standard deviation
    rolling_median = df[column].rolling(window=window, center=True).median()
    rolling_std = df[column].rolling(window=window, center=True).std()
    
    # Handle NaN values from rolling calculations
    rolling_median = rolling_median.fillna(df[column].median())
    rolling_std = rolling_std.fillna(df[column].std())
    
    # Create a copy of the dataframe
    result = df.copy()
    
    # Identify outliers
    lower_bound = rolling_median - threshold * rolling_std
    upper_bound = rolling_median + threshold * rolling_std
    outliers = (result[column] < lower_bound) | (result[column] > upper_bound)
    
    # Replace outliers with the corresponding rolling median
    result.loc[outliers, column] = rolling_median[outliers]
    
    return result

# Method 4: Using interpolation (maintains time series continuity)
def replace_outliers_interpolate(df, column='wh', window=10, threshold=3):
    """Replace outliers with interpolated values"""
    # Calculate rolling median and standard deviation
    rolling_median = df[column].rolling(window=window, center=True).median()
    rolling_std = df[column].rolling(window=window, center=True).std()
    
    # Handle NaN values from rolling calculations
    rolling_median = rolling_median.fillna(df[column].median())
    rolling_std = rolling_std.fillna(df[column].std())
    
    # Create a copy of the dataframe
    result = df.copy()
    
    # Identify outliers
    lower_bound = rolling_median - threshold * rolling_std
    upper_bound = rolling_median + threshold * rolling_std
    outliers = (result[column] < lower_bound) | (result[column] > upper_bound)
    
    # Set outliers to NaN
    result.loc[outliers, column] = np.nan
    
    # Interpolate the NaN values (using 'linear' instead of 'time')
    result[column] = result[column].interpolate(method='linear')
    
    # If any NaN values remain at the edges, fill them with nearby values
    result[column] = result[column].fillna(method='ffill').fillna(method='bfill')
    
    return result

# Example usage:
# df_cleaned = replace_outliers_interpolate(df, column='wh', window=24, threshold=3)

def print_gap_metrics(actual, ridge_predicted, composite_predicted, gap_number):
    """
    Print metrics for a specific gap
    
    :param actual: Actual values in the gap
    :param ridge_predicted: Ridge model predictions
    :param composite_predicted: Composite model predictions
    :param gap_number: Index of the current gap
    """
    print(f"\nGap {gap_number} Metrics:")
    
    metrics = []
    ridge = None
    composite = None
    model_labels = ['Inverted ridge regression', 'Composite model']
    for predicted, model_label in zip(
            [ridge_predicted, composite_predicted], model_labels):
        print(f"{model_label}")
        metrics.append(f"{model_label}"),

        mae_metric = mean_absolute_error(actual, predicted)
        print(f"Mean absolute error - {mae_metric:.2f}")
        metrics.append(f"Mean absolute error - {mae_metric:.2f}")


        rmse_metric = (mean_squared_error(actual, predicted)) ** 0.5
        print(f"Root mean squared error - {rmse_metric:.2f}")
        metrics.append(f"Root mean squared error - {rmse_metric:.2f}")

        median_ae_metric = median_absolute_error(actual, predicted)
        print(f"Median absolute error - {median_ae_metric:.2f}")
        metrics.append(f"Median absolute error - {median_ae_metric:.2f}")
        if ridge is None:
            ridge = median_ae_metric + rmse_metric + mae_metric
        else:
            composite = median_ae_metric + rmse_metric + mae_metric
    return metrics, ridge, composite


def plot_gap_result(subset, gap_indices, gap_number):
    """
    Plot results for a specific gap interactively
    
    :param subset: DataFrame containing the gap and surrounding data
    :param gap_indices: Indices of the gap within the subset
    :param gap_number: Index of the current gap
    """
    # Create masked array for visualizing the gap
    gap_mask = np.zeros(len(subset), dtype=bool)
    gap_mask[gap_indices] = True
    
    # Create a masked array for the gap
    masked_wh = np.ma.masked_array(subset['wh'].values, mask=~gap_mask)
    
    # Calculate the gap duration for the title
    gap_start = subset.iloc[gap_indices[0]]['date']
    gap_end = subset.iloc[gap_indices[-1]]['date']
    duration_minutes = len(gap_indices)
    
    # Create plot with lower DPI to reduce memory usage
    plt.figure(figsize=(10, 6), dpi=80)
    
    # Plot actual values
    plt.plot(subset['date'], subset['wh'], c='blue',
             alpha=0.5, label='Actual values', linewidth=1)
    
    #Plot the gap filled values
    plt.plot(subset['date'], subset['ridge'], c='orange',
             alpha=0.8, label='Ridge gap-filling', linewidth=1)
    
    plt.plot(subset['date'], subset['composite'], c='red',
             alpha=0.8, label='Composite gap-filling', linewidth=1)
    
    # Plot the gap itself more prominently
    plt.plot(subset.iloc[gap_indices]['date'], masked_wh[gap_indices], 'ko', 
             markersize=3, label='Gap locations')
    
    # Add vertical lines to indicate gap boundaries
    plt.axvline(x=gap_start, color='gray', linestyle='--', alpha=0.7)
    plt.axvline(x=gap_end, color='gray', linestyle='--', alpha=0.7)
    
    plt.title(f'Gap {gap_number}: {duration_minutes} minutes ({gap_start} to {gap_end})')
    plt.grid(True, alpha=0.3)
    plt.legend()
    plt.tight_layout()
    
    # Show the plot interactively instead of saving
    plt.draw()
    plt.pause(0.1)  # Small pause to ensure the plot is drawn
    print(f"Displayed interactive plot for gap {gap_number}")
    # Add a small delay so the plot is visible
    time.sleep(0.5)  # Add a short delay to give time to view the plot


def get_composite_pipeline(window_size):
    """
    The function returns prepared pipeline of 5 models

    :return: Pipeline object
    """
    # Create node with a 150-minute window size
    node_1 = PipelineNode('lagged')
    node_1.parameters = {'window_size': window_size}
    
    # Create node with a 100-minute window size
    node_2 = PipelineNode('lagged')
    node_2.parameters = {'window_size': window_size}
    
    # Linear models from lagged features
    node_linear_1 = PipelineNode('linear', nodes_from=[node_1])
    node_linear_2 = PipelineNode('linear', nodes_from=[node_2])

    # Final ridge regression combining both linear models
    node_final = PipelineNode('ridge', nodes_from=[node_linear_1,
                                                  node_linear_2])
    return Pipeline(node_final)


def get_simple_pipeline(window_size):
    """ 
    Function returns simple pipeline with lagged features and ridge regression
    """
    # Create node with a 200-minute window size
    node_lagged = PipelineNode('lagged')
    node_lagged.parameters = {'window_size': window_size}
    
    # Ridge regression from lagged features
    node_ridge = PipelineNode('ridge', nodes_from=[node_lagged])
    return Pipeline(node_ridge)


def add_chart_to_excel(file_name, metrics, ridge, comp, gap_size, window_size, ChartType=LineChart):
    # Load workbook and select sheet (after saving data)
    wb = load_workbook(file_name)
    ws = wb["Data"]

    row = 2
    for met in metrics:
        ws[f'W{row}'] = met
        row += 1
    row += 2
    ws[f'W{row}'] = f'Ridge: {ridge}'
    row += 1
    ws[f'W{row}'] = f'Composite: {comp}'
    row += 2
    ws[f'W{row}'] = f'Gap size: {gap_size}'
    row += 1
    ws[f'W{row}'] = f'Window size: {window_size}'
    row += 1

    # Set the style for date formatting and make it a str
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.value = cell.value.strftime("%Y-%m-%d %H:%M")

    # Create Line Chart
    chart1 = ChartType()
    chart1.type = "line"
    chart1.style = 10
    chart1.title = "Gap filling"

    # Axis titles
    chart1.y_axis.title = "-100 = Gap Filled (with_gap)"
    chart1.x_axis.title = "Date time"

    # **Ensure axis labels (tick marks) are visible**
    chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart1.x_axis.majorTickMark = "out"  # Show tick marks outward

    # **Ensure X-axis labels are positioned at the bottom**
    chart1.x_axis.tickLblPos = "low"  # Position X-axis labels at the bottom

    # **Ensure axis titles are positioned correctly**
    chart1.y_axis.title.txPr = None  

    # **Set Y-axis data range** (assuming the remaining column is the Y-values after removing columns B and E)
    col = 4
    if ridge - comp > 0:
        col = 5
    data = Reference(ws, min_col=col, min_row=1, max_col=col, max_row=ws.max_row)  # Y-values in column B
    chart1.add_data(data, titles_from_data=True)

  # # **Set X-axis to the date values in column 1 (starting from row 2 to the last row)**
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Dates in column A
    chart1.x_axis.majorGridlines = None  # Remove gridlines for clarity
    chart1.set_categories(categories)  # Use set_categories if it works

    # **Increase chart size to fit labels properly**
    chart1.width = 25  # Increase width (default is ~15)
    chart1.height = 15  # Increase height (default is ~7.5)

    chart1.layout = Layout(
          manualLayout=ManualLayout(
            h= 0.9,   # Reduce height of the plot area
            w= 0.95,    # Reduce width of the plot area
            y= 0.01,
            x= 0.01
        )
    )

    # Y-axis settings (auto-scaling, tick marks, etc.)
    y_values = [value for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=ws.max_row, values_only=True) 
                for value in row if value is not None]
    max_y = max(y_values) if y_values else 1
    min_y = min(y_values) if y_values else 0
    chart1.y_axis.scaling.min = min_y - 1
    chart1.y_axis.scaling.max = max_y * 1.2
    chart1.y_axis.majorUnit = round((max_y - min_y) / 10, 2)

    # Ensure Y-axis labels are displayed properly
    chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart1.y_axis.minorTickMark = "none"  # Hide minor tick marks
    chart1.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks
    chart1.y_axis.delete = False  # Ensure axis is visible

    # Ensure X-axis tick marks and labels are displayed
    chart1.x_axis.majorTickMark = "out"  # Show tick marks outward
    chart1.x_axis.tickLblPos = "low"  # Ensure labels are at the bottom
    chart1.x_axis.delete = False  # Make sure X-axis is visible
    chart1.x_axis.majorUnit = 1  # Display one tick per hour
    chart1.x_axis.minorUnit = 1  # Prevent minor ticks from interfering

    chart1.x_axis.txPr = chart1.x_axis.title.text.rich
    chart1.x_axis.txPr.properties.rot = "-2700000"
    chart1.x_axis.title = None

    # Create Line Chart for 2cnd y axis
    chart2 = ChartType()
    chart2.type = "line"
    chart2.style = 10
    chart2.title = "Gap filling"

    # Axis titles
    chart2.y_axis.title = "Wh (ridge | comp)"
    chart2.x_axis.title = "Date time"

    # **Ensure axis labels (tick marks) are visible**
    chart2.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart2.x_axis.majorTickMark = "out"  # Show tick marks outward

    # **Ensure X-axis labels are positioned at the bottom**
    chart2.x_axis.tickLblPos = "low"  # Position X-axis labels at the bottom

    # **Ensure axis titles are positioned correctly**
    #####chart2.x_axis.title.txPr = None  # Reset title formatting (places it outside)
    chart2.y_axis.title.txPr = None  

    # **Set Y-axis data range** (assuming the remaining column is the Y-values after removing columns B and E)
    data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)  # Y-values in column B
    chart2.add_data(data, titles_from_data=True)

  # # **Set X-axis to the date values in column 1 (starting from row 2 to the last row)**
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Dates in column A
    chart2.x_axis.majorGridlines = None  # Remove gridlines for clarity
    chart2.y_axis.majorGridlines = None 
    chart2.set_categories(categories)  # Use set_categories if it works

    # **Increase chart size to fit labels properly**
    chart2.width = 25  # Increase width (default is ~15)
    chart2.height = 15  # Increase height (default is ~7.5)

    chart2.layout = Layout(
        manualLayout=ManualLayout(
            h= 0.9,   # Reduce height of the plot area
            w= 0.95,    # Reduce width of the plot area
            y= 0.01,
            x= 0.01
        )
    )

    # Y-axis settings (auto-scaling, tick marks, etc.)
    y_values = [value for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, max_row=ws.max_row, values_only=True) 
                for value in row if value is not None]
    max_y = max(y_values) if y_values else 1
    min_y = min(y_values) if y_values else 0
    chart2.y_axis.scaling.min = min_y - 1
    chart2.y_axis.scaling.max = max_y * 1.2
    chart2.y_axis.majorUnit = round((max_y - min_y) / 10, 2)

    # Ensure Y-axis labels are displayed properly
    chart2.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart2.y_axis.minorTickMark = "none"  # Hide minor tick marks
    chart2.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks
    chart2.y_axis.delete = False  # Ensure axis is visible

    # Ensure X-axis tick marks and labels are displayed
    chart2.x_axis.majorTickMark = "out"  # Show tick marks outward
    chart2.x_axis.tickLblPos = "low"  # Ensure labels are at the bottom
    chart2.x_axis.delete = False  # Make sure X-axis is visible
    chart2.x_axis.majorUnit = 1  # Display one tick per hour
    chart2.x_axis.minorUnit = 1  # Prevent minor ticks from interfering

    chart2.x_axis.txPr = chart2.x_axis.title.text.rich
    chart2.x_axis.txPr.properties.rot = "-2700000"
    chart2.x_axis.title = 'XXXXX'
    chart2.y_axis.axId = 200
    chart1.legend.position = 'b'

        
        
    chart1.y_axis.crosses = "max"
    chart1 += chart2

    ws.add_chart(chart1, "G2")
    
    
    
    chart3 = ChartType()
    chart3.type = "line"
    chart3.style = 10
    chart3.title = "Compare Gap Filling"

    # Axis titles
    chart3.y_axis.title = "Ridge vs Composite"
    chart3.x_axis.title = "Date time"

    # Ensure axis labels (tick marks) are visible
    chart3.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart3.x_axis.majorTickMark = "out"  # Show tick marks outward

    # Ensure X-axis labels are positioned at the bottom
    chart3.x_axis.tickLblPos = "low"  # Position X-axis labels at the bottom

    # Ensure axis titles are positioned correctly
    chart3.y_axis.title.txPr = None
    
    # Increase chart size to fit labels properly
    chart3.width = 25  # Increase width (default is ~15)
    chart3.height = 15  # Increase height (default is ~7.5)

    #('layoutTarget', 'xMode', 'yMode', 'wMode', 'hMode', 'x', 'y', 'w', 'h')
    chart3.layout = Layout(
        manualLayout=ManualLayout(
            h= 0.9,   # Reduce height of the plot area
            w= 0.95,    # Reduce width of the plot area
            y= 0.01,
            x= 0.01
        )
    )
    
       # Y-axis settings (auto-scaling, tick marks, etc.)
    y_values = [value for row in ws.iter_rows(min_col=6, max_col=6, min_row=2, max_row=ws.max_row, values_only=True) 
                for value in row if value is not None]
    max_y = max(y_values) if y_values else 1
    min_y = min(y_values) if y_values else 0
    chart3.y_axis.scaling.min = min_y - 1
    chart3.y_axis.scaling.max = max_y * 1.2
    chart3.y_axis.majorUnit = round((max_y - min_y) / 10, 2)

    # Ensure Y-axis labels are displayed properly
    chart3.y_axis.majorTickMark = "out"  # Show tick marks outward
    chart3.y_axis.minorTickMark = "none"  # Hide minor tick marks
    chart3.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks
    chart3.y_axis.delete = False
    chart3.legend = None

    # Ensure X-axis tick marks and labels are displayed
    chart3.x_axis.majorTickMark = "out"  # Show tick marks outward
    chart3.x_axis.tickLblPos = "low"  # Ensure labels are at the bottom
    chart3.x_axis.delete = False  # Make sure X-axis is visible
    chart3.x_axis.majorUnit = 1  # Display one tick per hour
    chart3.x_axis.minorUnit = 1  # Prevent minor ticks from interfering

    chart3.x_axis.txPr = chart3.x_axis.title.text.rich
    chart3.x_axis.txPr.properties.rot = "-2700000"
    chart3.x_axis.title = None
    #!!!!!!!!!!

    data = Reference(ws, min_col=6, min_row=2, max_col=6, max_row=ws.max_row)
    chart3.add_data(data, titles_from_data=True)

#  # Set Y-axis data range (use column F for the difference between Ridge and Composite)
#     data = Reference(ws, min_col=6, min_row=2, max_col=6, max_row=ws.max_row)  # Y-values in column F
#     chart3.add_data(data, titles_from_data=False)

    # Set X-axis to the date values in column A (starting from row 2 to the last row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Dates in column A
    chart3.x_axis.majorGridlines = None  # Remove gridlines for clarity
    chart3.set_categories(categories) 
    # Style the lines
    # s1 = c1.series[0]
    # s1.marker.symbol = "triangle"
    # s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
    # s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

    # s1.graphicalProperties.line.noFill = True

    s2 = chart3.series[0]
    s2.graphicalProperties.line.solidFill = "00AAAA"
    ####s2.graphicalProperties.line.dashStyle = "sysDot"
    s2.graphicalProperties.line.width = 50050 # width in EMUs

    #s2 = c1.series[2]
    s2.smooth = True # Make the line smooth

    ws.add_chart(chart3, "G32")

    
    wb.save(file_name)
    wb.close()


def extend_dataset_with_medians(df, time_column, value_column, n_periods=500):
    """
    Extends a time series dataset by adding median values at the beginning and end.
    
    Parameters:
    -----------
    df : pandas.DataFrame
        The original time series data
    time_column : str
        Name of the column containing datetime values
    value_column : str
        Name of the column containing the values to be gap-filled
    n_periods : int
        Number of periods to add at each end (default: 500)
    
    Returns:
    --------
    pandas.DataFrame
        Extended dataset with median values at beginning and end
    """
    # Calculate the median value
    median_value = df[value_column].median()
    
    # Find time frequency (assuming regular intervals)
    #time_freq = df[time_column].diff().median()
    time_freq = pd.Timedelta(minutes=1)
    
    # Create new dates before the first date
    first_date = df[time_column].min()
    dates_before = pd.date_range(end=first_date - time_freq, periods=n_periods, freq=time_freq)
    
    # Create new dates after the last date
    last_date = df[time_column].max()
    dates_after = pd.date_range(start=last_date + time_freq, periods=n_periods, freq=time_freq)
    
    gap_col = 'with_gap'
    # Create DataFrames for the extensions
    start_df = pd.DataFrame({
        time_column: dates_before,
        value_column: median_value,
        gap_col: median_value,
    })
    
    end_df = pd.DataFrame({
        time_column: dates_after,
        value_column: median_value,
        gap_col: median_value,
    })
    
    # Combine everything
    extended_df = pd.concat([start_df, df, end_df], ignore_index=True)
    
    return extended_df


def run_gapfilling_by_segments(file_path, plot_individual_gaps=True, min_gap_size=10, df=None):
    # Load the data
    """
    Main entry point to fill gaps in a time series using gap filling algorithms.

    Parameters
    ----------
    file_path : str
        The path to the CSV file containing the time series to fill.
    plot_individual_gaps : bool, optional
        Whether to plot each gap individually with context around it.
    min_gap_size : int, optional
        The minimum gap size in minutes to plot individual gaps.
    df : pandas.DataFrame, optional
        The DataFrame containing the time series to fill. If None, the CSV file
        will be read using pandas.

    Returns
    -------
    pandas.DataFrame
        The DataFrame with the filled gaps.

    Notes
    -----
    The function will write the filled gaps to a new Excel file in the same
    directory as the input CSV file. The file name will be the same as the input
    CSV file with "_all_gaps" appended to the end. If the total gap size filled
    by the composite model is greater than the total gap size filled by the
    ridge model, the file name will be modified to include "_comp" instead of
    "_ridge". The file will contain a single sheet with the filled gaps, and a
    line chart will be added to the sheet showing the difference between the
    two models.

    The function will also print the metrics for each gap, including the mean
    absolute error (MAE), the mean squared error (MSE), the root mean squared
    percentage error (RMSPE), and the R-squared value. The function will also
    print the total gap size filled by each model, and the difference between
    the two models.
    """
    print("Loading data...")
    padding = 500
    # Read the data with efficient dtypes
    if df is not None:
        dataframe = df.copy()
    else:
        dataframe = pd.read_csv(file_path, dtype={'wh': 'float32', 'with_gap': 'float32'})

    dataframe = extend_dataset_with_medians(dataframe, 'date', 'wh', n_periods=padding)

    dataframe['date'] = pd.to_datetime(dataframe['date'], infer_datetime_format=True)

    print("Preprocessing time series...")
    # Convert date to index and fill missing timestamps
    dataframe = dataframe.set_index("date")
    full_range = pd.date_range(start=dataframe.index.min(), end=dataframe.index.max(), freq="T")
    dataframe = dataframe.reindex(full_range)

    # Process missing values
    missing_mask = dataframe["wh"].isna()
    dataframe["wh"] = dataframe["wh"].fillna(0)
    dataframe.loc[missing_mask, "with_gap"] = -100.0

    # Reset index
    dataframe = dataframe.reset_index().rename(columns={"index": "date"})

    print("Identifying gaps in the time series...")
    # Find all gaps (continuous sequences of -100.0 values)
    is_gap = dataframe['with_gap'] == -100.0
    gap_indices = np.where(is_gap)[0]
    
    if len(gap_indices) == 0:
        print("No gaps found in the data.")
        return None, 'N/A'  # No gaps to fill
    
    print(f"Found {len(gap_indices)} missing values across multiple gaps.")
    
    # Initialize result columns with actual values
    dataframe['ridge'] = dataframe['wh'].copy()
    dataframe['composite'] = dataframe['wh'].copy()
    
    # Group continuous gap indices to identify separate gaps
    print("Grouping continuous gaps...")
    gap_groups = []
    current_group = [gap_indices[0]]
    
    for i in range(1, len(gap_indices)):
        if gap_indices[i] == gap_indices[i-1] + 1:
            # This gap is continuous with the previous one
            current_group.append(gap_indices[i])
        else:
            # This is the start of a new gap
            gap_groups.append(current_group)
            current_group = [gap_indices[i]]
    
    # Add the last group if it exists
    if current_group:
        gap_groups.append(current_group)
    
    print(f"Identified {len(gap_groups)} separate gaps to process.")
    
    
    metricx = []
    df_xcel = pd.DataFrame()
    ttl_comp = 0
    ttl_ridge = 0
    max_gap = 7200
    min_gap = 9999999999999999
    # Process each gap separately to minimize memory usage
    for i, gap_group in enumerate(gap_groups):
        window_size = 0
        gap_size = len(gap_group)
        if gap_size < min_gap:
            min_gap = gap_size
        if gap_size > max_gap:  # 5 day
            print(f"Skipping gap {i+1} of {len(gap_groups)} (size: {gap_size} minutes): too long (>= 1 day)")
            start_idx = gap_group[0]
            end_idx = gap_group[-1]
            # Extract the gap
            subset = dataframe[start_idx:end_idx]
            # Replace Infs with NaNs
            subset.replace([np.inf, -np.inf], np.nan, inplace=True)  # Convert Inf to NaN
            md = dataframe['wh'].median()
            subset.fillna(md, inplace=True)  # Fill NaNs after replacing Infs
            subset['wh'] = md
            subset['ridge'] = 0.0 #dataframe['wh'].median()
            subset['composite'] = 0.0 #dataframe['wh'].median()
            buffer_size = 0
            before_after = 0
            # if df_xcel.empty:
            #     df_xcel = subset.copy()
            # else:
            #     df_xcel = pd.concat([df_xcel, subset], ignore_index=True)
            # df_xcel.at[df_xcel.index[-1], "with_gap"] = -150.0 
            # window_size = 0
            # # Copy filled values back to main dataframe
            # # We only copy the actual gap values that were filled, not the entire subset
            # #dataframe.loc[original_positions, 'wh'] = subset.iloc[subset_gap_indices]['ridge'].values
            # continue
        else:
            #####window_size = 300 #gap_size * 3 if gap_size * 3 < 5000 else 5000
            # Dynamically adjust window size based on gap size
            window_size = min(gap_size * 3, 500) if gap_size > 60 else gap_size
            before_after = 1000
            metricx.append(f" GAP {i+1} of {len(gap_groups)} (size: {gap_size} minutes)    {gap_size}")
            print(f"{table} Processing gap {i+1} of {len(gap_groups)} (size: {gap_size} minutes)... {window_size}")
            # Maximum window size used in any model
            max_window_size = window_size
            buffer_size = max_window_size * 5  # Extra buffer for safety

            # Calculate the start and end indices with buffer for window size
            start_idx = max(0, gap_group[0] - buffer_size)
            end_idx = min(len(dataframe) - 1, gap_group[-1] + buffer_size)

            # Extract the subset of data around the gap
            subset = dataframe[start_idx:end_idx]

            # Replace Infs with NaNs
            subset.replace([np.inf, -np.inf], np.nan, inplace=True)  # Convert Inf to NaN
            subset.fillna(subset.median(), inplace=True)  # Fill NaNs after replacing Infs
            ##subset['with_gap'] = -100

            # Apply gap filling algorithms to just this subset
            gap_arr = subset['with_gap'].values
            median = np.nanmedian(gap_arr)
            # Create a copy of the array
            gap_array = np.copy(gap_arr)
            # Replace NaN values with median
            np.nan_to_num(gap_array, nan=median, copy=False)

            print(f"  Applying ridge regression model to gap {i+1}...")
            # Ridge pipeline - simpler model
            ridge_pipeline = get_simple_pipeline(window_size)
            ridge_gapfiller = ModelGapFiller(gap_value=-100.0, pipeline=ridge_pipeline)
            subset['ridge'] = ridge_gapfiller.forward_inverse_filling(gap_array)

            # Clean up to free memory
            del ridge_pipeline, ridge_gapfiller
            gc.collect()

            print(f"  Applying composite model to gap {i+1}...")
            #Composite pipeline - more complex model
            composite_pipeline = get_composite_pipeline(window_size)
            composite_gapfiller = ModelGapFiller(gap_value=-100.0, pipeline=composite_pipeline)
            #  #!!!! XXXXXonly use ridge for now
            # subset['composite'] = subset['ridge']
            subset['composite'] = composite_gapfiller.forward_inverse_filling(gap_array) ####.forward_filling(gap_array)  #!!! .forward_inverse_filling(gap_array)
            #Clean up to free memory
            del composite_pipeline, composite_gapfiller
            gc.collect()

            # Calculate the max value of the 'wh' column
            max_wh = subset['wh'].max()

            # Apply random clipping between 90% and 100% of the max value of 'wh' for each row
            np.random.seed(42)  # For reproducibility
            random_percentages = np.random.uniform(0.9, 1.0, size=len(subset))  # Random percentages between 90% and 100%

            # Apply the clipping using .iloc on the index
            subset['composite'] = [
                min(row.composite, random_percentages[i] * max_wh)
                for i, row in enumerate(subset.itertuples())
            ]
            
            subset['ridge'] = [
                min(row.ridge, random_percentages[i] * max_wh)
                for i, row in enumerate(subset.itertuples())
            ]
        
        # Identify the gap positions within the subset
        subset_gap_mask = subset['with_gap'] == -100.0
        subset_gap_indices = np.where(subset_gap_mask)[0]
        
        # Map subset positions back to original dataframe positions
        original_positions = subset_gap_indices + start_idx
        

        # After gap filling, replace negative values with small random positive numbers
        min_value = 0.1  # Lower bound of random range
        max_value = 2.0   # Upper bound of random range

        # Generate random replacements for ridge predictions
        negative_ridge_mask = subset['ridge'] < 0
        if negative_ridge_mask.any():
            random_values = np.random.uniform(min_value, max_value, size=negative_ridge_mask.sum())
            subset.loc[negative_ridge_mask, 'ridge'] = random_values

        # Generate random replacements for composite predictions
        negative_composite_mask = subset['composite'] < 0
        if negative_composite_mask.any():
            random_values = np.random.uniform(min_value, max_value, size=negative_composite_mask.sum())
            subset.loc[negative_composite_mask, 'composite'] = random_values
        
        if df_xcel.empty:
            df_xcel = subset.copy()
        else:
            df_xcel = pd.concat([df_xcel, subset], ignore_index=True)
        df_xcel.at[df_xcel.index[-1], "with_gap"] = -200.0 
        # Copy filled values back to main dataframe
        # We only copy the actual gap values that were filled, not the entire subset
        dataframe.loc[original_positions, 'ridge'] = subset.iloc[subset_gap_indices]['ridge'].values
        dataframe.loc[original_positions, 'composite'] = subset.iloc[subset_gap_indices]['composite'].values
        dataframe.loc[original_positions, 'wh'] = subset.iloc[subset_gap_indices]['wh'].values
        
        # Calculate and print metrics for this specific gap
        if len(subset_gap_indices) > 0:
            actual = subset.iloc[subset_gap_indices]['wh'].values
            ridge_predicted = subset.iloc[subset_gap_indices]['ridge'].values
            composite_predicted = subset.iloc[subset_gap_indices]['composite'].values
            
            metrics, ridge, comp = print_gap_metrics(actual, ridge_predicted, composite_predicted, i+1)
            metricx.append(f"diff: {ridge - comp} ridge: {ridge} composite: {comp} metrics: {metrics}")
            ttl_ridge += ridge
            ttl_comp += comp
            # Plot this gap individually if it's large enough
            if plot_individual_gaps and gap_size >= min_gap_size:
                # We only plot context around the gap, not the entire subset
                context_size = min(buffer_size, before_after)  # Show at most 60 data points before/after
                plot_start = int(max(0, subset_gap_indices[0] - context_size))
                plot_end = int(min(len(subset) - 1, subset_gap_indices[-1] + context_size))
                
                plot_subset = subset.iloc[plot_start:plot_end+1].copy()
                
                # Adjust gap indices for the new reduced plotting subset
                plot_gap_indices = [idx - plot_start for idx in subset_gap_indices 
                                   if plot_start <= idx <= plot_end]
                
                # Plot this specific gap
                with plt.style.context('seaborn-v0_8-whitegrid'):  # Use a nice style
                    plot_gap_result(plot_subset, plot_gap_indices, i+1)
                    # Allow the user to press Enter to continue or just wait
                    # when in interactive mode
                    print("Showing plot... (close the plot window or press Enter to continue)")
                    try:
                        # This will return immediately if Enter is pressed
                        # or timeout after 5 seconds
                        plt.pause(2)
                    except Exception:
                        pass
            else:
                context_size = min(buffer_size, before_after)  # Show at most 60 data points before/after
                plot_start = int(max(0, subset_gap_indices[0] - context_size))
                plot_end = int(min(len(subset) - 1, subset_gap_indices[-1] + context_size))
                
                # plot_subset = subset.iloc[plot_start:plot_end+1].copy()
                # total_ridge = plot_subset.loc[plot_subset['with_gap'] == -100, 'ridge'].sum()
                # print(total_ridge)
                # file_name = f'{data_dir}\\{table}_subset_zero_gaps_{i}.xlsx'
                # if total_ridge > 0:
                #     file_name = f'{data_dir}\\{table}_subset_gaps_{i}.xlsx'
                # if ridge - comp < 0:
                #     file_name = file_name.replace(".xlsx",f"_ridge_{i}.xlsx")
                # else:
                #     file_name = file_name.replace(".xlsx",f"_comp_{i}.xlsx")
                                                 
                
                # plot_subset['ridge_comp_diff'] = plot_subset['ridge'] - plot_subset['composite']
                # plot_subset.to_excel(file_name, index=False, sheet_name="Data")
                # add_chart_to_excel(file_name, metrics, ridge, comp, gap_size, window_size)

        # Clean up subset to free memory
        del subset, subset_gap_indices, original_positions
        gc.collect()

        print(f" {table} Completed gap {i+1}.")
    
    print("All gaps have been filled successfully.")
    file_name = os.path.join(data_dir, f'{table}_all_gaps.xlsx')
    if min_gap > max_gap:
        ttl_ridge = ttl_comp = 0.0
        type = 'median'
        suffix = '_median.xlsx'
    elif ttl_ridge - ttl_comp < 0:
        type = 'ridge'
        suffix = '_ridge.xlsx'
    else:
        type = 'composite'
        suffix = '_comp.xlsx'
    file_name = file_name.replace(".xlsx", suffix)
    
    if not df_xcel.empty:
        df_xcel['ridge_comp_diff'] = df_xcel['ridge'] - df_xcel['composite']
        df_xcel.to_excel(file_name, index=False, sheet_name="Data")
        add_chart_to_excel(file_name, metricx, ttl_ridge, ttl_comp, gap_size, window_size, ChartType=LineChart)
    # remove padding values
    return dataframe[padding:(padding *-1)], type


def hyper_gaps(table, data_dir, minutes=30):
    gap_csv_path = os.path.join(data_dir, f'{table}_gaps.csv')
    if os.path.isfile(gap_csv_path):
        df = pd.read_csv(gap_csv_path)
        df[["wh", "with_gap"]] = df[["wh", "with_gap"]].astype("float32")
        df["date"] = pd.to_datetime(df["date"])
        df_cleaned = replace_outliers_interpolate(df, column='wh', window=minutes, threshold=3)
        if not df_cleaned.empty:
            path = os.path.join(data_dir, f'{table.replace("gb_", "gb_clean_")}')
            df_cleaned.to_csv(path + r"_gaps.csv", mode="w", index=False, header=True)
        return df, df_cleaned
    conn = engines[1].raw_connection()  # Get raw psycopg2 connection
    with conn.cursor() as cursor:
        # Step 1: Get all hypertables in the "eyedro" schema
        cursor.execute(f"""select ts date, (abs(wh_p1) + abs(wh_p2) + abs(wh_p3)) wh, (abs(wh_p1) + abs(wh_p2) + abs(wh_p3)) with_gap 
                       from eyedro.{table}
                       where wh_p1 is not null and wh_p2 is not null and wh_p3 is not null
                       AND NOT wh_p1::text = 'NaN' AND NOT wh_p1::text = 'NaN' AND NOT wh_p3::text = 'NaN'
                       order by ts;""")

        rows = cursor.fetchall()
        print(f"Processing hypertable {table} Rows: {len(rows)}")

        df = pd.DataFrame(rows, columns=["date", "wh", "with_gap"])
        df[["wh", "with_gap"]] = df[["wh", "with_gap"]].astype("float32")
        df["date"] = pd.to_datetime(df["date"])

        # Save results to CSV
        if not df.empty:
            path = os.path.join(data_dir, table)
            df.to_csv(path + r"_gaps.csv", mode="w", index=False, header=True)
        # Close connection
        cursor.close()
        conn.close()
    df_cleaned = replace_outliers_interpolate(df, column='wh', window=minutes, threshold=3)
    return df, df_cleaned


def update_gaps_table(tn, db_eng):
    ####### = db.set_local_defaultdb_engine() if local else db.set_azure_defaultdb_engine()

    # save data to gaps table
    res, err = db.sql_execute(f"""
        WITH ordered_epochs AS (
            SELECT epoch_secs, 
                LAG(epoch_secs) OVER (ORDER BY epoch_secs) AS prev_epoch
            FROM eyedro.{tn[0]}
        )
        SELECT epoch_secs, prev_epoch, (epoch_secs - prev_epoch) AS diff_seconds
        FROM ordered_epochs
        WHERE (epoch_secs - prev_epoch) != 60;
    """, db_eng)

    if err:
        logging.error(err)
        return None, err

    gaps = []
    for row in res:
        # Insert into database, ensuring uniqueness
        res1, err = db.sql_execute(f"""
            INSERT INTO {const.GB_GAPS_TABLE} (hypertable_name, epoch_secs, prev_epoch, diff_seconds, days)
            VALUES ('{tn[0]}', {row[0]}, {row[1]}, {row[2]}, '{row[2] / (60 * 60 * 24):.2f}')
            ON CONFLICT (hypertable_name, epoch_secs, prev_epoch, deleted) DO NOTHING RETURNING 1;
        """, db_eng)
        if err: 
            logging.error(err)
            return None, err
        sn = tn[0][3:]
        ts_from = datetime.fromtimestamp(row[1], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
        ts_to = datetime.fromtimestamp(row[0], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
        days = f'{row[2] / (60 * 60 * 24):.2f}'
        print(f"{sn} from: {ts_from} to: {ts_to} days: {days}")
        gaps.append([sn, row[1], row[0], row[2], ts_from, ts_to, days])
    return gaps, None


def concurrent_update_gaps(local=True, csv_path=None, run_dt=datetime.now().date()):
    """
    Concurrently update gaps table for all GB tables.
    
    If a gaps csv file is provided, it will be read instead of updating the gaps table.

    Parameters:
        local (bool): Whether to use a local database connection or an Azure database connection. Defaults to True.
        csv_path (str): The path to the csv file containing the gaps data. If provided, the function will read from the csv file instead of updating the gaps table.
        run_dt (datetime.date): The date to run the gap filling for. Defaults to the current date.

    Returns:
        A pandas DataFrame containing the gaps data, or None if there is an error.
        An error message if there is an error.
    """
    if csv_path and os.path.exists(csv_path):
        df = pd.read_csv(csv_path)
        return df, None

    def process_chunk(chunk, param1):
        return [update_gaps_table(item, db_eng=param1) for item in chunk]  # Apply function to each item
    eng = db.set_local_defaultdb_engine() if local else db.set_azure_defaultdb_engine()

    # Ensure the gaps table exists
    sql = f'select epoch_secs from {const.GB_GAPS_TABLE} limit 1;'
    res, err = db.sql_execute(sql, eng)
    if err:
        res, err = db.sql_execute(const.SQL_GB_GAPS_TABLE, eng)
        if err:
            logging.error(f'{sql} ERROR: {err}')
            return None, err 
    else:
        # soft delete existing gaps if not done today
        sql = f'select min(updated_at) from {const.GB_GAPS_TABLE} where deleted is true;'
        res, err = db.sql_execute(sql, eng)
        if err:
            logging.error(f'{sql} ERROR: {err}')
            return None, err
        if res[0][0] is None or res[0][0].date() < run_dt:
            sql = f'{const.SQL_GB_GAPS_DELETE}'
            res, err = db.sql_execute(sql, eng)
            if err:
                logging.error(f'{sql} ERROR: {err}')
                return None, err

    # Fetch hypertable names
    sql = """SELECT hypertable_name FROM timescaledb_information.hypertables WHERE hypertable_name LIKE 'gb_%' ORDER BY hypertable_name;"""
    gb_tn_list, err = db.sql_execute(sql, eng)
    if err:
        logging.error(err)
        return None, err

    # set chunks from gb table names list
    num_parts = 10
    chunks = [list(chunk) for chunk in np.array_split(gb_tn_list, num_parts)]  # Ensure list format

    with ThreadPoolExecutor(max_workers=num_parts) as executor:
        results = list(executor.map(partial(process_chunk, param1=eng), chunks))
    
    gaps = [item for sublist in results for item in sublist]
    # Filter out tuples where the first element is empty or None
    filtered_data = [array for array, value in gaps]
    # Remove the outer list and keep the nested lists
    gap_list = [sublist for array in filtered_data for sublist in array]
    cols=['gb_sn', 'epoch_secs', 'prev_epoch', 'diff_seconds', 'start_ts', 'end_ts', 'days']
    df = pd.DataFrame(gap_list, columns=cols)
    if csv_path:
        df.to_csv(csv_path, index=False)
    # Flatten results
    return df, None

#!!!! run to update eyedro.gb_gaps table
# res = concurrent_update_gaps(local=True)
# logging.debug(res)
# pass
#!!!! run to update eyedro.gb_gaps table

#!!!! gp_fill_sn_list manually generated TODO: make it automatic
gp_fill_sn_list = [
['009-8064a','2023-03-13','1900-01-00','0.5'],
['009-8064d','2004-07-22','1900-01-00','0.5'],
['009-80741','2024-07-24','1900-01-00','0.5'],
['009-80742','2024-11-25','1900-01-00','0.5'],
['009-80747','2024-08-23','1900-01-00','1'],
['009-8074a','2023-08-24','1900-01-00','1'],
['009-80755','2023-09-11','1900-01-00','1'],
['009-8075b','2023-09-29','1900-01-00','1'],
['009-80780','2023-09-22','1900-01-00','1'],
['009-80785','2023-08-21','1900-01-00','0.5'],
['009-8078a','2024-01-24','1900-01-00','1'],
['009-8079a','2023-07-18','1900-01-00','1'],
['009-807b3','2023-07-20','1900-01-00','1'],
['009-807b4','2023-08-14','1900-01-00','1'],
['009-807c4','2023-09-27','1900-01-00','1'],
['009-807d7','2024-09-25','1900-01-00','1'],
['009-807d8','2024-09-25','1900-01-00','1'],
['009-807f5','2023-04-27','1900-01-00','0.5'],
['009-80820','2023-10-12','1900-01-00','1'],
['009-80822','2023-06-13','1900-01-00','0.5'],
['009-80823','2024-11-17','1900-01-00','1'],
['009-8082b','2024-09-04','1900-01-00','1'],
['009-8082d','2024-03-27','1900-01-00','0.5'],
['009-80830','2024-10-11','1900-01-00','1'],
['009-80832','2024-06-27','1900-01-00','1'],
['009-80833','2023-04-17','1900-01-00','0.5'],
['009-80836','2023-09-22','1900-01-00','1'],
['009-80839','2023-09-11','1900-01-00','1'],
['009-8083a','2024-10-10','1900-01-00','1'],
['009-8083b','2024-04-08','1900-01-00','0.5'],
['009-8083c','2024-04-19','1900-01-00','0.5'],
['009-8083e','2024-01-23','1900-01-00','0.5'],
['009-8083f','2024-01-17','1900-01-00','1'],
['009-80843','2024-06-19','1900-01-00','1'],
['009-80844','2023-12-01','1900-01-00','1'],
['009-80846','2024-08-20','1900-01-00','0.5'],
['009-80848','2023-07-23','1900-01-00','1'],
['009-8084b','2024-02-19','1900-01-00','0.5'],
['009-8084d','2024-08-06','1900-01-00','1'],
['009-8084e','2021-08-24','1900-01-00','0.5'],
['009-8085b','2021-12-18','1900-01-00','1'],
['009-80862','2022-03-22','1900-01-00','1'],
['009-80864','2022-03-22','1900-01-00','1'],
['009-80866','2023-05-28','1900-01-00','0.5'],
['009-8086f','2022-12-12','1900-01-00','1'],
['009-80874','2023-07-11','1900-01-00','1'],
['009-80875','2023-07-26','2024-09-30','1'],
['009-8087b','2023-04-04','1900-01-00','1'],
['009-80882','2022-07-27','1900-01-00','1'],
['009-80883','2022-10-02','1900-01-00','1'],
['009-80889','2021-09-18','1900-01-00','1'],
['009-8088d','2021-03-25','1900-01-00','1'],
['009-8088e','2022-10-27','1900-01-00','1'],
['009-80890','2022-07-08','1900-01-00','1'],
['009-80891','2022-05-06','2025-01-20','1'],
['009-80896','2022-05-26','1900-01-00','1'],
['009-80898','2022-05-26','1900-01-00','1'],
['009-80899','2022-05-26','1900-01-00','1'],
['009-8089b','2022-05-26','1900-01-00','1'],
['009-8089d','2022-01-28','1900-01-00','1'],
['009-8089e','2021-09-17','1900-01-00','1'],
['009-8089f','2022-05-12','1900-01-00','1'],
['009-808a8','2022-12-20','1900-01-00','1'],
['009-808a9','2022-04-24','1900-01-00','1'],
['009-808b5','2022-05-26','1900-01-00','1'],
['009-808b9','2022-03-09','1900-01-00','1'],
['009-808bb','2022-03-03','1900-01-00','0.5'],
['009-808bd','2022-02-11','2024-03-31','1'],
['009-808c6','2022-09-05','1900-01-00','0.5'],
['009-808d9','2022-05-18','1900-01-00','0.5'],
['009-808e0','2021-08-13','1900-01-00','1'],
['009-808e1','2021-08-13','1900-01-00','1'],
['009-808e9','2022-04-25','1900-01-00','1'],
['009-808ec','2022-05-23','1900-01-00','1'],
['009-808f0','2023-08-12','1900-01-00','1'],
['009-808f1','2022-09-02','1900-01-00','1'],
['009-808f4','2022-03-22','1900-01-00','1'],
['009-808f6','2022-02-25','1900-01-00','1'],
['009-808f7','2022-02-02','1900-01-00','1'],
['009-808f8','2022-02-02','1900-01-00','1'],
['009-808fd','2024-10-27','2024-05-31','0.5'],
['009-808fe','2024-10-22','1900-01-00','1'],
['009-8090a','2022-04-10','2025-03-12','1'],
['009-8090c','2022-03-01','1900-01-00','0.5'],
['009-8090d','2022-03-22','1900-01-00','1'],
['009-80911','2022-03-14','1900-01-00','1'],
['009-80912','2022-03-15','2024-07-05','1'],
['009-80914','2022-04-08','2024-06-01','1'],
['009-80923','2022-03-08','1900-01-00','1'],
['009-80929','2021-12-27','1900-01-00','0.5'],
['009-8092d','2022-02-03','1900-01-00','1'],
['009-8092f','2022-03-15','1900-01-00','0.5'],
['009-80935','2021-12-19','1900-01-00','0.5'],
['009-80954','2022-04-06','1900-01-00','1'],
['009-80958','2022-04-22','1900-01-00','0.5'],
['009-8095f','2021-11-02','1900-01-00','0.5'],
['009-80960','2022-04-22','1900-01-00','1'],
['009-80961','2021-10-15','1900-01-00','0.5'],
['009-80962','2021-12-14','1900-01-00','0.5'],
['009-8097f','2021-10-14','1900-01-00','1'],
['009-80981','2021-10-20','1900-01-00','1'],
['009-80982','2022-05-23','1900-01-00','1'],
['009-80984','2021-05-08','1900-01-00','1'],
['009-8098d','2021-10-27','1900-01-00','1'],
['009-809ac','2021-10-04','1900-01-00','1'],
['009-809ad','2022-05-05','2024-03-25','1'],
['009-809b5','2021-09-02','1900-01-00','1'],
['009-809b8','2021-10-15','1900-01-00','1'],
['009-809c9','2021-09-22','1900-01-00','1'],
['009-809ca','2022-01-13','1900-01-00','0.5'],
['009-809cb','2021-09-07','1900-01-00','1'],
['009-809d4','2023-05-01','1900-01-00','1'],
['009-809df','2022-01-24','1900-01-00','0.5'],
['009-809e1','2021-12-18','1900-01-00','0.5'],
['009-809e9','2021-07-13','1900-01-00','1'],
['009-809ee','2021-12-11','1900-01-00','1'],
['009-809f0','2021-07-29','1900-01-00','1'],
['009-809f3','2021-09-29','1900-01-00','1'],
['009-80a00','2021-06-17','1900-01-00','1'],
['009-80a01','2022-07-13','1900-01-00','0.5'],
['009-80a0b','2021-06-10','1900-01-00','0.5'],
['009-80a0c','2022-04-12','1900-01-00','0.5'],
['009-80a0d','2021-04-15','1900-01-00','1'],
['009-80a0f','2021-04-15','1900-01-00','1'],
['009-80a14','2022-03-22','1900-01-00','1'],
['009-80a16','2021-06-03','1900-01-00','1'],
['009-80a19','2021-07-05','1900-01-00','1'],
['009-80a1c','2021-12-16','1900-01-00','1'],
['009-80a20','2021-12-16','1900-01-00','1'],
['009-80a24','2022-06-24','1900-01-00','1'],
['009-80a27','2021-05-27','1900-01-00','1'],
['009-80a29','2021-06-03','1900-01-00','1'],
['009-80a2e','2022-06-19','1900-01-00','1'],
['009-80a2f','2021-10-06','1900-01-00','1'],
['009-80a30','2021-04-01','1900-01-00','1'],
['009-80a35','2021-04-21','1900-01-00','1'],
['009-80a37','2021-12-12','1900-01-00','1'],
['009-80a38','2022-02-25','1900-01-00','1'],
['009-80a39','2022-02-25','1900-01-00','1'],
['009-80a3c','2021-06-02','1900-01-00','1'],
['009-80a3e','2021-06-02','1900-01-00','1'],
['009-80a40','2021-06-02','1900-01-00','1'],
['009-80a4e','2021-01-26','1900-01-00','1'],
['009-80a4f','2024-08-22','1900-01-00','1'],
['009-80a50','2021-03-24','2024-07-26','1'],
['009-80a52','2022-02-25','1900-01-00','1'],
['009-80a6d','2022-02-25','1900-01-00','1'],
['009-80a6f','2022-02-25','1900-01-00','1'],
['009-80a73','2021-07-18','1900-01-00','0.5'],
['009-80a75','2021-01-05','1900-01-00','0.5'],
['009-80a99','2022-02-08','1900-01-00','0.5'],
['009-80a9d','2023-05-27','1900-01-00','1'],
['009-80a9e','2021-09-16','1900-01-00','1'],
['009-80a9f','2021-10-15','1900-01-00','0.5'],
['009-80aa1','2021-10-29','1900-01-00','0.5'],
['009-80aa5','2021-12-09','1900-01-00','0.5'],
['009-80ab8','2021-08-11','1900-01-00','1'],
['009-80abb','2021-03-14','1900-01-00','1'],
['009-80ac4','2022-02-25','1900-01-00','1'],
['009-80af1','2021-03-14','1900-01-00','1'],
['009-80af4','2021-10-14','1900-01-00','0.5'],
['009-80af5','2021-12-22','1900-01-00','1'],
['009-80b11','2021-03-24','1900-01-00','1'],
['009-80b17','2022-03-03','1900-01-00','1'],
['009-80b19','2022-03-15','1900-01-00','1'],
['009-80b22','2022-03-31','1900-01-00','1'],
['009-80b23','2022-03-01','1900-01-00','1'],
['009-80b27','2021-05-26','1900-01-00','0.5'],
['009-80b2c','2021-01-18','1900-01-00','0.5'],
['009-80b2f','2021-12-18','1900-01-00','1'],
['009-80b33','2022-07-05','1900-01-00','0.5'],
['009-80b58','2021-01-21','1900-01-00','0.5'],
['009-80b66','2021-03-09','1900-01-00','0.5'],
['009-80b6b','2021-02-10','1900-01-00','1'],
['009-80b7a','2021-10-20','1900-01-00','0.5'],
['009-80b80','2021-03-15','1900-01-00','1'],
['009-80b86','2021-05-14','1900-01-00','1'],
['009-80b96','2021-04-07','1900-01-00','0.5'],
['009-80da4','2021-02-10','1900-01-00','1'],
['009-80da6','2021-02-10','1900-01-00','1'],
['009-80daa','2021-04-13','1900-01-00','1'],
['009-80db4','2022-04-06','1900-01-00','1'],
['009-80db6','2022-04-05','1900-01-00','1'],
['009-80db9','2021-02-16','1900-01-00','1'],
['009-80dba','2021-02-20','2025-02-14','0.5'],
['009-80dbb','2020-12-01','1900-01-00','1'],
['009-80dc1','2022-01-20','1900-01-00','1'],
['009-80dc5','2021-11-19','1900-01-00','1'],
['009-80dcd','2020-12-12','1900-01-00','1'],
['009-80dd1','2022-04-29','2024-07-30','1'],
['009-80dd3','2021-01-19','1900-01-00','1'],
['009-80dde','2021-02-16','2024-10-25','1'],
['009-80de3','2020-12-10','1900-01-00','0.5'],
['009-80de5','2022-08-17','1900-01-00','1'],
['009-80de7','2021-08-03','1900-01-00','1'],
['009-80de8','2024-05-30','1900-01-00','1'],
['009-80df1','2021-05-18','1900-01-00','0.5'],
['009-80df9','2021-06-08','1900-01-00','0.5'],
['009-80e10','2021-08-07','1900-01-00','1'],
['009-80e17','2021-02-25','1900-01-00','1'],
['009-80e18','2021-09-03','1900-01-00','0.5'],
['009-80e1c','2022-07-19','1900-01-00','1'],
['009-80e29','2021-02-23','1900-01-00','0.5'],
['b12-0045e','2021-10-07','1900-01-00','1'],
['b12-00464','2022-04-16','1900-01-00','1'],
['b12-004b7','2023-06-20','1900-01-00','1'],
['b12-00613','2021-11-17','1900-01-00','1'],
]

#['009-8064a','2023-03-13','1900-01-00','0.5'],
# for el in gp_fill_sn_list:
#     sn = el[0]
#     act_dt = el[1]
#     deact_dt = el[2]
#     if deact_dt == '1900-01-00':
#         deact_dt = None
#     factor = float(el[3])
#     print(sn, act_dt, deact_dt, factor)

ttl_wh = []
# Main function to run the example
def get_gb_gaps(gb_sn_list=gp_fill_sn_list, data_dir=const.GB_GAPS_DATA_DIR, run_dt=datetime.now().date()):
    # Start with a clean memory state
    gc.collect()
    print("Starting memory-efficient gap filling process...") 
    # TODO: get list of gbs from dfaultdb public.gb_api_no_dups  fileter 006-
    for el in gb_sn_list: #['gb_0098082b']:
        if el[0] < 'b12-004b7':
            continue
        table = 'gb_' + el[0].replace('-', '')
        act_dt = el[1]
        deact_dt = el[2]
        if deact_dt == '1900-01-00':
            deact_dt = None
        factor = float(el[3])
        print(table, act_dt, deact_dt, factor)

        df, df_cleaned = hyper_gaps(table,data_dir)
        df_2024 = df[df['date'].dt.year == 2024]
        df_cl_2024 = df_cleaned[df_cleaned['date'].dt.year == 2024]

        if df_cl_2024.empty:
            sum_wh = [table, 0, 0, 'N/A', -1, -1, -1, -1]
            ttl_wh.append(sum_wh)
            print(f"No 2024 data found for table {table}.")
            df_ttl_wh = pd.DataFrame(ttl_wh, columns=["table", "wh", "wh_cleaned", "type", 'wh_2024', 'wh_median', 'wh_ridge', 'wh_comp'])
            df_ttl_wh.to_csv(os.path.join(data_dir, 'ttl_wh.csv'), mode="w", index=False, header=True)
            continue
        else:
            sum_wh = [table, df_2024['wh'].sum(), df_cl_2024['wh'].sum()]
            path = os.path.join(data_dir, f'{table.replace("gb_", "gb_clean_")}')
            df_cl_2024.to_csv(path + r"_gaps_2024.csv", mode="w", index=False, header=True)

        file_path = f'{table}_subset_gaps.csv'
        # Run the gap filling with our memory-efficient approach
        full_path = os.path.join(data_dir, file_path)
        df_gf_cl_2024, type = run_gapfilling_by_segments(full_path, plot_individual_gaps=False, min_gap_size=10, df=df_cl_2024)

        if df_gf_cl_2024 is None:
            sum_wh.append('N/A')
            sum_wh.append(-1)
            sum_wh.append(-1)
            sum_wh.append(-1)
            sum_wh.append(-1)
            ttl_wh.append(sum_wh)
            print(f"No gaps found for table {table}.")
            df_ttl_wh = pd.DataFrame(ttl_wh, columns=["table", "wh", "wh_cleaned", "type", 'wh_2024', 'wh_median', 'wh_ridge', 'wh_comp'])
            df_ttl_wh.to_csv(os.path.join(data_dir, 'ttl_wh.csv'), mode="w", index=False, header=True)
            continue

        df2024 = df_gf_cl_2024[df_gf_cl_2024['date'].dt.year == 2024]
        file_path = file_path.replace('gb_', 'gb_all_')
        full_path = os.path.join(data_dir, file_path)
        df2024.to_csv(full_path, mode="w", index=False, header=True)
        wh = df2024[df2024['with_gap'] > -100]['wh'].sum()
        wh_med = df2024[df2024['with_gap'] < -99]['wh'].sum()
        wh_ridge = df2024[df2024['with_gap'] < -99]['ridge'].sum()
        wh_comp = df2024[df2024['with_gap'] < -99]['composite'].sum()

        sum_wh.append(type)
        sum_wh.append(wh)
        sum_wh.append(wh_med)
        sum_wh.append(wh_ridge)
        sum_wh.append(wh_comp)
        ttl_wh.append(sum_wh)
        df_ttl_wh = pd.DataFrame(ttl_wh, columns=["table", "wh", "wh_cleaned", "type", 'wh_2024', 'wh_median', 'wh_ridge', 'wh_comp'])
        df_ttl_wh.to_csv(os.path.join(data_dir, 'ttl_wh.csv'), mode="w", index=False, header=True)
        print("\nProcess completed successfully.")
        print("Interactive plots were displayed during processing.")

        # Keep the plots open until the user closes the program
        print("Press Ctrl+C to exit...")
        plt.ioff()  # Turn off interactive mode
        plt.show(block=True)  # This will block until all plot windows are closed

    for x in ttl_wh:
        print(x)

# filtered_gb_sn_df = all_gb_api_sn_df[0][
#     ~all_gb_api_sn_df[0]["gb_serial"].str.startswith(const.GB_GATEWAY_PREFIX)
# ]

FILTERED_GB_SN_PATH=const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt.isoformat())
filtered_gb_sn_df = pd.read_csv(FILTERED_GB_SN_PATH)
arr = (filtered_gb_sn_df[['gb_serial','epoch_utc']].values.astype(np.int32)).tolist()

pass
