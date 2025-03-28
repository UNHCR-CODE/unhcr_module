from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from functools import partial
import io
import itertools
import json
import logging
import threading
import numpy as np
import os
import re
import time
import winsound
from numpy import sort
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
import pandas as pd
from psycopg2.extras import execute_values
from psycopg2 import DatabaseError
import requests
from sqlalchemy import text
import sys
import zipfile

from unhcr import constants as const
from unhcr import db
from unhcr import utils
from unhcr import gb_eyedro



# local testing ===================================
if const.LOCAL:  # testing with local python files
    const, db, utils, gb_eyedro = const.import_local_libs(
        mods=[
            ["constants", "const"],
            ["db", "db"],
            ["utils", "utils"],
            ["gb_eyedro", "gb_eyedro"],
        ]
    )

utils.log_setup(level="INFO", log_file="unhcr.gb_1miny.log", override=True)
logging.info(
    f"{sys.argv[0]} Process ID: {os.getpid()}   Log Level: {logging.getLevelName(logging.getLogger().getEffectiveLevel())}"
)

if not utils.is_version_greater_or_equal('0.4.7'):
    logging.error(
        "This version of the script requires at least version 0.4.6 of the unhcr module."
    )
    exit(47)

#!!! this came from Istvans excel 2024-02-27 "New Top 20 - Copy.xlsx" 
# Actually the tables that I could get data for (add eyedro.gb_ for AZURE defaultdb table name)
# Here is the list of errors:
"""
    2025-02-27 10:14:12,349 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980829 ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:14:28,330 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 0098084C ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:14:46,387 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 0098085D ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:14:48,485 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 0098086A ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:14:48,798 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 0098086C ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:14:49,104 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 0098086D ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:15:14,057 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 009808DF ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:15:24,188 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980953 ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:15:35,050 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 009809B7 ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:16:13,562 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980A03 ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:16:45,553 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980A9A ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:16:55,433 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980AAF ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:16:55,763 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980ABA ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:17:07,954 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980B12 ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:17:20,336 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980B2E ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
2025-02-27 10:18:23,851 - ERROR - 2025-02-27T00:00:00+00:00 ZZZ 00980B93 ERRORS: [[1210, 'API Error. Contact Eyedro Admin.'], [899, 'DeviceSerial invalid, or device owner has not claimed device.']]
"""

genset_gbs_org = ['00980864', '0098086a', '0098086c', '0098086d', '009807c4', '009807d3', '009807d8', '00980827', '00980828', '00980829', '0098082e', '0098082f', '0098082d', '00980830', '00980848', '00980849', '0098084c', '0098084f', '00980857', '0098085c', '0098085b', '0098085d', '00980aa1', '00980890', '0098087b', '00980891', '0098087c', '009808c7', '009808b5', '009808c6', '00980893', '00980892', '009808e8', '009808dc', '009808df', '009808db', '00980907', '00980914', '00980953', '009808f8', '009808f1', '009809ad', '00980954', '009809b5', '009809b7', '009809c3', '009809c4', '009809e5', '009809b8', '009809ea', '009809f3', '009809f4', '009809e6', '009809e9', '009809fd', '00980a02', '00980a03', '00980a01', '00980a06', '00980a07', '00980a14', '00980a05', '00980a17', '00980a27', '00980a20', '00980a21', '00980a24', '00980a50', '00980a40', '00980a52', '00980a74', '00980aa0', '00980a9a', '00980aa2', '00980a53', '00980aa5', '00980abb', '00980aaf', '00980aba', '00980ac9', '00980acb', '00980b11', '00980af4', '00980b12', '00980b1c', '00980b1e', '00980b2e', '00980b2a', '00980b29', '00980b2f', '00980b36', '00980b37', '00980b58', '00980b35', '00980b6b', '00980b5e', '00980b6e', '00980b6c', '00980b74', '00980b76', '00980b6f', '00980b75', '00980b80', '00980b81', '00980b77', '00980b89', '00980b8a', '00980b8b', '00980b87', '00980b91', '00980b96', '00980b93', '00980b90', '00980b8c', '00980da4', '00980b97', '00980da6', '00980dbb', '00980dd3', '00980da7', '00980dcb', '00980dcd', '00980dd6', '00980de3', '00980dd4', '00980de8', '00980e05', '00980dfc', '00980e09', '00980e08', '00980e10', '00980e0e', '00980e13', '00980e1f', '00980e22', '00980e27', 'b120045e', '00980e29', '00980e28', 'b120045f', '00980e2a', 'b1200464', 'b1200465', 'b1200631', 'b1200461', '009809f5', '00980a2a', '00980a29', '00980b13', '00980de5', '00980b6a', '00980b98', '00980b84', '009808c8', '00980ddb', '009809cb', '00980e1c', '00980b33', '0098097b', '00980850', '00980889', '00980826', '00980845', '0098084e', '00980a19', '00980aa3', '00980ac1', '00980b1a', '00980b34', '00980b70', '00980b7a', '00980b86', '00980da0', '00980dd5', '00980df1', '00980e18', 'b1200462']

genset_gbs = [
'009807c4',
'009807d3',
'009807d8',
'00980826',
'00980827',
'00980828',
'00980829',
'0098082d',
'0098082e',
'0098082f',
'00980830',
'00980845',
'00980848',
'00980849',
'0098084c',
'0098084e',
'0098084f',
'00980850',
'00980857',
'0098085b',
'0098085c',
'0098085d',
'00980864',
'0098086a',
'0098086c',
'0098086d',
'0098087b',
'0098087c',
'00980889',
'00980890',
'00980891',
'00980892',
'00980893',
'009808c6',
'009808c7',
'009808c8',
'009808b5',
'009808db',
'009808dc',
'009808df',
'009808e8',
'009808f1',
'009808f8',
'00980907',
'00980914',
'00980953',
'00980954',
'0098097b',
'009809ad',
'009809b5',
'009809b7',
'009809b8',
'009809c3',
'009809c4',
'009809cb',
'009809e5',
'009809e6',
'009809e9',
'009809ea',
'009809f3',
'009809f4',
'009809f5',
'009809fd',
'00980a01',
'00980a02',
'00980a03',
'00980a05',
'00980a06',
'00980a07',
'00980a14',
'00980a17',
'00980a19',
'00980a20',
'00980a21',
'00980a24',
'00980a27',
'00980a29',
'00980a2a',
'00980a40',
'00980a50',
'00980a52',
'00980a53',
'00980a74',
'00980a9a',
'00980aa0',
'00980aa1',
'00980aa2',
'00980aa3',
'00980aa5',
'00980abb',
'00980aba',
'00980aaf',
'00980ac1',
'00980ac9',
'00980acb',
'00980af4',
'00980b11',
'00980b12',
'00980b13',
'00980b1a',
'00980b1c',
'00980b1e',
'00980b29',
'00980b2a',
'00980b2e',
'00980b2f',
'00980b33',
'00980b34',
'00980b35',
'00980b36',
'00980b37',
'00980b58',
'00980b5e',
'00980b6a',
'00980b6b',
'00980b6c',
######################'00980b6d',
'00980b6e',
'00980b6f',
'00980b70',
'00980b74',
'00980b75',
'00980b76',
'00980b77',
'00980b7a',
'00980b80',
'00980b81',
'00980b84',
'00980b86',
'00980b87',
'00980b89',
'00980b8a',
'00980b8b',
'00980b8c',
'00980b90',
'00980b91',
'00980b93',
'00980b96',
'00980b97',
'00980b98',
'00980da0',
'00980da4',
'00980da6',
'00980da7',
'00980dbb',
'00980dcb',
'00980dcd',
'00980dd3',
'00980dd4',
'00980dd5',
'00980dd6',
'00980de3',
'00980de5',
'00980de8',
'00980df1',
'00980dfc',
'00980e05',
'00980e08',
'00980e09',
'00980e0e',
'00980e10',
'00980e13',
'00980e18',
'00980e1c',
'00980e1f',
'00980e22',
'00980e27',
'00980e28',
'00980e29',
'00980e2a',
'b120045e',
'b120045f',
'b1200461',
'b1200462',
'b1200464',
'b1200465',
'b1200631'
]

#2025-03-07
all_api_gbs = sorted(['00980639',	'0098063D',	'00980640',	'00980641',	'00980644',	'00980645',	'0098064A',	'0098064C',	'0098064D',	'0098064E',	'00980741',	'00980742',	'00980747',	'0098074A',	'00980753',	'00980755',	'00980756',	'00980757',	'00980759',	'0098075A',	'0098075B',	'00980760',	'00980780',	'00980781',	'00980785',	'00980786',	'00980789',	'0098078A',	'0098078B',	'0098079A',	'009807B3',	'009807B4',	'009807B5',	'009807C4',	'009807D2',	'009807D4',	'009807D5',	'009807D6',	'009807D7',	'009807D8',	'009807D9',	'009807DB',	'009807DE',	'009807F5',	'009807F6',	'009807FB',	'00980820',	'00980822',	'00980823',	'00980825',	'00980826',	'00980827',	'00980828',	'0098082A',	'0098082B',	'0098082D',	'0098082E',	'0098082F',	'00980830',	'00980831',	'00980832',	'00980833',	'00980834',	'00980835',	'00980836',	'00980838',	'00980839',	'0098083A',	'0098083B',	'0098083C',	'0098083E',	'0098083F',	'00980840',	'00980841',	'00980842',	'00980843',	'00980844',	'00980845',	'00980846',	'00980847',	'00980848',	'00980849',	'0098084A',	'0098084B',	'0098084C',	'0098084D',	'0098084E',	'0098084F',	'00980850',	'00980855',	'00980856',	'00980857',	'00980858',	'00980859',	'0098085A',	'0098085B',	'0098085C',	'0098085E',	'0098085F',	'00980860',	'00980861',	'00980862',	'00980863',	'00980864',	'00980865',	'00980866',	'00980867',	'00980868',	'00980869',	'0098086B',	'0098086E',	'0098086F',	'00980870',	'00980871',	'00980872',	'00980873',	'00980874',	'00980875',	'00980876',	'00980877',	'00980878',	'00980879',	'0098087A',	'0098087A',	'0098087B',	'0098087C',	'0098087D',	'0098087F',	'00980880',	'00980881',	'00980882',	'00980883',	'00980885',	'00980887',	'00980888',	'00980889',	'0098088A',	'0098088B',	'0098088C',	'0098088D',	'0098088E',	'00980890',	'00980891',	'00980892',	'00980893',	'00980894',	'00980895',	'00980896',	'00980897',	'00980898',	'00980899',	'0098089A',	'0098089B',	'0098089C',	'0098089D',	'0098089E',	'0098089F',	'009808A8',	'009808A9',	'009808A9',	'009808AA',	'009808AF',	'009808B0',	'009808B1',	'009808B2',	'009808B3',	'009808B4',	'009808B5',	'009808B6',	'009808B7',	'009808B8',	'009808B9',	'009808BA',	'009808BB',	'009808BC',	'009808BD',	'009808BE',	'009808BF',	'009808C6',	'009808C7',	'009808C8',	'009808D5',	'009808D8',	'009808D9',	'009808DA',	'009808DB',	'009808DC',	'009808DD',	'009808DE',	'009808E0',	'009808E1',	'009808E2',	'009808E3',	'009808E5',	'009808E6',	'009808E8',	'009808E9',	'009808EA',	'009808EB',	'009808EC',	'009808EF',	'009808F0',	'009808F1',	'009808F2',	'009808F4',	'009808F5',	'009808F6',	'009808F7',	'009808F8',	'009808FC',	'009808FD',	'009808FE',	'00980901',	'00980902',	'00980905',	'00980906',	'00980907',	'0098090A',	'0098090B',	'0098090C',	'0098090D',	'0098090E',	'0098090F',	'00980910',	'00980911',	'00980912',	'00980914',	'00980915',	'00980916',	'00980918',	'00980919',	'0098091A',	'0098091B',	'00980921',	'00980923',	'00980925',	'00980926',	'00980927',	'00980929',	'0098092A',	'0098092B',	'0098092C',	'0098092D',	'0098092F',	'00980931',	'00980932',	'00980933',	'00980935',	'00980937',	'0098093A',	'0098093B',	'00980952',	'00980954',	'00980957',	'00980958',	'00980959',	'0098095A',	'0098095B',	'0098095F',	'00980960',	'00980961',	'00980962',	'00980966',	'00980967',	'00980978',	'0098097A',	'0098097B',	'0098097C',	'0098097E',	'0098097F',	'00980981',	'00980982',	'00980984',	'00980985',	'00980986',	'00980987',	'00980988',	'00980989',	'0098098D',	'009809A8',	'009809A8',	'009809AC',	'009809AD',	'009809AE',	'009809B5',	'009809B6',	'009809B8',	'009809C3',	'009809C4',	'009809C6',	'009809C7',	'009809C8',	'009809C9',	'009809CA',	'009809CB',	'009809CC',	'009809D4',	'009809DA',	'009809DF',	'009809E1',	'009809E4',	'009809E5',	'009809E6',	'009809E8',	'009809E9',	'009809EA',	'009809ED',	'009809EE',	'009809EF',	'009809F0',	'009809F2',	'009809F2',	'009809F3',	'009809F4',	'009809F5',	'009809F6',	'009809F7',	'009809F9',	'009809FA',	'009809FB',	'009809FC',	'009809FD',	'009809FE',	'009809FF',	'00980A00',	'00980A01',	'00980A02',	'00980A05',	'00980A06',	'00980A07',	'00980A09',	'00980A0A',	'00980A0B',	'00980A0C',	'00980A0D',	'00980A0F',	'00980A10',	'00980A11',	'00980A12',	'00980A13',	'00980A14',	'00980A15',	'00980A16',	'00980A17',	'00980A18',	'00980A19',	'00980A1A',	'00980A1B',	'00980A1C',	'00980A1E',	'00980A1F',	'00980A20',	'00980A21',	'00980A22',	'00980A24',	'00980A25',	'00980A26',	'00980A27',	'00980A28',	'00980A29',	'00980A2A',	'00980A2B',	'00980A2C',	'00980A2D',	'00980A2E',	'00980A2F',	'00980A30',	'00980A31',	'00980A32',	'00980A33',	'00980A34',	'00980A35',	'00980A36',	'00980A37',	'00980A38',	'00980A39',	'00980A3C',	'00980A3D',	'00980A3E',	'00980A3F',	'00980A40',	'00980A47',	'00980A48',	'00980A49',	'00980A4C',	'00980A4D',	'00980A4E',	'00980A4F',	'00980A50',	'00980A51',	'00980A52',	'00980A53',	'00980A57',	'00980A66',	'00980A67',	'00980A68',	'00980A69',	'00980A6A',	'00980A6B',	'00980A6D',	'00980A6E',	'00980A6F',	'00980A72',	'00980A73',	'00980A74',	'00980A75',	'00980A76',	'00980A97',	'00980A99',	'00980A9B',	'00980A9C',	'00980A9D',	'00980A9E',	'00980A9F',	'00980AA0',	'00980AA1',	'00980AA2',	'00980AA3',	'00980AA4',	'00980AA5',	'00980AAE',	'00980AB3',	'00980AB7',	'00980AB8',	'00980AB9',	'00980ABA',	'00980ABB',	'00980AC1',	'00980AC2',	'00980AC3',	'00980AC4',	'00980AC8',	'00980AC9',	'00980ACB',	'00980ACC',	'00980AD1',	'00980ADE',	'00980ADF',	'00980AE3',	'00980AE6',	'00980AEC',	'00980AEE',	'00980AF1',	'00980AF4',	'00980AF5',	'00980AFE',	'00980AFF',	'00980B00',	'00980B01',	'00980B11',	'00980B13',	'00980B14',	'00980B17',	'00980B19',	'00980B1A',	'00980B1C',	'00980B1E',	'00980B21',	'00980B22',	'00980B23',	'00980B24',	'00980B27',	'00980B28',	'00980B2A',	'00980B2C',	'00980B2F',	'00980B31',	'00980B33',	'00980B34',	'00980B35',	'00980B35',	'00980B36',	'00980B37',	'00980B38',	'00980B58',	'00980B5C',	'00980B5E',	'00980B5F',	'00980B60',	'00980B62',	'00980B63',	'00980B64',	'00980B66',	'00980B6A',	'00980B6B',	'00980B6C',	'00980B6E',	'00980B6F',	'00980B70',	'00980B74',	'00980B75',	'00980B76',	'00980B77',	'00980B78',	'00980B79',	'00980B7A',	'00980B7F',	'00980B80',	'00980B81',	'00980B84',	'00980B85',	'00980B86',	'00980B87',	'00980B8A',	'00980B8B',	'00980B8C',	'00980B90',	'00980B91',	'00980B95',	'00980B96',	'00980B97',	'00980B98',	'00980DA1',	'00980DA2',	'00980DA4',	'00980DA5',	'00980DA7',	'00980DA8',	'00980DA9',	'00980DAA',	'00980DAD',	'00980DB4',	'00980DB6',	'00980DB7',	'00980DB9',	'00980DBA',	'00980DBB',	'00980DBF',	'00980DC1',	'00980DC2',	'00980DC3',	'00980DC4',	'00980DC5',	'00980DC6',	'00980DCB',	'00980DCC',	'00980DCD',	'00980DCF',	'00980DD0',	'00980DD1',	'00980DD2',	'00980DD3',	'00980DD4',	'00980DD5',	'00980DD6',	'00980DD7',	'00980DD8',	'00980DD9',	'00980DDB',	'00980DDC',	'00980DDD',	'00980DDE',	'00980DDF',	'00980DE0',	'00980DE3',	'00980DE4',	'00980DE5',	'00980DE6',	'00980DE7',	'00980DE8',	'00980DEC',	'00980DEE',	'00980DF1',	'00980DF4',	'00980DF6',	'00980DF9',	'00980DFC',	'00980DFE',	'00980E05',	'00980E08',	'00980E09',	'00980E0B',	'00980E0D',	'00980E0E',	'00980E0F',	'00980E10',	'00980E11',	'00980E13',	'00980E14',	'00980E17',	'00980E18',	'00980E19',	'00980E1A',	'00980E1B',	'00980E1C',	'00980E1D',	'00980E1F',	'00980E22',	'00980E23',	'00980E24',	'00980E26',	'00980E28',	'00980E29',	'00980E2A',	'91880100',	'B120045D',	'B120045E',	'B120045F',	'B1200461',	'B1200462',	'B1200463',	'B1200464',	'B1200465',	'B1200466',	'B12004B7',	'B12005E8',	'B12005ED',	'B1200613',	'B120061A',	'B1200631',	'B1200632',	'B1400275',	'B1400276'])

top_20_gbs = [
'00980639',
'00980640',
'00980641',
'00980644',
'00980645',
'0098064A',
'0098064C',
'0098064D',
'0098064E',
'00980741',
'00980742',
'00980747',
'0098074A',
'00980753',
'00980754',
'00980756',
'00980757',
'00980759',
'0098075A',
'0098075B',
'00980760',
'009808C9',
'00980E27',
'00980780',
'00980781',
'00980785',
'00980786',
'00980787',
'00980789',
'0098078A',
'0098078B',
'0098079A',
'009807B3',
'009807B4',
'009807B5',
'009807C4',
'009807D2',
'009807D3',
'009807D4',
'009807D5',
'009807D6',
'009807D7',
'009807D8',
'009807D9',
'009807DB',
'009807DD',
'009807DE',
'009807F5',
'009807F6',
'009807FB',
'00980818',
'00980820',
'00980822',
'00980823',
'00980824',
'00980825',
'00980826',
'00980827',
'00980828',
'00980829',
'0098082A',
'0098082B',
'0098082D',
'0098082E',
'0098082F',
'00980830',
'00980831',
'00980832',
'00980833',
'00980834',
'00980835',
'00980836',
'00980837',
'00980838',
'00980839',
'0098083A',
'0098083B',
'0098083C',
'0098083D',
'0098083E',
'0098083F',
'00980840',
'00980841',
'00980842',
'00980843',
'00980844',
'00980845',
'00980846',
'00980847',
'00980848',
'00980849',
'0098084A',
'0098084B',
'0098084C',
'0098084D',
'0098084E',
'0098084F',
'00980850',
'00980851',
'00980852',
'00980853',
'00980854',
'00980855',
'00980B6E',
'00980856',
'00980857',
'00980858',
'00980859',
'0098085A',
'0098085B',
'0098085C',
'0098085D',
'0098085E',
'0098085F',
'00980860',
'00980861',
'00980862',
'00980863',
'00980864',
'00980865',
'00980866',
'00980867',
'00980868',
'00980869',
'0098086A',
'0098086B',
'0098086C',
'0098086D',
'0098086E',
'0098086F',
'00980870',
'00980871',
'00980872',
'00980873',
'00980874',
'00980875',
'00980876',
'00980877',
'00980878',
'00980879',
'0098087A',
'0098087B',
'0098087C',
'0098087D',
'0098087F',
'00980880',
'00980881',
'00980882',
'00980883',
'00980884',
'00980885',
'00980886',
'00980887',
'00980888',
'00980889',
'009808C9',
'0098088A',
'0098088B',
'0098088C',
'009808DF',
'0098088D',
'0098088E',
'0098088F',
'00980890',
'00980891',
'00980892',
'00980893',
'00980894',
'00980895',
'00980896',
'00980897',
'00980898',
'00980899',
'0098089A',
'0098089B',
'0098089C',
'0098089D',
'0098089E',
'0098089F',
'009808A8',
'009808A9',
'009808AA',
'009808AF',
'009808B0',
'009808B1',
'009808B2',
'009808B3',
'009808B4',
'009808B5',
'009808B6',
'009808B7',
'009808B8',
'009808B9',
'009808BA',
'009808BB',
'009808BC',
'009808BD',
'009808BE',
'009808BF',
'009808C0',
'009808C6',
'009808C7',
'009808C8',
'009808D5',
'009808D7',
'009808D8',
'009808D9',
'009808DA',
'009808DB',
'009808DC',
'009808DD',
'009808DE',
'009808DF',
'009808E0',
'009808E1',
'009808E2',
'009808E3',
'009808E5',
'009808E6',
'009808E7',
'009808E8',
'009808E9',
'009808EA',
'009808EB',
'009808EC',
'009808ED',
'009808EE',
'009808EF',
'009808F0',
'009808F1',
'009808F2',
'009808F4',
'009808F5',
'009808F6',
'009808F7',
'009808F8',
'009808F9',
'009808FB',
'009808FC',
'009808FD',
'009808FE',
'00980901',
'00980902',
'00980903',
'00980904',
'00980905',
'00980906',
'00980907',
'0098090A',
'0098090B',
'0098090C',
'0098090D',
'0098090E',
'0098090F',
'00980910',
'00980911',
'00980912',
'00980914',
'00980915',
'00980916',
'00980918',
'00980919',
'0098091A',
'0098091B',
'00980921',
'00980923',
'00980924',
'00980925',
'00980926',
'00980927',
'00980929',
'0098092A',
'0098092B',
'0098092C',
'0098092D',
'0098092F',
'00980931',
'00980932',
'00980933',
'00980935',
'00980936',
'00980937',
'0098093A',
'0098093B',
'00980952',
'00980953',
'00980954',
'00980957',
'00980958',
'00980959',
'0098095A',
'0098095B',
'0098095E',
'0098095F',
'00980960',
'00980961',
'00980962',
'00980966',
'00980967',
'00980978',
'0098097A',
'0098097B',
'0098097C',
'0098097D',
'0098097E',
'0098097F',
'00980981',
'00980982',
'00980984',
'00980985',
'00980986',
'00980987',
'00980988',
'00980989',
'0098098D',
'0098098E',
'009809A8',
'009809AB',
'009809AC',
'009809AD',
'009809AE',
'009809B5',
'009809B6',
'009809B7',
'009809B8',
'009809C3',
'009809C4',
'009809C6',
'009809C7',
'009809C8',
'009809C9',
'009809CA',
'009809CB',
'009809CC',
'009809D4',
'009809DA',
'009809DF',
'009809E0',
'009809E1',
'009809E2',
'009809E4',
'009809E5',
'009809E6',
'009809E8',
'009809E9',
'009809EA',
'009809EB',
'009809EC',
'009809ED',
'009809EE',
'009809EF',
'009809F0',
'009809F1',
'009809F2',
'009809F3',
'009809F4',
'009809F5',
'009809F6',
'009809F7',
'009809F9',
'009809FA',
'009809FB',
'009809FC',
'009809FD',
'009809FE',
'009809FF',
'00980A00',
'00980A01',
'00980A02',
'00980A03',
'00980A05',
'00980A06',
'00980A07',
'00980A08',
'00980A09',
'00980A0A',
'00980A0B',
'00980A0C',
'00980A0D',
'00980A0E',
'00980A0F',
'00980A10',
'00980A11',
'00980A12',
'00980A13',
'00980A14',
'00980A15',
'00980A16',
'00980A17',
'00980A18',
'00980A19',
'00980A1A',
'00980A1B',
'00980A1C',
'00980A1E',
'00980A1F',
'00980A20',
'00980A21',
'00980A22',
'00980A24',
'00980A25',
'00980A26',
'00980A27',
'00980A28',
'00980A29',
'00980A2A',
'00980A2B',
'00980A2C',
'00980A2D',
'00980A2E',
'00980A2F',
'00980A30',
'00980A31',
'00980A32',
'00980A33',
'00980A34',
'00980A35',
'00980A36',
'00980A37',
'00980A38',
'00980A39',
'00980A3C',
'00980A3D',
'00980A3E',
'00980A3F',
'00980A40',
'00980A47',
'00980A48',
'00980A49',
'00980A4B',
'00980A4C',
'00980A4D',
'00980A4E',
'00980A4F',
'00980A50',
'00980A51',
'00980A52',
'00980A53',
'00980A54',
'00980A57',
'00980A59',
'00980A5A',
'00980A66',
'00980A67',
'00980A68',
'00980A69',
'00980A6A',
'00980A6B',
'00980A6C',
'00980A6D',
'00980A6E',
'00980A6F',
'00980A72',
'00980A73',
'00980A74',
'00980A75',
'00980A76',
'00980A97',
'00980A99',
'00980A9A',
'00980A9B',
'00980A9C',
'00980A9D',
'00980A9E',
'00980A9F',
'00980AA0',
'00980AA1',
'00980AA2',
'00980AA3',
'00980AA4',
'00980AA5',
'00980AAE',
'00980AAF',
'00980AB2',
'00980AB3',
'00980AB7',
'00980AB8',
'00980AB9',
'00980ABA',
'00980ABB',
'00980AC1',
'00980AC2',
'00980AC3',
'00980AC4',
'00980AC8',
'00980AC9',
'00980ACA',
'00980ACB',
'00980ACC',
'00980AD1',
'00980ADE',
'00980ADF',
'00980AE3',
'00980AE6',
'00980AEC',
'00980AEE',
'00980AF1',
'00980AF4',
'00980AF5',
'00980AFE',
'00980AFF',
'00980B00',
'00980B01',
'00980B11',
'00980B12',
'00980B13',
'00980B14',
'00980B16',
'00980B17',
'00980B19',
'00980B1A',
'00980B1C',
'00980B1D',
'00980B1E',
'00980B21',
'00980B22',
'00980B23',
'00980B24',
'00980B26',
'00980B27',
'00980B28',
'00980B29',
'00980B2A',
'00980B2C',
'00980B2E',
'00980B2F',
'00980B31',
'00980B33',
'00980B34',
'00980B35',
'00980B36',
'00980B37',
'00980B38',
'00980B58',
'00980B5C',
'00980B5E',
'00980B5F',
'00980B60',
'00980B62',
'00980B63',
'00980B64',
'00980B66',
'00980B67',
'00980B6A',
'00980B6B',
'00980B6C',
'00980B6E',
'00980B6F',
'00980B70',
'00980B74',
'00980B75',
'00980B76',
'00980B77',
'00980B78',
'00980B79',
'00980B7A',
'00980B7F',
'00980B80',
'00980B81',
'00980B82',
'00980B84',
'00980B85',
'00980B86',
'00980B87',
'00980B88',
'00980B89',
'00980B8A',
'00980B8B',
'00980B8C',
'00980B8D',
'00980B8E',
'00980B90',
'00980B91',
'00980B92',
'00980B93',
'00980B95',
'00980B96',
'00980B97',
'00980B98',
'00980DA0',
'00980DA1',
'00980DA2',
'00980DA4',
'00980DA5',
'00980DA6',
'00980DA7',
'00980DA8',
'00980DA9',
'00980DAA',
'00980DAD',
'00980DB4',
'00980DB6',
'00980DB7',
'00980DB9',
'00980DBA',
'00980DBB',
'00980DBF',
'00980DC1',
'00980DC2',
'00980DC3',
'00980DC4',
'00980DC5',
'00980DC6',
'00980DCB',
'00980DCC',
'00980DCD',
'00980DCF',
'00980DD0',
'00980DD1',
'00980DD2',
'00980DD3',
'00980DD4',
'00980DD5',
'00980DD6',
'00980DD7',
'00980DD8',
'00980DD9',
'00980DDB',
'00980DDC',
'00980DDD',
'00980DDE',
'00980DDF',
'00980DE0',
'00980DE3',
'00980DE4',
'00980DE5',
'00980DE6',
'00980DE7',
'00980DE8',
'00980DEC',
'00980DEE',
'00980DF1',
'00980DF4',
'00980DF6',
'00980DF9',
'00980DFC',
'00980DFE',
'00980E05',
'00980E08',
'00980E09',
'00980E0B',
'00980E0D',
'00980E0E',
'00980E0F',
'00980E10',
'00980E11',
'00980E13',
'00980E14',
'00980E17',
'00980E18',
'00980E19',
'00980E1A',
'00980E1B',
'00980E1C',
'00980E1D',
'00980E1F',
'00980E22',
'00980E23',
'00980E24',
'00980E26',
'00980E27',
'00980E28',
'00980E29',
'00980E2A',
'B120045D',
'B120045E',
'B120045F',
'B1200461',
'B1200462',
'B1200463',
'B1200464',
'B1200465',
'B1200466',
'B12004B6',
'B12004B7',
'B12005E8',
'B12005E9',
'B12005EB',
'B12005ED',
'B1200617',
'B1200631',
'B1200632',
]

api_missing_gbs = sort(list(set(top_20_gbs) - set(all_api_gbs)))

no_genset_gbs = sort(list(set(all_api_gbs) - set(genset_gbs)))
#!!! this is from E:\_UNHCR\CODE\DATA\gaps\eyedro_data_gaps.xlsx -- Azure gb tables with data gaps, that are not gensets
gbs_not_genset = ['gb_00980789', 'gb_0098082b', 'gb_00980858', 'gb_00980885', 'gb_0098088c', 'gb_0098088d', 'gb_00980890', 'gb_00980892', 'gb_00980898', 'gb_0098089a', 'gb_0098089c', 'gb_0098089e', 'gb_0098089f', 'gb_009808b0', 'gb_009808b1', 'gb_009808b6', 'gb_009808b9', 'gb_009808bb', 'gb_009808be', 'gb_009808bf', 'gb_009808f1', 'gb_0098090a', 'gb_0098090c', 'gb_00980912', 'gb_00980929', 'gb_00980958', 'gb_009809e9', 'gb_009809ea', 'gb_00980a21', 'gb_00980a2c', 'gb_00980a3e', 'gb_00980a4f', 'gb_00980a74', 'gb_00980aa1', 'gb_00980af4', 'gb_00980b2a', 'gb_00980b6e', 'gb_00980b81', 'gb_00980b89', 'gb_00980da0', 'gb_00980da2', 'gb_00980dfe', 'gb_00980e0d', '00980af5', '00980b11', '00980b35', '00980b6e', '00980da0', '00980da6', '00980db4', '00980dc4', '00980dc6', '00980dd7', '00980ddd', '00980df4', '00980dfe', '00980e0d', '00980e22']


all_gbs = sort(list(set(all_api_gbs).union(set(top_20_gbs))))


bad_gbs = [
'00980B1D',
'009809EB',
'009809EC',
'00980B26',
'00980B2E',
'00980754',
'009808C0',
'00980787',
'00980A03',
'009808C9',
'00980B67',
'00980A08',
'009808DF',
'00980A0E',
'009807DD',
'00980B82',
'009808E7',
'00980818',
'00980B88',
'00980824',
'009808ED',
'009808EE',
'00980B8D',
'00980B8E',
'00980829',
'00980B92',
'00980B93',
'009808F9',
'009808FB',
'00980837',
'0098083D',
'00980A4B',
'00980A54',
'00980924',
'0098084C',
'00980A59',
'00980851',
'00980A6C',
'00980852',
'00980853',
'00980854',
'00980936',
'00980A9A',
'00980953',
'0098085D',
'0098095E',
'00980AAF',
'00980AB2',
'0098086A',
'0098086C',
'0098097D',
'0098086D',
'00980ABA',
'00980ACA',
'0098098E',
'009809AB',
'B12004B6',
'B12005E9',
'009809B7',
'B12005EB',
'B12005ED',
'00980884',
'B1200617',
'00980886',
'00980B12',
'00980B16',
'0098088F',
'009809E0',
'009809E2',
]

engines = db.set_db_engines()

def log_gb_errors(errs):
    ###logging.error(f"ERRORS: {errs}")
    if len(errs) > 1 and any('DeviceSerial invalid' in str(item) for item in errs[1]):
        return errs[1]
    elif len(errs) > 2 and any('Invalid DateStartSecUtc' in str(item) for item in errs[2]):
        ####logging.info('Invalid DateStartSecUtc NORMAL END OF DATA')
        return errs[2]
    elif any('API' in str(item) for item in errs[0]):
        return errs[0]

    return errs

def get_previous_midnight_epoch(epoch: int = None) -> int:
    if epoch:
        dt = datetime.fromtimestamp(epoch, timezone.utc)  # Convert epoch to UTC datetime
    else:
        dt = datetime.now(timezone.utc)  # Get current UTC time

    # Set the time to midnight (00:00:00) UTC
    midnight_utc = dt.replace(hour=0, minute=0, second=0, microsecond=0)  # Ensure midnight is in UTC

    return int(midnight_utc.timestamp())


def meter_response_empty(serial, epoch=None):
    try:
        EYEDRO_KEY_GET_DATA_EMPTY = "URcLQ4MNDKgCPOacW8PB4jbTxBdEXvk3sajrD7SU"  # empty API
        if epoch:
            meter_url = "https://api.eyedro.com/Unhcr/DeviceData?DeviceSerial=" + str(serial) + "&DateStartSecUtc=" + str(epoch) + f"&DateNumSteps=1440&UserKey={EYEDRO_KEY_GET_DATA_EMPTY}"
        else:
            meter_url = "https://api.eyedro.com/Unhcr/DeviceData?DeviceSerial=" + str(serial) + f"&DateNumSteps=1440&UserKey={EYEDRO_KEY_GET_DATA_EMPTY}"
        response = requests.get(meter_url, timeout=600)
        response.raise_for_status()  # Raise an exception for HTTP errors
        return json.loads(response.text)
    except requests.exceptions.HTTPError as http_err:
        logging.error(f"meter_response:ZZZ {serial} {epoch} HTTP error occurred ERROR: {http_err}")
        return None
    except requests.exceptions.ConnectionError as conn_err:
        logging.error(f"meter_response: ZZZ {serial} {epoch} Connection error occurred ERROR: {conn_err}")
        return None
    except requests.exceptions.Timeout as timeout_err:
        logging.error(f"meter_response: ZZZ {serial} {epoch} Timeout error occurred ERROR: {timeout_err}")
        return None
    except json.JSONDecodeError as json_err:
        logging.error(f"meter_response: ZZZ {serial} {epoch} JSON decoding error occurred ERROR: {json_err}")
        return None
    except Exception as e:
        logging.error(f"meter_response: ZZZ {serial} {epoch} An unexpected error occurred: {e}")
        return None


def get_last_com_epoch(s_num):
    # Get UTC midnight for today
    midnight_utc = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    # Convert to epoch (seconds since Unix epoch)
    midnight_epoch = int(midnight_utc.timestamp())
    data = meter_response_empty(s_num)
    if data is None:
        ###logging.error(f"ZZZ {s_num} ERROR: NO DATA")
        return midnight_epoch, None
    if data['Errors']:
        err = log_gb_errors(data['Errors'])
        ##logging.error(f"ZZZ {s_num}  {midnight_epoch} ERRORS: {err}")
        return midnight_epoch, data
    last_comm = data['LastCommSecUtc']
    return get_previous_midnight_epoch(last_comm), data

# #!!!!!!!!!!!!
# for s_num in bad_gbs:
#     epoch, data = get_last_com_epoch(s_num)
#     if len(data['Errors']) != 0:
#         print(f"ERROR: {s_num}  {data['Errors']}")
#     else:
#         print(f"OK {s_num}   {data['LastCommSecUtc']}")
#     print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
#     time.sleep(1)

# pass
# #!!!!!!!!

#!!!!!!!!!!!!!!!!!! CHECK FOR HYPERGAPS !!!!!!!!!!!!!!!!!!!
def hyper_gb_gaps(eng):
    """
    Collects all hypergaps in the "eyedro" schema and writes the results to a CSV file.
    
    The function first gets all hypertables in the "eyedro" schema, then runs a query on each one
    to detect gaps in the data. The results are collected in a list and then written to a CSV file.
    
    The query is a window query that calculates the difference between the current epoch and the previous
    epoch for each row in the table. If the difference is not equal to 60 (i.e., there is a gap), the row
    is included in the results.
    
    The results are written to a CSV file in the "data/gaps" directory. The file is named "eyedro_data_gaps.csv".
    """
    conn = eng # Get raw psycopg2 connection
    with conn.cursor() as cursor:
        # Step 1: Get all hypertables in the "eyedro" schema
        cursor.execute("""
            SELECT table_name FROM information_schema.tables 
            WHERE table_schema = 'eyedro' 
            AND table_type = 'BASE TABLE'
            order by table_name;
        """)
        hypertables = [row[0] for row in cursor.fetchall()]

        # Step 2: Run the query on each hypertable and collect results
        all_results = []
        x  = 0
        for table in hypertables:
            x += 1
            # if x < 622: #!!!! for broken run, appends results
            #     continue
            query = f"""
            WITH ordered_epochs AS (
                SELECT epoch_secs, 
                    LAG(epoch_secs) OVER (ORDER BY epoch_secs) AS prev_epoch
                FROM eyedro.{table}
            )
            SELECT '{table}' AS hypertable, epoch_secs, prev_epoch, (epoch_secs - prev_epoch) AS diff_seconds
            FROM ordered_epochs
            WHERE (epoch_secs - prev_epoch) != 60;
            """

            cursor.execute(query)
            rows = cursor.fetchall()
            print(f"Processing hypertable {table} {x} of {len(hypertables)}  gaps: {len(rows)}")
            for row in rows:
                all_results.append(row)
        # Close connection
        cursor.close()
        conn.close()
    return all_results

# # #!!!!! get all the gb data gaps
# src = 'local'
# if src == 'local':
#     eng = engines[1].raw_connection()
# else:
#     eng = engines[0].raw_connection()
# res = hyper_gb_gaps(eng)
# #Step 3: Convert results to DataFrame
# df = pd.DataFrame(res, columns=["hypertable", "epoch_secs", "prev_epoch", "diff_seconds"])
# df['src'] = src

# #Save results to CSV
# df.to_csv(r"E:\_UNHCR\CODE\DATA\gaps\eyedro_data_gaps.csv", index=False)

# #Print summary
# print(df.head())
# print(f"✅ Data gaps found in {len(df)} rows. Results saved to 'eyedro_data_gaps.csv'.")
# pass

#!!!!!!!!!!!!!!!!!! CHECK FOR HYPERGAPS !!!!!!!!!!!!!!!!!!!

def sanitize_filename(filename, replacement="_"):
    # Define illegal characters for Windows filenames
    illegal_chars = r'[<>:"/\\|?*\x00-\x1F]'  # Includes control characters (0-31)
    
    # Replace illegal characters with the specified replacement (default: "_")
    sanitized = re.sub(illegal_chars, replacement, filename)
    
    # Trim trailing dots and spaces (not allowed in Windows filenames)
    return sanitized.strip(" .")

def phase_imbalance(engine, gbs):
    
    
    # wb = Workbook(write_only=True)
    # ws = wb.create_sheet()

    # rows = [
    #     ('Number', 'Batch 1', 'Batch 2'),
    #     (2, 10, 30),
    #     (3, 40, 60),
    #     (4, 50, 70),
    #     (5, 20, 10),
    #     (6, 10, 40),
    #     (7, 50, 30),
    # ]


    # for row in rows:
    #     ws.append(row)


    # chart1 = BarChart()
    # chart1.type = "col"
    # chart1.style = 10
    # chart1.title = "Bar Chart"
    # chart1.y_axis.title = 'Test number'
    # chart1.x_axis.title = 'Sample length (mm)'

    # data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
    # cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    # chart1.add_data(data, titles_from_data=True)
    # chart1.set_categories(cats)
    # chart1.shape = 4
    # ws.add_chart(chart1, "A10")

    # from copy import deepcopy

    # chart2 = deepcopy(chart1)
    # chart2.style = 11
    # chart2.type = "bar"
    # chart2.title = "Horizontal Bar Chart"

    # ws.add_chart(chart2, "G10")


    # chart3 = deepcopy(chart1)
    # chart3.type = "col"
    # chart3.style = 12
    # chart3.grouping = "stacked"
    # chart3.overlap = 100
    # chart3.title = 'Stacked Chart'

    # ws.add_chart(chart3, "A27")


    # chart4 = deepcopy(chart1)
    # chart4.type = "bar"
    # chart4.style = 13
    # chart4.grouping = "percentStacked"
    # chart4.overlap = 100
    # chart4.title = 'Percent Stacked Chart'

    # ws.add_chart(chart4, "G27")

    # wb.save("bar.xlsx")
    
    
    
    
    gb_office_path = r'E:\_UNHCR\CODE\DATA\gb-office.csv'
    # Load CSV into DataFrame
    df_gb_office = pd.read_csv(gb_office_path, dtype={"serial": str}, encoding="ISO-8859-1")  # Ensure serial is treated as a string

    out_path = r'E:\UNHCR\OneDrive - UNHCR\Energy Team\Concept development\AZURE DATA\phase_imbalance\gensets'
    dt = datetime.now().date().isoformat()

    try:
        conn = engine.raw_connection()  # Get raw psycopg2 connection
        with conn.cursor() as cur:
            for serial in gbs:
                try:
                    # last 180 days, more than 1 kwh in an hour
                    sql = f"""
                    WITH wh AS (
                        SELECT 
                            DATE_TRUNC('hour', ts) AS ts_hour,
                            DATE(ts) AS dt,
                            EXTRACT(HOUR FROM ts) AS hr_utc,
                            ROUND(SUM(wh_p1)::numeric, 3) AS wh_p1, 
                            ROUND(SUM(wh_p2)::numeric, 3) AS wh_p2, 
                            ROUND(SUM(wh_p3)::numeric, 3) AS wh_p3
                        FROM eyedro.gb_{serial}
                        where ts > now() - interval '180 days'
                        GROUP BY 1, 2, 3
                    ),
                    hourly_avg AS (
                        SELECT 
                            hr_utc,
                            ROUND(AVG(wh_p1)::numeric, 3) AS avg_wh_p1,
                            ROUND(AVG(wh_p2)::numeric, 3) AS avg_wh_p2,
                            ROUND(AVG(wh_p3)::numeric, 3) AS avg_wh_p3,
                            ROUND(
                            ((GREATEST(AVG(wh_p1)::numeric, AVG(wh_p2)::numeric, AVG(wh_p3)::numeric) - 
                            LEAST(AVG(wh_p1)::numeric, AVG(wh_p2)::numeric, AVG(wh_p3)::numeric)) / 
                            NULLIF(GREATEST(AVG(wh_p1)::numeric, AVG(wh_p2)::numeric, AVG(wh_p3)::numeric), 0)), 3
                        ) AS phase_imbalance
                        FROM wh
                        GROUP BY hr_utc
                    )
                    select * from hourly_avg 
                    where avg_wh_p1 + avg_wh_p2 + avg_wh_p3 > 1
                    order by hr_utc
                    """

                    cur.execute(sql)
                    res = cur.fetchall()

                    # Convert to DataFrame only if there is data
                    if res:
                        df_result = pd.DataFrame(res, columns=[desc[0] for desc in cur.description])
                        print (df_result)
                        # Search for a specific serial
                        matching_row = df_gb_office[df_gb_office["serial"] == serial.upper()]
                        print (matching_row)
                        if matching_row.empty:
                            fn = f'{out_path}\\{serial}_phase_imbalance_{dt}.xlsx'
                        else:
                            prefix = sanitize_filename(f'{matching_row.values[0][4]}_{matching_row.values[0][5]}_')
                            fn = f'{out_path}\\{prefix}{serial}_phase_imbalance_{dt}.xlsx'

                        df_result["phase_imbalance"] = pd.to_numeric(df_result["phase_imbalance"])
                        df_result["avg_wh_p1"] = pd.to_numeric(df_result["avg_wh_p1"])
                        df_result["avg_wh_p2"] = pd.to_numeric(df_result["avg_wh_p2"])
                        df_result["avg_wh_p3"] = pd.to_numeric(df_result["avg_wh_p3"])
                        df_result["hr_utc"] = pd.to_numeric(df_result["hr_utc"])

                        with pd.ExcelWriter(fn, engine="openpyxl") as writer:
                            df_result.to_excel(writer, sheet_name=prefix, index=False)  # ✅ Save to CSV
                        # Load workbook and sheet
                        wb = load_workbook(fn)
                        ws = wb[prefix]

                        # Apply style to "phase_imbalance" column (column E, starting from row 5)
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
                            for cell in row:
                                cell.number_format = "0.0%"

                        # Create Column Chart
                        chart = BarChart()
                        chart.type = "col"
                        chart.style = 10
                        chart.title = "Phase Imbalance"

                        # **Ensure axis titles are positioned correctly**
                        chart.y_axis.title = "Phase Imbalance (%)"
                        chart.x_axis.title = "Hour (UTC)"

                        # **Ensure axis labels (tick marks) are visible**
                        #chart.y_axis.majorGridlines = None  # Remove unnecessary gridlines
                        #chart.y_axis.minorGridlines = None  # Ensure no minor gridlines
                        chart.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart.x_axis.majorTickMark = "out"  # Show tick marks outward

                        # **Ensure X-axis title is positioned properly**
                        chart.x_axis.title.txPr = None  # Reset title formatting (places it outside)
                        chart.y_axis.title.txPr = None  

                        # **Ensure data is treated as a single series**
                        data = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
                        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

                        chart.add_data(data, titles_from_data=False)  # Single series, no multi-color
                        chart.set_categories(categories)  # Ensure hours appear on X-axis

                        # **Increase chart size to fit labels properly**
                        chart.width = 20  # Increase width (default is ~15)
                        chart.height = 10  # Increase height (default is ~7.5)

                        chart.layout = Layout(
                            manualLayout=ManualLayout(
                                x=0.999,  # Shift plot area slightly right
                                #y=1.05,  # Shift plot area slightly up
                                h=0.95,  # Reduce height of the plot area
                                w=0.95   # Reduce width of the plot area
                            )
                        )

                        # Enable auto-scaling (default behavior in openpyxl)
                        chart.y_axis.scaling.min = None  # Allow automatic minimum value
                        chart.y_axis.scaling.max = None  # Allow automatic maximum value
                        # !!!!!values = [row[0] for row in ws.iter_rows(min_col=5, min_row=2, max_row=ws.max_row, max_col=5, values_only=True)]
                        # max_value = max(values) if values else 1  # Avoids errors if empty
                        # chart.y_axis.scaling.max = max_value * 1.1  # Scale up by 10%

                        # Ensure Y-axis major tick marks are visible (auto-scaled)
                        chart.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart.y_axis.minorTickMark = "none"  # Hide minor tick marks

                        # Force display of axis values (ensures ticks are labeled)
                        chart.y_axis.majorGridlines = None  # Keep clean gridlines
                        chart.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks

                        # Extract Y-axis values correctly
                        # y_values = [row[0] for row in ws.iter_rows(min_col=5, max_col=5, min_row=2, max_row=ws.max_row, values_only=True) if row[0] is not None]
                        # # Prevent errors when max() is called on empty data
                        # max_y = max(y_values) if y_values else 1  
                        # min_y = min(y_values) if y_values else 0  # Ensure we get the lowest value  
                        # Apply auto-scaling while ensuring labels are displayed
                        # chart.y_axis.scaling.min = min_y  # Start from lowest value
                        # chart.y_axis.scaling.max = max_y * 1.1  # Scale up by 10%
                        # chart.y_axis.majorUnit = round((max_y - min_y) / 10, 2)
                        chart.y_axis.scaling.min = 0  # Start from lowest value
                        chart.y_axis.scaling.max = 1.05
                        chart.y_axis.majorUnit = round(1 / 10, 2)  # Ensure reasonable spacing

                        # Ensure tick labels are displayed properly
                        chart.y_axis.majorTickMark = "out"
                        chart.y_axis.minorTickMark = "none"
                        chart.y_axis.tickLblPos = "nextTo"  # Position labels properly

                        # Force Y-axis labels to display (Excel sometimes hides them)
                        chart.y_axis.delete = False  # Ensure axis is visible

                        chart.x_axis.majorTickMark = "out"  # Show tick marks outward
                        chart.x_axis.tickLblPos = "nextTo"  # Ensure labels are placed properly
                        chart.x_axis.delete = False  # Make sure X-axis is visible
                        chart.x_axis.majorUnit = 1  # Ensure every hour is displayed
                        chart.x_axis.minorUnit = 1  # Prevent minor ticks from interfering
                        # Get the first data series (bars)
                        series = chart.series[0]  

                        # Set a solid fill color (change "0000FF" to your desired color)
                        series.graphicalProperties.solidFill = "0000FF"  # Blue

                        # **Improve styling**
                        chart.gapWidth = 30  # Adjust spacing between bars
                        chart.overlap = 0  # Ensure no overlap
                        chart.legend = None  # Remove legend (single series)

                        # Create Line Chart
                        chart1 = LineChart()
                        chart1.type = "line"
                        chart1.style = 10
                        chart1.title = "Wh by Phase"

                        # **Ensure axis titles are positioned correctly**
                        chart1.y_axis.title = "Phase (Wh)"
                        chart1.x_axis.title = "Hour (UTC)"

                        # **Ensure axis labels (tick marks) are visible**
                        chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart1.x_axis.majorTickMark = "out"  # Show tick marks outward

                        # **Ensure X-axis title is positioned properly**
                        chart1.x_axis.title.txPr = None  # Reset title formatting (places it outside)
                        chart1.y_axis.title.txPr = None  

                        # **Ensure data is treated as a single series**
                        data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=ws.max_row)

                        ####categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

                        chart1.add_data(data, titles_from_data=True)  # Single series, no multi-color
                        ####chart1.set_categories(categories)  # Ensure hours appear on X-axis

                        # **Increase chart size to fit labels properly**
                        chart1.width = 20  # Increase width (default is ~15)
                        chart1.height = 10  # Increase height (default is ~7.5)

                        chart1.layout = Layout(
                            manualLayout=ManualLayout(
                                x=0.999,  # Shift plot area slightly right
                                #y=1.05,  # Shift plot area slightly up
                                h=0.95,  # Reduce height of the plot area
                                w=0.95   # Reduce width of the plot area
                            )
                        )

                        # Enable auto-scaling (default behavior in openpyxl)
                        chart1.y_axis.scaling.min = None  # Allow automatic minimum value
                        chart1.y_axis.scaling.max = None  # Allow automatic maximum value
                        # !!!!!!! values = [row[0] for row in ws.iter_rows(min_col=5, min_row=2, max_row=ws.max_row, max_col=5, values_only=True)]
                        # max_value = max(values) if values else 1  # Avoids errors if empty
                        # chart1.y_axis.scaling.max = max_value * 1.1  # Scale up by 10%

                        # Ensure Y-axis major tick marks are visible (auto-scaled)
                        chart1.y_axis.majorTickMark = "out"  # Show tick marks outward
                        chart1.y_axis.minorTickMark = "none"  # Hide minor tick marks

                        # Force display of axis values (ensures ticks are labeled)
                        chart1.y_axis.majorGridlines = None  # Keep clean gridlines
                        chart1.y_axis.tickLblPos = "nextTo"  # Ensure labels appear near ticks

                        # Extract Y-axis values correctly
                        # Extract all values from columns 2, 3, and 4
                        y_values = [value for row in ws.iter_rows(min_col=2, max_col=4, min_row=2, max_row=ws.max_row, values_only=True) 
                            for value in row if value is not None]
                        # Prevent errors when max() is called on empty data
                        max_y = max(y_values) if y_values else 1
                        min_y = min(y_values) if y_values else 0  # Ensure we get the lowest value
                        #Apply auto-scaling while ensuring labels are displayed
                        chart1.y_axis.scaling.min = min_y  # Start from lowest value
                        chart1.y_axis.scaling.max = max_y * 1.1  # Scale up by 10%
                        chart1.y_axis.majorUnit = round((max_y - min_y) / 10, 2)
                        # chart1.y_axis.scaling.min = 0  # Start from lowest value
                        # chart1.y_axis.scaling.max = 1.05
                        # chart1.y_axis.majorUnit = round(1 / 10, 2)  # Ensure reasonable spacing

                        # Ensure tick labels are displayed properly
                        chart1.y_axis.majorTickMark = "out"
                        chart1.y_axis.minorTickMark = "none"
                        chart1.y_axis.tickLblPos = "nextTo"  # Position labels properly

                        # Force Y-axis labels to display (Excel sometimes hides them)
                        chart1.y_axis.delete = False  # Ensure axis is visible

                        chart1.x_axis.majorTickMark = "out"  # Show tick marks outward
                        chart1.x_axis.tickLblPos = "nextTo"  # Ensure labels are placed properly
                        chart1.x_axis.delete = False  # Make sure X-axis is visible
                        chart1.x_axis.majorUnit = 1  # Ensure every hour is displayed
                        chart1.x_axis.minorUnit = 1  # Prevent minor ticks from interfering

                        chart_sheet_name = "Charts"
                        ws_chart = wb.create_sheet(chart_sheet_name)
                        # **Add chart to Excel**
                        ws_chart.add_chart(chart, "C2")
                        ws_chart.add_chart(chart1, "C23")

                        wb.save(fn)

                        logging.info(f"Saved CSV for {serial}: {fn}")
                    else:
                        logging.info(f"No data for {serial}, skipping CSV creation.")

                except DatabaseError as e:
                    logging.error(f"Database error for {serial}: {e}")
                except Exception as e:
                    logging.error(f"Unexpected error for {serial}: {e}")

    except DatabaseError as e:
        logging.error(f"Database connection error: {e}")
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
    finally:
        if conn:
            conn.close()  # ✅ Always close the connection

# !!!!!!! phase_imbalance
# phase_imbalance(engines[1], genset_gbs)
# pass


#!!!!!!!!!!


# #!!!!! we have a gaps table
# gaps_csv = r"E:\_UNHCR\CODE\DATA\gaps\gaps.csv"
# #!!!! get data from table and save to csv
# # Set database engine
# eng = db.set_local_defaultdb_engine()

# # Define the table where results will be stored
# gb_gaps_table = "eyedro.gb_gaps"  # Change this to your actual table name

# #!!! get data from gaps table and save to csv
# sql = f"SELECT hypertable_name, epoch_secs, prev_epoch, diff_seconds FROM {gb_gaps_table};"
# res, err = db.sql_execute(sql, eng)

# # Check if we got results
# if res:
#     # Process data into a DataFrame
#     data = []
#     for row in res:
#         from_timestamp = datetime.fromtimestamp(row[2], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
#         to_timestamp = datetime.fromtimestamp(row[1], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
#         days = row[3] / (60 * 60 * 24)  # Convert seconds to days
        
#         data.append([row[0][3:], row[1], row[2], row[3], from_timestamp, to_timestamp, round(days, 2)])

#     # Create DataFrame
#     df = pd.DataFrame(data, columns=["gb_serial", "epoch_secs", "prev_epoch", "diff_seconds", "from_timestamp", "to_timestamp", "days"])
    
#     # Save to CSV
#     df.to_csv(gaps_csv, index=False)

# pass
# #!!!!! we have a gaps table

# #!!!!! save data to gaps table
# # Ensure the table exists
# db.sql_execute(f"""
#     CREATE TABLE IF NOT EXISTS {gb_gaps_table} (
#         hypertable_name TEXT,
#         epoch_secs BIGINT,
#         prev_epoch BIGINT,
#         diff_seconds INT,
#         days varchar(10),
#         UNIQUE (epoch_secs, prev_epoch, diff_seconds)
#     );
# """, eng)


# # Fetch hypertable names
# sql = """SELECT hypertable_name FROM timescaledb_information.hypertables WHERE hypertable_name LIKE 'gb_%' ORDER BY hypertable_name;"""
# res, err = db.sql_execute(sql, eng)

# # Loop through hypertables
# for tn in res:
#     print('')
#     res, err = db.sql_execute(f"""
#         WITH ordered_epochs AS (
#             SELECT epoch_secs, 
#                 LAG(epoch_secs) OVER (ORDER BY epoch_secs) AS prev_epoch
#             FROM eyedro.{tn[0]}
#         )
#         SELECT epoch_secs, prev_epoch, (epoch_secs - prev_epoch) AS diff_seconds
#         FROM ordered_epochs
#         WHERE (epoch_secs - prev_epoch) != 60;
#     """, eng)

#     if res:
#         for row in res:
#             print(f"{tn[0]} from: {datetime.fromtimestamp(row[1], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} "
#                   f"to: {datetime.fromtimestamp(row[0], tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} "
#                   f"days: {row[2] / (60 * 60 * 24):.2f}")

#             # Insert into database, ensuring uniqueness
#             db.sql_execute(f"""
#                 INSERT INTO {gb_gaps_table} (hypertable_name, epoch_secs, prev_epoch, diff_seconds, days)
#                 VALUES ('{tn[0]}', {row[0]}, {row[1]}, {row[2]}, '{row[2] / (60 * 60 * 24):.2f}')
#                 ON CONFLICT (epoch_secs, prev_epoch, diff_seconds) DO NOTHING;
#             """, eng)
#     else:
#         print(f"{tn[0]} !!!!!!!!!!!!")
# pass
# #!!!!! save data to gaps table



def upsert_zip(table):
    zip_path = r"E:\UNHCR\OneDrive - UNHCR\Energy Team\Concept development\AZURE DATA\BACKUPS\eyedro\local" + "\\"
    engine = db.set_local_defaultdb_engine()
    conn = engine.raw_connection()  # Get raw psycopg2 connection
    cur = conn.cursor()
    table_name = f"gb_{serial}".lower()
    #!!!! cur.execute(f"""select count(*) from eyedro.{table_name}""")
    #!!!!! cur.execute(f"""SET timescaledb.max_tuples_decompressed_per_dml_transaction = 200000;""")
    cur.execute(f'delete from eyedro.gb_{serial} where a_p1 is null and a_p2 is null and a_p3 is null')
    print(f"delete from eyedro.gb_{serial} cnt: {cur.rowcount}")  #cur.fetchone()[0]
    conn.commit()
    #!!!!!
    cur.close()
    return

    if res == 0:
        # ✅ Create a single temporary table BEFORE the loop
        cur.execute(f"CREATE TEMP TABLE if not exists temp_{table_name} (LIKE eyedro.gb_00980639 INCLUDING ALL);")
        cur.execute(f"TRUNCATE temp_{table_name};")
        conn.commit()
        zip_filename = f"{zip_path}{table_name}.csv.zip"
        # Open ZIP file and extract CSV in memory
        try:
            with zipfile.ZipFile(zip_filename, "r") as z:
                with z.open(f'{table_name}.csv') as f:
                    cur.copy_expert(f"COPY temp_{table_name} FROM STDIN WITH CSV HEADER", io.TextIOWrapper(f, encoding="utf-8"))

                    # Insert only new rows, ignore duplicates
                    cur.execute(f"""
                        INSERT INTO eyedro.{table_name}
                        SELECT * FROM temp_{table_name}
                        ON CONFLICT (ts, epoch_secs) DO NOTHING;
                    """)
                    inserted_rows = cur.rowcount
                    cur.execute(f"drop table if exists temp_{table_name};")
                    conn.commit()
                    print(f"'{serial}',{inserted_rows}")
        except Exception as e:
            if conn:
                conn.rollback()  # Rollback on failure
            logging.error(f"Database error during table creation: {e}")
    else:
        pass #print(f'XXXXX{serial}       {res}')
    cur.close()


# #!!!!!! ZIP ess_zip(chunk):
#     return [upsert_zip(item) for item in chunk]  # Apply function to each item

# num_parts = 5
# chunks = [list(chunk) for chunk in np.array_split(all_api_gbs, num_parts)]  # Ensure list format

# with ThreadPoolExecutor(max_workers=num_parts) as executor:
#     results = list(executor.map(process_zip, chunks))

# # Flatten the nested list of results
# final_output = [item for sublist in results for item in sublist]

# pass
# #!!!!!!


def get_ttl_kwh(table_name, start_yr=2020, end_yr=2025):
    try:
        engine = db.set_local_defaultdb_engine()
        conn = engine.raw_connection()  # Get raw psycopg2 connection
        cur = conn.cursor()
        sn = (table_name[:6] + "-" + table_name[6:])[3:].upper()
        sql = f"""WITH stats AS (
    SELECT 
        EXTRACT(YEAR FROM ts) AS yr,
        AVG(ABS(wh_p1) + ABS(wh_p2) + ABS(wh_p3)) AS mean_wh,
        STDDEV(ABS(wh_p1) + ABS(wh_p2) + ABS(wh_p3)) AS std_wh
    FROM eyedro.{table_name}
    WHERE a_p1 IS NOT NULL AND a_p2 IS NOT NULL AND a_p3 IS NOT NULL
    AND a_p1 NOT IN ('NaN'::FLOAT8) AND a_p2 NOT IN ('NaN'::FLOAT8) AND a_p3 NOT IN ('NaN'::FLOAT8)
    GROUP BY yr
    )
    SELECT 
        ys.yr,
        '{sn}' AS gb_serial,
            FLOOR(COALESCE(SUM(
                CASE 
                    WHEN (ABS(wh_p1) + ABS(wh_p2) + ABS(wh_p3)) NOT IN ('NaN'::FLOAT8)  
                    AND (ABS(wh_p1) + ABS(wh_p2) + ABS(wh_p3)) > s.mean_wh + (3 * s.std_wh) 
                    OR (ABS(wh_p1) + ABS(wh_p2) + ABS(wh_p3)) < s.mean_wh - (3 * s.std_wh)
                    THEN COALESCE(s.mean_wh, 0)  -- Replace outliers with mean, ensuring no NULL
                    ELSE COALESCE(ABS(wh_p1) + ABS(wh_p2) + ABS(wh_p3), 0)  -- Ensure no NULL values and take absolute values
                END
            ) / 1000, 0.0)) AS ttl
        FROM generate_series(2020, 2025) ys(yr)
        LEFT JOIN eyedro.{table_name} g
            ON ys.yr = EXTRACT(YEAR FROM g.ts)
        LEFT JOIN stats s
            ON ys.yr = s.yr
        AND g.a_p1 IS NOT NULL AND g.a_p2 IS NOT NULL AND g.a_p3 IS NOT NULL
        AND g.a_p1 NOT IN ('NaN'::FLOAT8) AND g.a_p2 NOT IN ('NaN'::FLOAT8) AND g.a_p3 NOT IN ('NaN'::FLOAT8)
        GROUP BY ys.yr
        ORDER BY ys.yr;
                """
        cur.execute(sql)
        res = cur.fetchall()
        conn.commit()
        cur.close()
    except Exception as e:
        return None
    return res


def get_ttl_kwh_concur(table_name):
    res = get_ttl_kwh(table_name)
    if res:
        print(f" device_serial: {table_name}")

        vals = [res[0][1]] + [val[2] for val in res]
        return vals
    else:
        return None


# #!!!!!! Get cleaned ttl KWH for every gb in local DB
# def process_kwh(chunk):
#     return [get_ttl_kwh_concur(item[0]) for item in chunk]  # Apply function to each item

# sns, err = db.sql_execute("""SELECT hypertable_name FROM timescaledb_information.hypertables 
# where hypertable_schema = 'eyedro' and hypertable_name like 'gb_%';""", db.set_local_defaultdb_engine())

# num_parts = 20
# chunks = [list(chunk) for chunk in np.array_split(sns, num_parts)]  # Ensure list format

# with ThreadPoolExecutor(max_workers=num_parts) as executor:
#     results = list(executor.map(process_kwh, chunks))

# # Flatten the nested list of results
# final_list = [item for sublist in results for item in sublist]

# df = pd.DataFrame(final_list, columns=["gb_serial", "kwh_2020", "kwh_2021", "kwh_2022", "kwh_2023", "kwh_2024", "kwh_2025"])
# df.to_csv(r"E:\_UNHCR\CODE\DATA\gaps\all_gbs_kwh_2024.csv", index=False)
# pass
# #!!!!!!

def meter_response_null(serial, epoch=None):
    try:
        EYEDRO_KEY_GET_DATA_NULL = "UNkQm84Wyt61agj9jZShDFtS6fN8k7jAkxER2cSU"  # null API
        print("EyeDro Endpoint and Key Set", serial, epoch)
        if epoch:
            meter_url = "https://api.eyedro.com/Unhcr/DeviceData?DeviceSerial=" + str(serial) + "&DateStartSecUtc=" + str(epoch) + f"&DateNumSteps=1440&UserKey={EYEDRO_KEY_GET_DATA_NULL}"
        else:
            meter_url = "https://api.eyedro.com/Unhcr/DeviceData?DeviceSerial=" + str(serial) + f"&DateNumSteps=1440&UserKey={EYEDRO_KEY_GET_DATA_NULL}"
        print(meter_url)
        #https://api.eyedro.com/Unhcr/DeviceData?DeviceSerial=00980AA1&DateNumSteps=1440&UserKey=UNkQm84Wyt61agj9jZShDFtS6fN8k7jAkxER2cSU
        #https://api.eyedro.com/Unhcr/DeviceData?DeviceSerial=B1200631&DateStartSecUtc=1725147900&DateNumSteps=1440&UserKey=URcLQ4MNDKgCPOacW8PB4jbTxBdEXvk3sajrD7SU UNkQm84Wyt61agj9jZShDFtS6fN8k7jAkxER2cSU
        #https://api.eyedro.com/Unhcr/DeviceData?DeviceSerial=00980640&DateStartSecUtc=1725147900&DateNumSteps=1440&UserKey=URcLQ4MNDKgCPOacW8PB4jbTxBdEXvk3sajrD7SU
        response = requests.get(meter_url, timeout=600)
        response.raise_for_status()  # Raise an exception for HTTP errors
        return json.loads(response.text)
    except requests.exceptions.HTTPError as http_err:
        logging.error(f"meter_response:ZZZ {serial} {epoch} HTTP error occurred ERROR: {http_err}")
        return None
    except requests.exceptions.ConnectionError as conn_err:
        logging.error(f"meter_response: ZZZ {serial} {epoch} Connection error occurred ERROR: {conn_err}")
        return None
    except requests.exceptions.Timeout as timeout_err:
        logging.error(f"meter_response: ZZZ {serial} {epoch} Timeout error occurred ERROR: {timeout_err}")
        return None
    except json.JSONDecodeError as json_err:
        logging.error(f"meter_response: ZZZ {serial} {epoch} JSON decoding error occurred ERROR: {json_err}")
        return None
    except Exception as e:
        logging.error(f"meter_response: ZZZ {serial} {epoch} An unexpected error occurred: {e}")
        return None


def get_db_epoch(serial, engine, max):
    # Get midnight timestamp (00:00:00) for today in UTC
    epoch_midnight_today = get_previous_midnight_epoch()
    epoch, err = db.get_gb_epoch(serial, engine, max=max)
    if err:
        logging.error(f'get_db_epoch {serial} ERROR:{err}')
    if epoch:
        dt = datetime.fromtimestamp(epoch, tz=timezone.utc) + timedelta(days=1) # Convert epoch to datetime
        midnight = dt.replace(hour=0, minute=0, second=0, microsecond=0)  # Set time to midnight
        epoch = int(midnight.timestamp()) 
    else:
        epoch = epoch_midnight_today
    return epoch


def map_gb(device_data):
    # Rename "A" to "a_p1", "a_p2", "a_p3"
    parameter_mapping = {"A": "a_p", "V": "v_p", "PF": "pf_p", "Wh": "wh_p"}
    records = {}

    for param, lists in device_data.items():
        param_prefix = parameter_mapping.get(param, param)  # Default to param name if not in mapping
        for i, dataset in enumerate(lists):
            for timestamp, value in dataset:
                ts_str = datetime.fromtimestamp(timestamp, tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
                if ts_str not in records:
                    records[ts_str] = {"ts": ts_str, "epoch_secs": timestamp}  # Initialize timestamp row
                records[ts_str][f"{param_prefix}{i + 1}"] = value
    # Convert records dict to DataFrame
    df = pd.DataFrame.from_dict(records, orient="index")
    # !!! no need for this, but might use it later
    # # Ensure missing columns (V & PF) exist with NaN
    # expected_columns = ["ts"] + [f"{p}{i}" for p in ["a_p", "v_p", "pf_p", "wh_p"] for i in range(1, 4)]
    # for col in expected_columns:
    #     if col not in df:
    #         df[col] = None  # Or df[col] = 0 for default values

    # if amps are missing, drop the row Eyedro send wh when there are no amps
    # Define the required columns
    required_columns = ['a_p1', 'a_p2', 'a_p3']

    # Check if all required columns are in the DataFrame
    if all(col in df.columns for col in required_columns):
        # If all required columns exist, drop rows with NaN in those columns
        df = df.dropna(subset=required_columns)
    else:
        # If any required columns are missing, remove all rows
        df = df.iloc[0:0]  # This clears the DataFrame
        # Sort by timestamp
    return df.sort_values("ts").reset_index(drop=True)


def error_alarm():
    for _ in range(3):  # Repeat 3 times
        winsound.Beep(900, 500)  # High tone (900 Hz for 500 ms)
        winsound.Beep(600, 500)  # Low tone (600 Hz for 500 ms)


def update_gb_db(serial, df, engine, msg=''):
    ###! df = df.fillna(0)
    """
    Updates the database with new data for a given serial number using an UPSERT operation.

    This function takes a DataFrame containing device data and inserts it into a PostgreSQL table
    named "gb_<serial>". If a record with the same timestamp and epoch_secs already exists, it updates
    the existing record with new values for each parameter. The function commits the transaction after
    successfully inserting/updating the records, and logs the number of inserted and updated rows.

    Args:
        serial (str): The serial number of the device.
        df (pd.DataFrame): A DataFrame containing the device data to be inserted/updated.
        engine (sqlalchemy.engine.base.Engine): The SQLAlchemy engine connected to the PostgreSQL database.

    Returns:
        tuple: A tuple containing a list with the number of inserted and updated records, and an error object (if any).
               If an error occurs, the function returns (None, error).
    """

    columns_str = ", ".join(df.columns)
    # Use ON CONFLICT for UPSERT (Insert or Update)
    upsert_sql = f"""
    WITH insert_attempt AS (
        INSERT INTO eyedro.gb_{serial} ({columns_str})
        VALUES %s
        ON CONFLICT (ts, epoch_secs) DO UPDATE SET
            a_p1 = EXCLUDED.a_p1,
            a_p2 = EXCLUDED.a_p2,
            a_p3 = EXCLUDED.a_p3,
            v_p1 = EXCLUDED.v_p1,
            v_p2 = EXCLUDED.v_p2,
            v_p3 = EXCLUDED.v_p3,
            pf_p1 = EXCLUDED.pf_p1,
            pf_p2 = EXCLUDED.pf_p2,
            pf_p3 = EXCLUDED.pf_p3,
            wh_p1 = EXCLUDED.wh_p1,
            wh_p2 = EXCLUDED.wh_p2,
            wh_p3 = EXCLUDED.wh_p3
        RETURNING xmax = 0 AS inserted
    )
    SELECT COUNT(*) FILTER (WHERE inserted) AS inserted_count FROM insert_attempt;
    """

    conn = engine.raw_connection()  # Get raw psycopg2 connection
    try:
        with conn.cursor() as cur:
            execute_values(cur, upsert_sql, df.to_records(index=False).tolist())

            # Get the inserted count from RETURNING
            inserted_count = cur.fetchone()[0]  # Fetch single row result

            # Use rowcount to get total affected rows (inserts + updates)
            total_affected = len(df)  

            # Calculate updated count
            updated_count = total_affected - inserted_count  

            # Commit transaction
            conn.commit()

    except DatabaseError as e:
        conn.rollback()  # Rollback on failure
        logging.error(f"ZZZ {serial} update_gb_db Database error during UPSERT: {e}")
        return None, e
    except Exception as e:
        conn.rollback()  # Rollback on any other failure
        logging.error(f"ZZZ {serial} update_gb_db Unexpected error: {e}")
        return None, e
    finally:
        conn.close()
    logging.info(f"{serial} {df["ts"].min()} {df["ts"].max()} Inserted: {inserted_count}, Updated: {updated_count} ✅ {msg}")
    return [inserted_count, updated_count], None


epoch_2024 = 1704067199
epoch_2023 = 1672502399
epoch_2022 = 1640937599
epoch_2021 = 1609372799
epoch_2020 = 1577807999

def upsert_gb_data(s_num, engine, epoch_cutoff = epoch_2023, epoch_start = None, MAX_EMPTY=30, msg=''):
    no_data_cnt = 0
    ttl_cnt = 0
    err_cnt = 0
    ttl_inserted = 0
    ttl_updated = 0
    ttl_empty = 0
    ##!!! error_alarm()
    epoch, data = get_last_com_epoch(s_num)
    if epoch < epoch_cutoff or s_num == '00980639':
        logging.info(f"ZZZ {s_num} last com before cutoff: {datetime.fromtimestamp(epoch, timezone.utc).isoformat()}\n{datetime.fromtimestamp(epoch_cutoff, timezone.utc).isoformat()}" )
        return s_num, ttl_cnt, err_cnt, ttl_inserted, ttl_updated, ttl_empty
    #!!!!! epoch = get_db_epoch(s_num, engine, max=True)
    #!!!!! epoch = 1740535260
    if data and "DeviceData" in data:
        df = map_gb(data["DeviceData"])
        if not df.empty:
            cnt, err = update_gb_db(s_num, df, engine, msg)
            if err:
                logging.error(f"ZZZ {s_num}  {epoch} ERRORS: {err}")
                return s_num, ttl_cnt, err_cnt, ttl_inserted, ttl_updated, ttl_empty
            ttl_inserted += cnt[0]
            ttl_updated += cnt[1]
            ttl_cnt += 1
            print(cnt, ttl_cnt, epoch, s_num)
        ####!!!! max_epoch_db = get_db_epoch(s_num, engine, max=True)
        #!!! min_epoch_db = get_db_epoch(s_num, local_engine, max=False)
    
    logging.info(f"{s_num} last com: {datetime.fromtimestamp(epoch, timezone.utc).isoformat()} cutoff: {datetime.fromtimestamp(epoch_cutoff, timezone.utc).isoformat()}" )
    if data and "Errors" in data:
        if len(data["Errors"]) == 0:
            epoch -= 86400
        else:
            logging.error(f"last_comm TRY AGAIN ZZZ {s_num}  {epoch} ERRORS: {data['Errors']}")
    elif data is None:
        logging.error(f"data None TRY AGAIN ZZZ {s_num}  {epoch}")
    ####epoch = min_epoch_db # start where we left off
    if epoch_start:
        epoch = epoch_start
        try:
            sql = f'select min(epoch_secs), min(ts) from eyedro.gb_{s_num};'
            res, err = db.sql_execute(sql, engine)
            if res[0][0] and res[0][0] < epoch_start:
                epoch = get_previous_midnight_epoch(res[0][0])
        except: 
            pass

    while epoch >= epoch_cutoff and err_cnt < 5:  #2024-01-01
        # if ttl_cnt > 7:
        #     #!!!!if epoch <= max_epoch_db and min_epoch_db <= epoch_cutoff:
        #     #    return ttl_cnt, err_cnt, ttl_inserted, ttl_updated
        #     return ttl_cnt, err_cnt, ttl_inserted, ttl_updated

        #     epoch = min_epoch_db
        #     ttl_cnt = -200
        data = meter_response_empty(s_num, epoch)
        ttl_cnt += 1
        # these first 3 errors will retry
        if data is None:
            no_data_cnt += 1
            err_cnt += 1
            logging.error(f"{err_cnt} ZZZ {s_num} {epoch} ERROR: NO DATA: {no_data_cnt}")
            ###epoch -= 86400
            time.sleep(.5)
            continue
        if 'Errors' in data and len(data['Errors']) != 0:
            gb_err = log_gb_errors(data['Errors'])
            if 'Invalid DateStartSecUtc' in str(gb_err[1]):
                logging.info(f'Invalid DateStartSecUtc NORMAL END OF DATA  {s_num} {epoch} {ttl_cnt} {err_cnt} {ttl_inserted} {ttl_updated} {ttl_empty}')
                return s_num, ttl_cnt, err_cnt, ttl_inserted, ttl_updated, ttl_empty
            err_cnt += 1
            logging.error(f"{err_cnt} ZZZ {s_num}  {epoch} ERRORS: {gb_err}")
            if 'API Error' in gb_err[1]:
                print(f'{s_num} {epoch}  ERROR: API Error. Contact Eyedro Admin.')
                time.sleep(.5)
            if  'API busy' in gb_err[1]:
                print(f'{s_num} {epoch}  ERROR: API Error. Contact Eyedro Admin.')
                time.sleep(2)
            continue
        if 'DeviceData' not in data:
            err_cnt += 1
            no_data_cnt += 1
            logging.error(f"{err_cnt} ZZZ {s_num} {epoch} ERROR: NO DATA IN RESPONSE: {no_data_cnt}")
            time.sleep(.5)
            continue
        else:
            df = map_gb(data["DeviceData"])
            if not df.empty:
                cnt, err = update_gb_db(s_num, df, engine, msg)
                err_cnt = 0
                #!!! if gaps:
                #     columns = ['a_p1', 'a_p2', 'a_p3']
                #     existing_columns = [col for col in columns if col in df.columns] 
                #     if existing_columns:
                #         df_filtered = df[df[existing_columns].isna().any(axis=1)]  # Select rows with NaN in any of the existing columns
                #         df_filtered.loc[:, ~df_filtered.columns.isin(['ts', 'epoch_secs'])] = None

                #         cnt1, err1 = update_gb_db(s_num, df_filtered, engine, msg)
                #         print(epoch,datetime.fromtimestamp(epoch, timezone.utc).date(), s_num, err, err1)
                if cnt:
                    ttl_inserted += cnt[0]
                    ttl_updated += cnt[1]
            else:
                ttl_empty += 1
                if ttl_empty > MAX_EMPTY:
                    return s_num, ttl_cnt, err_cnt, ttl_inserted, ttl_updated, ttl_empty
                logging.info(f"ZZZ {s_num} data empty: {datetime.fromtimestamp(epoch, timezone.utc).isoformat()}\n{datetime.fromtimestamp(epoch_cutoff, timezone.utc).isoformat()}" )
        epoch -= 86400
        # event = keyboard.read_event(suppress=True) if keyboard.is_pressed('space') else None
        # if event and event.event_type == keyboard.KEY_DOWN:
        #     print("Key pressed! Skipping execution and continuing loop.")
        #     break

    return s_num, ttl_cnt, err_cnt, ttl_inserted, ttl_updated, ttl_empty

    ##!!!!!!!!!!!!!!!!!!!!!



#!!!!!! upload new gb data to db concurrent -- update all_api_gbs using unhcr_module\gb_serial_nums.py
global_counter = itertools.count(0)
counter_lock = threading.Lock()
def process_chunk_new(chunk, param1, param2, param3, param4):
    with counter_lock:
        chunk_count = next(param4)  # Thread-safe chunk count
    chunk_size = len(chunk)
    print(f"Processing chunk #{chunk_count} with {chunk_size} items")

    results = []
    item_counter = itertools.count(0)
    item_lock = threading.Lock()
    for item in chunk:
        with item_lock:
            item_count = next(item_counter)  # Thread-safe item count

        print(f"Processing item #{item_count} in chunk #{chunk_count}")
        msg = f'{chunk_count}:{item_count} of {chunk_size}'
        result = [upsert_gb_data(s_num=item, engine=param1, epoch_cutoff=param2, MAX_EMPTY=param3, msg=msg) for item in chunk]  # Apply function to each item
        results.append(result)
    return results

def process_chunk_start(chunk, param1, param2, param3, param4):
    with counter_lock:
        chunk_count = next(param4)  # Thread-safe chunk count

    chunk_size = len(chunk)
    print(f"Processing chunk #{chunk_count} with {chunk_size} items")
    results = []
    item_counter = itertools.count(0)
    item_lock = threading.Lock()
    for item in chunk:
        with item_lock:
            item_count = next(item_counter)  # Thread-safe item count
        print(f"Processing chunk {chunk_count} item #{item_count} ")
        msg = f'{chunk_count}:{item_count} of {chunk_size}'
        result = upsert_gb_data(s_num=item, engine=param1, epoch_start=param2, MAX_EMPTY=param3, msg=msg)
        results.append(result)
    return results

run_dt = datetime.now().date()
FILTERED_GB_SN_PATH=const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, (run_dt - timedelta(days=1)).isoformat()) #const.add_csv_dt(const.ALL_API_GBS_CSV_PATH, run_dt.isoformat())

if os.path.exists(FILTERED_GB_SN_PATH):
    filtered_gb_sn_df = pd.read_csv(FILTERED_GB_SN_PATH)
else:
    filtered_gb_sn_df = []

sn_array = sorted(filtered_gb_sn_df["gb_serial"].str.replace('-', '').tolist())

num_parts = 20
chunks = [list(chunk) for chunk in np.array_split(sn_array, num_parts)]  # Ensure list format

dt_start = now = datetime.now()
cutoff = datetime.now(timezone.utc) - timedelta(days=9)
epoch_cutoff = get_previous_midnight_epoch(int(cutoff.timestamp()))
eng = db.set_azure_defaultdb_engine()
MAX_EMPTY = 3

# counter = Value('i', 0)
# with ThreadPoolExecutor(max_workers=num_parts) as executor:
#     results = list(executor.map(partial(process_chunk_new, param1=eng, param2=epoch_cutoff, param3=MAX_EMPTY, param4=global_counter), chunks))

epoch_start = epoch_2024 + 86400
with ThreadPoolExecutor(max_workers=num_parts) as executor:
    results = list(executor.map(partial(process_chunk_start, param1=eng, param2=epoch_start, param3=MAX_EMPTY, param4=global_counter), chunks))


# Flatten results
final_output = [item for sublist in results for item in sublist]

# Compute elapsed time
elapsed = datetime.now() - dt_start

# Format as hh:mm:ss
formatted_elapsed = str(elapsed).split('.')[0] 
print('total time:', formatted_elapsed)
pass
#!!!!!! upload new gb data to db concurrent

#!!!!!!!!!!!!!!! BAD CODE ????????
complete = []
if not os.path.exists('complete.txt'):
    open('complete.txt', "w").close()
if not os.path.exists('complete_app.txt'):
    open('complete_app.txt', "w").close()

with open("complete.txt", "r") as f:
    content = f.read().strip()
    complete = list(map(str, content.split(","))) if content else []

if not complete: 
    complete = ['00980AA1','009807C4','00980892','00980AF4','00980953']

serials =  list(set(genset_gbs) - set(complete) - set(['00980864', '0098086a', '0098086c', '0098086d', '009807c4', '009807d3', '009807d8', '00980827', '00980828', '00980829', '0098082e', '0098082f', '0098082d', '00980830', '00980848', '00980849', '0098084c', '0098084f', '00980857', '0098085c', '0098085b', '0098085d', '00980aa1', '00980890', '0098087b', '00980891', '0098087c', '009808c7', '009808b5', '009808c6', '00980893', '00980892', '009808e8', '009808dc', '009808df', '009808db', '00980907', '00980914', '00980953', '009808f8', '009808f1']))

xx = 0
engine = engines[1]
for serial in sort(serials):
    xx += 1
    epoch = get_db_epoch(serial, engine, max=True)
    no_data_cnt = 0
    while epoch > 1704067199:   # Dec 31 2023
        dt = datetime.fromtimestamp(epoch, tz=timezone.utc).isoformat()
        data = meter_response_empty(serial, epoch)
        if data is None:
            no_data_cnt += 1
            logging.error(f"{dt} ZZZ {serial} ERROR: NO DATA: {no_data_cnt}")
            complete.append(serial)
            continue
        if data['Errors']:
            err = log_gb_errors(data['Errors'])
            complete.append(serial)
            with open("complete.txt", "w") as f:
                f.write(",".join(map(err, complete)))
            continue

        res, err = update_gb_db(serial, map_gb(data["DeviceData"]), engine)
        if err:
            logging.error(f"ZZZ {serial} gb_1min update_gb_db ERROR: {err}")
            complete.append(serial)
            with open("complete.txt", "w") as f:
                f.write(",".join(map(err, complete)))
            continue

        logging.info(f"gb_1min GB: {xx} {serial} Date: {dt} Rows: {res}  {epoch}")
        # previous day
        epoch -= 86400
with open("complete_app.txt", "w") as f:
    f.write(",".join(map(str, complete)))
pass

"""


# Get list of tables in schema
cur.execute("SELECT tablename FROM pg_catalog.pg_tables WHERE schemaname = 'eyedro'")
tables = [row[0] for row in cur.fetchall()]

# Build the query
query_parts = []
for table in tables:
    query_parts.append(f"(SELECT ts, epoch_secs, '{table}' AS source_table, * FROM {table})")

query = " FULL OUTER JOIN ".join(query_parts)

final_query = f"SELECT * FROM {query};"
print(final_query)  # Print the generated query

# Execute the query
cur.execute(final_query)
results = cur.fetchall()

# Print first few rows
for row in results[:5]:
    print(row)

# Close connection
cur.close()
conn.close()

    
    
"""
